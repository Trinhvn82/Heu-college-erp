import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\ctdt.xlsx')
#wb.get_sheet_names()
sheet = wb['ctdt-update']

conn = psycopg2.connect(database="ERP",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO sms_ctdtmonhoc (hocky, monhoc_id, ctdt_id) VALUES (%s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
#for r in range(sheet.max_row-1, sheet.max_row):
    hk = sheet.cell(r,1).value
    id1 = sheet.cell(r,2).value
    id2 = sheet.cell(r,3).value
    values = (hk, id1, id2)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

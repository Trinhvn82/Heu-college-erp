import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\ctdt.xlsx')
#wb.get_sheet_names()
sheet = wb['Sheet1']

conn = psycopg2.connect(database="temp8",
                        host="localhost",
                        user="postgres",
                        password="123654",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO sms_ctdt (ten, khoa, khoahoc) VALUES (%s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
    ten = sheet.cell(r,1).value
    khoa = sheet.cell(r,2).value
    khoahoc = sheet.cell(r,3).value
    

    values = (ten, khoa, khoahoc)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

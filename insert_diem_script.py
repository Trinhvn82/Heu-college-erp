import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\diemthi.xlsx')
#wb.get_sheet_names()
sheet = wb['diem']

conn = psycopg2.connect(database="SMS",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO teachers_diemthanhphan (sv_id, monhoc_id, tp_id, diem) VALUES (%s, %s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
    sv = sheet.cell(r,1).value
    monhoc = sheet.cell(r,3).value
    tp = sheet.cell(r,5).value
    diem = sheet.cell(r,7).value
    #print(ten)
    values = (sv, monhoc, tp , diem)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

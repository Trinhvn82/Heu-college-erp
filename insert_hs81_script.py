import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\hs81.xlsx')
#wb.get_sheet_names()
sheet = wb['hs81']

conn = psycopg2.connect(database="SMS",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO teachers_hs81 (sv_id, lop_id, hk, dondn, cntn, bangtn, xnct, cccd,cccdbo,cccdme,gsk,ghichu) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s)"""
print(sheet.max_row)

#for r in range(435, 463):
#276, 309
#310, 334
#335, 366
#367, 398
#399, 434
#435, 462
for r in range(2, sheet.max_row+1):
    v1 = sheet.cell(r,1).value
    v2 = sheet.cell(r,2).value
    v3 = sheet.cell(r,3).value
    v4 = sheet.cell(r,4).value
    v5 = sheet.cell(r,5).value
    v6 = sheet.cell(r,6).value
    v7 = sheet.cell(r,7).value
    v8 = sheet.cell(r,8).value
    v9 = sheet.cell(r,9).value
    v10 = sheet.cell(r,10).value
    v11 = sheet.cell(r,11).value
    v12 = sheet.cell(r,12).value
    

    values = (v1, v2, v3, v4, v5, v6, v7, v8, v9, v10, v11, v12)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\hsgv.xlsx')
#wb.get_sheet_names()
sheet = wb['Sheet1']

conn = psycopg2.connect(database="SMS",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO teachers_hsgv (ma, email, hoten, diachi, quequan, sdt, gioitinh, cccd, tthn, dantoc, loaihd, hocham, hocvi, stk, bank, branch) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
    ma = sheet.cell(r,1).value
    email = sheet.cell(r,2).value
    hoten = sheet.cell(r,3).value
    diachi = sheet.cell(r,4).value
    quequan = sheet.cell(r,5).value
    sdt = sheet.cell(r,6).value
    gioitinh = sheet.cell(r,7).value
    cccd = sheet.cell(r,8).value
    tthn = sheet.cell(r,9).value
    dantoc = sheet.cell(r,10).value
    loaihd = sheet.cell(r,11).value
    hocham = sheet.cell(r,12).value
    hocvi = sheet.cell(r,13).value
    stk = sheet.cell(r,14).value
    bank = sheet.cell(r,15).value
    branch = sheet.cell(r,16).value
    

    values = (ma, email, hoten, diachi, quequan, sdt, gioitinh, cccd, tthn, dantoc, loaihd, hocham, hocvi, stk, bank, branch)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

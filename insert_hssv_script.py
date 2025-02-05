import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\hssv.xlsx')
#wb.get_sheet_names()
sheet = wb['allsv']

conn = psycopg2.connect(database="temp1",
                        host="localhost",
                        user="postgres",
                        password="123654",
                        port="5432")
cursor = conn.cursor()



#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO sms_hssv (msv, hoten, lop, namsinh, gioitinh, dantoc, noisinh, quequan, diachi, xa, huyen, tinh, cccd, hotenbo, hotenme, sdths, sdtph, hs81_1, hs81_2, hs81_3, hs81_4, hs81_5, hs81_6, hs81_7, hs81_8, hs81_9, hs81_10, hs81_11, hs81_12, hs81_13, ghichu) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
print(sheet.max_row)

#for r in range(435, 463):
#276, 309
#310, 334
#335, 366
#367, 398
#399, 434
#435, 462
for r in range(3, sheet.max_row+1):
    v1 = sheet.cell(r,1).value
    v2 = sheet.cell(r,2).value
    v3 = sheet.cell(r,3).value
    v4 = sheet.cell(r,4).value
    v5 = sheet.cell(r,5).value
    v6 = sheet.cell(r,6).value
    v7 = sheet.cell(r,7).value
    v8 = sheet.cell(r,8).value
    v9 = sheet.cell(r,9).value
    v10 = sheet.cell(r,10).value
    v11 = sheet.cell(r,11).value
    v12 = sheet.cell(r,12).value
    v13 = sheet.cell(r,13).value
    v14 = sheet.cell(r,14).value
    v15 = sheet.cell(r,15).value
    v16 = sheet.cell(r,16).value
    v17 = sheet.cell(r,17).value
    v18 = sheet.cell(r,18).value
    v19 = sheet.cell(r,19).value
    v20 = sheet.cell(r,20).value
    v21 = sheet.cell(r,21).value
    v22 = sheet.cell(r,22).value
    v23 = sheet.cell(r,23).value
    v24 = sheet.cell(r,24).value
    v25 = sheet.cell(r,25).value
    v26 = sheet.cell(r,26).value
    v27 = sheet.cell(r,27).value
    v28 = sheet.cell(r,28).value
    v29 = sheet.cell(r,29).value
    v30 = sheet.cell(r,30).value
    v31 = sheet.cell(r,31).value
    

    values = (v1, v2, v3, v4, v5, v6, v7, v8, v9, v10, v11, v12, v13, v14, v15, v16, v17, v18, v19, v20, v21, v22, v23, v24, v25, v26, v27, v28, v29, v30,v31)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

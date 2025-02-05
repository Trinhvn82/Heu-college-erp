import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\lichhoc.xlsx')
#wb.get_sheet_names()
sheet = wb['lichhoc']

conn = psycopg2.connect(database="SMS",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO teachers_lichhoc (trungtam, diadiem, thoigian, sogio, hoc_from_gio, hoc_from_phut, hoc_to_gio, hoc_to_phut, lop_id, monhoc_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
#for r in range(sheet.max_row-1, sheet.max_row):
    d1 = sheet.cell(r,1).value
    d2 = sheet.cell(r,2).value
    d3 = sheet.cell(r,3).value
    d4 = sheet.cell(r,4).value
    d5 = sheet.cell(r,5).value
    d6 = sheet.cell(r,6).value
    d7 = sheet.cell(r,7).value
    d8 = sheet.cell(r,8).value
    d9 = sheet.cell(r,9).value
    d10 = sheet.cell(r,11).value

    values = (d1, d2, d3, d4, d5, d6, d7, d8, d9, d10)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\diemthi.xlsx')
#wb.get_sheet_names()
sheet = wb['Sheet1']

conn = psycopg2.connect(database="SMS",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO teachers_loaidiem (ten, ghichu) VALUES (%s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
    ten = sheet.cell(r,1).value
    ghichu = sheet.cell(r,2).value
    print(ten)
    values = (ten, ghichu)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

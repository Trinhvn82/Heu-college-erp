import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\lop.xlsx')
#wb.get_sheet_names()
sheet = wb['Sheet1']

conn = psycopg2.connect(database="temp1",
                        host="localhost",
                        user="postgres",
                        password="123654",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO sms_lop (ma, ten, trungtam, ctdt_id) VALUES (%s, %s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
    ma = sheet.cell(r,1).value
    ten = sheet.cell(r,2).value
    trungtam = sheet.cell(r,3).value
    ctdt_id = sheet.cell(r,4).value
    

    values = (ma, ten, trungtam, ctdt_id)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

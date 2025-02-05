import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\Đại học Điện lực\\monhoc.xlsx')
#wb.get_sheet_names()
sheet = wb['Sheet1']

conn = psycopg2.connect(database="ERP",
                        host="localhost",
                        user="postgres",
                        password="Hs123456",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO sms_monhoc (ma, ten, chuongtrinh, sotinchi, sogio_lt, sogio_th, sogio_kt) VALUES (%s, %s, %s, %s, %s, %s, %s)"""
print(sheet.max_row)

for r in range(4, sheet.max_row):
#for r in range(sheet.max_row-1, sheet.max_row):
    ma = sheet.cell(r,1).value
    ten = sheet.cell(r,2).value
    chuongtrinh = sheet.cell(r,3).value
    sotinchi = sheet.cell(r,4).value
    sogio_lt = sheet.cell(r,5).value
    sogio_th = sheet.cell(r,6).value
    sogio_kt = sheet.cell(r,7).value

    values = (ma, ten, chuongtrinh, sotinchi, sogio_lt, sogio_th, sogio_kt)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

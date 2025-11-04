import psycopg2
import openpyxl

#book = xlrd.open_workbook("D:\Đại học Điện lực\\monhoc.xlsx")
#sheet = book.sheet_by_name("Sheet1")
wb = openpyxl.load_workbook('D:\\Coding\\Python ANW\\xaphuong.xlsx')
#wb.get_sheet_names()
sheet = wb['3321 xã']

conn = psycopg2.connect(database="temp1",
                        host="localhost",
                        user="postgres",
                        password="Hs123654",
                        port="5432")
cursor = conn.cursor()

#database = psycopg2.connect (database = "RunP", user="postgres", password="", host="localhost", port="5432")

#cursor = database.cursor()

query = """INSERT INTO sms_xaphuong (ma,ten, tp) VALUES (%s, %s, %s)"""
print(sheet.max_row)

for r in range(2, sheet.max_row+1):
    ma = sheet.cell(r,1).value
    ten = sheet.cell(r,2).value
    tp = sheet.cell(r,6).value

    values = (ma, ten, tp)

    cursor.execute(query, values)

cursor.close()

conn.commit()

conn.close()

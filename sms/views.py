from datetime import datetime
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model
# --- Hàm tiện ích chuyển đổi ngày tự động cho import ---
def parse_date_auto(val):
    if pd.isnull(val) or str(val).strip() == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    # Nếu đúng định dạng YYYY-MM-DD
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        print("giá trị date 1" + s)
        pass
    # Nếu đúng định dạng DD/MM/YYYY hoặc D/M/YYYY
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        print("giá trị date 2" + s)
        pass
    # Nếu đúng định dạng D/M/YY
    try:
        return datetime.strptime(s, "%d/%m/%y").date()
    except Exception:
        print("giá trị date 3" + s)
        pass
    # Nếu đúng định dạng MM/DD/YYYY
    try:
        return datetime.strptime(s, "%m/%d/%Y").date()
    except Exception:
        pass
    # Nếu đúng định dạng khác, trả về None
    print("giá trị date 4" + s)
    return None

# ...existing code...
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Renter
from .forms_import import RenterImportForm
from django.db import IntegrityError


from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect

from django.urls import reverse
from django.utils import timezone
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_http_methods
from django.http import FileResponse
from django.core.exceptions import ValidationError
from notifications.signals import notify
from sms.utils.hashid_converter import HashidConverter

from django.template.loader import render_to_string
from decimal import Decimal
from django.contrib import messages



from django.contrib.auth import get_user_model
from django.db.models import Sum

import psycopg2
import openpyxl, xlsxwriter 
import shutil, os

from django.conf import settings
from django.http import HttpResponse, Http404, JsonResponse
import json

from django.contrib.auth.models import Group
from django.contrib.auth.models import Permission
from guardian.shortcuts import assign_perm, remove_perm


from guardian.decorators import permission_required_or_403
from django.http import HttpResponse


User = get_user_model()

from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from .forms import *

from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from datetime import datetime, date
from django.http import HttpResponseForbidden,HttpResponse
import pandas as pd
import locale

from PIL import Image
from pypdf import PdfReader
from pypdf.errors import PdfReadError


# ============================================
# Helper Functions for Ownership Validation
# ============================================

def get_object_or_forbidden(model, user, error_message, **kwargs):
    """
    Get object and validate ownership. Returns object or redirects with error message.
    
    Args:
        model: Django model class
        user: Current request user
        error_message: Message to display if access denied
        **kwargs: Query parameters (must include 'id' or 'pk')
    
    Returns:
        Tuple of (object, None) if valid, or (None, redirect_response) if invalid
    """
    try:
        obj = get_object_or_404(model, **kwargs)
        
        # Superuser has access to everything
        if user.is_superuser:
            return obj, None
        
        # Check ownership based on model type
        if model == Location:
            if obj.chu_id != user.id:
                messages.error(getattr(user, 'request', None) or user, error_message or "Bạn không có quyền truy cập vị trí này")
                return None, redirect('loc_list')
        
        elif model == House:
            if obj.loc.chu_id != user.id:
                messages.error(getattr(user, 'request', None) or user, error_message or "Bạn không có quyền truy cập nhà này")
                return None, redirect('loc_list')
        
        elif model == Renter:
            from django.db.models import Q
            # Renter must belong to user or have contract in user's house
            if not Renter.objects.filter(
                Q(chu_id=user.id) | Q(houserenter__house__loc__chu=user)
            ).filter(id=obj.id).exists():
                messages.error(getattr(user, 'request', None) or user, error_message or "Bạn không có quyền truy cập khách thuê này")
                return None, redirect('renter_list')
        
        elif model == HouseRenter:
            if obj.house.loc.chu_id != user.id:
                messages.error(getattr(user, 'request', None) or user, error_message or "Bạn không có quyền truy cập hợp đồng này")
                return None, redirect('loc_list')
        
        elif model == Hoadon:
            # Check if user is landlord or renter of this bill
            is_landlord = obj.house.loc.chu_id == user.id
            is_renter = False
            try:
                is_renter = obj.renter and obj.renter.user_id == user.id
            except (AttributeError, Renter.DoesNotExist):
                pass
            
            if not is_landlord and not is_renter:
                messages.error(getattr(user, 'request', None) or user, error_message or "Bạn không có quyền truy cập hóa đơn này")
                # Redirect based on user type
                try:
                    if user.renter:
                        return None, redirect('/renter/dashboard/')
                except (AttributeError, Renter.DoesNotExist):
                    pass
                return None, redirect('invoice_search')
        
        elif model == Thanhtoan:
            if obj.hoadon.house.loc.chu_id != user.id:
                messages.error(getattr(user, 'request', None) or user, error_message or "Bạn không có quyền truy cập thanh toán này")
                return None, redirect('invoice_search')
        
        return obj, None
        
    except Http404:
        messages.error(getattr(user, 'request', None) or user, "Không tìm thấy dữ liệu yêu cầu")
        # Return to appropriate list view
        if model == Location:
            return None, redirect('loc_list')
        elif model in [House, HouseRenter, Hoadon, Thanhtoan]:
            return None, redirect('loc_list')
        elif model == Renter:
            return None, redirect('renter_list')
        else:
            return None, redirect('index')


def block_renter_access(request, redirect_url='renter_dashboard', message="Bạn không có quyền truy cập chức năng này."):
    """Check if user is a renter and block access. Returns redirect response or None"""
    try:
        if request.user.is_renter:
            messages.error(request, message)
            return redirect(redirect_url)
    except (AttributeError, Renter.DoesNotExist):
        pass
    return None


def validate_location_access(request, loc_id):
    """Validate location access and return (location, None) or (None, error_response)"""
    return get_object_or_forbidden(
        Location, 
        request.user,
        f"Bạn không có quyền truy cập vị trí ID {loc_id}",
        id=loc_id
    )


def get_object_or_forbidden(model, request, error_message, **kwargs):
    """
    Get object and validate ownership. Returns object or redirects with error message.
    
    Args:
        model: Django model class
        request: Current HttpRequest object
        error_message: Message to display if access denied
        **kwargs: Query parameters (must include 'id' or 'pk')
    
    Returns:
        Tuple of (object, None) if valid, or (None, redirect_response) if invalid
    """
    try:
        obj = get_object_or_404(model, **kwargs)
        user = request.user
        
        # Superuser has access to everything
        if user.is_superuser:
            return obj, None
        
        # Check ownership based on model type
        if model == Location:
            if obj.chu_id != user.id:
                messages.error(request, error_message or "Bạn không có quyền truy cập vị trí này")
                return None, redirect('loc_list')
        
        elif model == House:
            if obj.loc.chu_id != user.id:
                messages.error(request, error_message or "Bạn không có quyền truy cập nhà này")
                return None, redirect('loc_list')
        
        elif model == Renter:
            from django.db.models import Q
            # Renter must belong to user or have contract in user's house
            if not Renter.objects.filter(
                Q(chu_id=user.id) | Q(houserenter__house__loc__chu=user)
            ).filter(id=obj.id).exists():
                messages.error(request, error_message or "Bạn không có quyền truy cập khách thuê này")
                return None, redirect('renter_list')
        
        elif model == HouseRenter:
            if obj.house.loc.chu_id != user.id:
                messages.error(request, error_message or "Bạn không có quyền truy cập hợp đồng này")
                return None, redirect('loc_list')
        
        elif model == Hoadon:
            # Check if user is landlord or renter of this bill
            is_landlord = obj.house.loc.chu_id == user.id
            is_renter = False
            try:
                is_renter = obj.renter and obj.renter.user_id == user.id
            except (AttributeError, Renter.DoesNotExist):
                pass
            
            if not is_landlord and not is_renter:
                messages.error(request, error_message or "Bạn không có quyền truy cập hóa đơn này")
                # Redirect based on user type
                try:
                    if user.renter:
                        return None, redirect('/renter/dashboard/')
                except (AttributeError, Renter.DoesNotExist):
                    pass
                return None, redirect('invoice_search')
        
        elif model == Thanhtoan:
            if obj.hoadon.house.loc.chu_id != user.id:
                messages.error(request, error_message or "Bạn không có quyền truy cập thanh toán này")
                return None, redirect('invoice_search')
        
        return obj, None
        
    except Http404:
        messages.error(request, "Không tìm thấy dữ liệu yêu cầu")
        # Return to appropriate list view
        if model == Location:
            return None, redirect('loc_list')
        elif model in [House, HouseRenter, Hoadon, Thanhtoan]:
            return None, redirect('loc_list')
        elif model == Renter:
            return None, redirect('renter_list')
        else:
            return None, redirect('index')


def validate_location_access(request, loc_id):
    """Validate location access and return (location, None) or (None, error_response)"""
    return get_object_or_forbidden(
        Location, 
        request,
        f"Bạn không có quyền truy cập vị trí ID {loc_id}",
        id=loc_id
    )


def validate_house_access(request, house_id):
    """Validate house access and return (house, None) or (None, error_response)"""
    return get_object_or_forbidden(
        House,
        request,
        f"Bạn không có quyền truy cập nhà ID {house_id}",
        id=house_id
    )


def validate_renter_access(request, renter_id):
    """Validate renter access and return (renter, None) or (None, error_response)"""
    return get_object_or_forbidden(
        Renter,
        request,
        f"Bạn không có quyền truy cập khách thuê ID {renter_id}",
        id=renter_id
    )


def validate_contract_access(request, hr_id):
    """Validate house-renter contract access and return (contract, None) or (None, error_response)"""
    return get_object_or_forbidden(
        HouseRenter,
        request,
        f"Bạn không có quyền truy cập hợp đồng ID {hr_id}",
        id=hr_id
    )


def validate_bill_access(request, bill_id):
    """Validate bill access and return (bill, None) or (None, error_response)"""
    return get_object_or_forbidden(
        Hoadon,
        request,
        f"Bạn không có quyền truy cập hóa đơn ID {bill_id}",
        id=bill_id
    )


def validate_payment_access(request, payment_id):
    """Validate payment access and return (payment, None) or (None, error_response)"""
    return get_object_or_forbidden(
        Thanhtoan,
        request,
        f"Bạn không có quyền truy cập thanh toán ID {payment_id}",
        id=payment_id
    )


#@login_required
def index(request):
    """Main entry point - redirects to welcome page after login"""
    if not request.user.is_authenticated:
        return redirect("renter_landing")
    
    # Check if user is a renter - use try/except to avoid AttributeError
    try:
        if request.user.is_renter:
            return redirect("/renter/dashboard/")
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if request.user.is_superuser:
        return redirect("shop-statistics")
    elif request.user.is_hv:
        return redirect("lichhoc_list")
    elif request.user.is_gv:
        return redirect("lichhoc_list")
    elif request.user.is_internalstaff:
        return redirect("lop_list")
    else:
        # Nếu là chủ nhà (có Location), luôn vào home_dashboard
        if request.user.is_landlord:
            return redirect("home_dashboard")


@login_required
def welcome_page(request):
    """Welcome page shown after successful login"""
    # Redirect renters to their dashboard
    try:
        if request.user.is_renter:
            return redirect("renter_dashboard")
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Get quick stats for landlords
    context = {
        'total_locations': Location.objects.filter(chu=request.user).count(),
        'total_houses': House.objects.filter(loc__chu=request.user).count(),
        'total_renters': Renter.objects.filter(chu_id=request.user.id).count(),
    }
    
    return render(request, 'sms/welcome.html', context)


@login_required
def home_dashboard(request):
    """Main dashboard with feature blocks"""
    # Block renter access
    try:
        if request.user.is_renter:
            messages.warning(request, "Bạn không có quyền truy cập trang này.")
            return redirect("renter_dashboard")
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    from django.db.models import Q, Count
    from datetime import datetime, timedelta
    
    # Calculate statistics
    locations = Location.objects.filter(chu=request.user)
    houses = House.objects.filter(loc__chu=request.user)
    renters = Renter.objects.filter(chu_id=request.user.id)
    
    # Active contracts (rent_to is None or in the future)
    active_contracts = HouseRenter.objects.filter(
        house__loc__chu=request.user
    ).filter(
        Q(rent_to__isnull=True) | Q(rent_to__gte=timezone.now().date())
    ).count()
    
    # Occupied houses
    occupied_houses = houses.filter(
        houserenter__rent_to__isnull=True
    ).distinct().count()
    
    # Pending bills (status = ChuaTT or QuaHan)
    pending_bills = Hoadon.objects.filter(
        house__loc__chu=request.user,
        status__in=['ChuaTT', 'QuaHan']
    ).count()
    
    # Pending issues
    pending_issues = IssueReport.objects.filter(
        house__loc__chu=request.user,
        status__in=['new', 'in_progress']
    ).count()
    
    # Unread notifications
    try:
        unread_notifications = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
    except:
        unread_notifications = 0
    
    # Recent activities (last 10)
    recent_activities = []
    
    # Recent bills
    recent_bills = Hoadon.objects.filter(
        house__loc__chu=request.user
    ).order_by('-ngay_tao')[:3]
    
    for bill in recent_bills:
        recent_activities.append({
            'icon': 'fa-file-invoice',
            'color': 'warning',
            'title': f'Hóa đơn {bill.house.ten}',
            'description': f'{bill.ten} - {bill.get_status_display()}',
            'time': bill.ngay_tao
        })
    
    # Recent issues
    recent_issues = IssueReport.objects.filter(
        house__loc__chu=request.user
    ).order_by('-created_at')[:3]
    
    for issue in recent_issues:
        recent_activities.append({
            'icon': 'fa-tools',
            'color': 'danger',
            'title': f'Sự cố: {issue.title}',
            'description': f'{issue.house.ten} - {issue.get_status_display()}',
            'time': issue.created_at
        })
    
    # Sort by time
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    recent_activities = recent_activities[:10]
    
    context = {
        'current_date': timezone.now(),
        'stats': {
            'total_locations': locations.count(),
            'total_houses': houses.count(),
            'occupied_houses': occupied_houses,
            'total_renters': renters.count(),
            'active_contracts': active_contracts,
            'pending_bills': pending_bills,
            'pending_issues': pending_issues,
            'unread_notifications': unread_notifications,
        },
        'recent_activities': recent_activities,
    }
    
    return render(request, 'sms/home_dashboard.html', context)

# Teacher Views

@login_required
def t_clas(request, teacher_id, choice):
    teacher1 = get_object_or_404(Hsgv, id=teacher_id)
    return render(request, 'info/t_clas.html', {'teacher1': teacher1, 'choice': choice})


@login_required()
def t_timetable(request, teacher_id):
    #asst = AssignTime.objects.filter(assign__teacher_id=teacher_id)
    class_matrix = [[True for i in range(12)] for j in range(6)]

    '''
    for i, d in enumerate(DAYS_OF_WEEK):
        t = 0
        for j in range(12):
            if j == 0:
                class_matrix[i][0] = d[0]
                continue
            if j == 4 or j == 8:
                continue
            try:
                a = asst.get(period=time_slots[t][0], day=d[0])
                class_matrix[i][j] = a
            except AssignTime.DoesNotExist:
                pass
            t += 1
    '''
    context = {
        'class_matrix': class_matrix,
    }
    return render(request, 'sms/t_timetable.html', context)


# Create your views here.
def teacher_list(request):
    teachers = TeacherInfo.objects.all()

    paginator = Paginator(teachers, 10)
    page = request.GET.get('page')
    paged_teachers = paginator.get_page(page)
    context = {
        "teachers": paged_teachers
    }
    return render(request, "sms/teacher_list.html", context)

@login_required
@permission_required('sms.view_hsgv',raise_exception=True)
def gv_list(request):
    if request.user.is_gv:
        teachers = Hsgv.objects.filter(user = request.user).order_by('hoten')
    else:
        teachers = Hsgv.objects.all().order_by('hoten')
    if request.method == "POST":
            query_name = request.POST.get('ten', None)
            if query_name:
                teachers = Hsgv.objects.filter(hoten__contains=query_name).order_by('hoten')
                messages.success(request, "Ket qua tim kiem voi ten co chua: " + query_name)
#                return render(request, 'product-search.html', {"results":results})

    paginator = Paginator(teachers, 20)
    page = request.GET.get('page')
    paged_teachers = paginator.get_page(page)
    context = {
        "teachers": paged_teachers
    }
    return render(request, "sms/gv_list.html", context)

@login_required
def ns_list(request):
    ns = Hsns.objects.all().order_by('hoten')
    if request.method == "POST":
            query_name = request.POST.get('ten', None)
            if query_name:
                ns = Hsns.objects.filter(hoten__contains=query_name).order_by('hoten')
                messages.success(request, "Ket qua tim kiem voi ten co chua: " + query_name)
#                return render(request, 'product-search.html', {"results":results})

    paginator = Paginator(ns, 20)
    page = request.GET.get('page')
    paged_ns = paginator.get_page(page)
    context = {
        "nss": paged_ns
    }
    return render(request, "sms/ns_list.html", context)

@login_required
def renter_list(request):
    from django.db.models import Q
    
    # Chặn renter truy cập danh sách khách thuê
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền xem danh sách khách thuê.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if request.user.is_superuser:
        renters = Renter.objects.all().order_by('hoten')
    else:
        # Chỉ hiển thị renter do chủ nhà này tạo ra
        renters = Renter.objects.filter(chu_id=request.user.id).order_by('hoten')
    
    if request.method == "POST":
            query_name = request.POST.get('name', None)
            if query_name:
                # Lọc search theo chủ nhà
                if request.user.is_superuser:
                    renters = Renter.objects.filter(hoten__contains=query_name).order_by('hoten')
                else:
                    renters = Renter.objects.filter(chu_id=request.user.id, hoten__contains=query_name).order_by('hoten')
                messages.success(request, "Ket qua tim kiem voi ten co chua: " + query_name)
#                return render(request, 'product-search.html', {"results":results})

    paginator = Paginator(renters, 20)
    page = request.GET.get('page')
    paged_renters = paginator.get_page(page)
    context = {
        "renters": paged_renters
    }
    return render(request, "sms/renter_list.html", context)

@login_required
@permission_required('sms.view_hssv',raise_exception=True)
def sv_list(request):
    from django.db.models import Q
    if request.user.is_hv:
        students = Hssv.objects.filter(user = request.user).order_by('msv')
    else:
        students = Hssv.objects.all().order_by('msv')
    if request.method == "POST":
            query_name = request.POST.get('ten', None)
            if query_name:
                query_name = query_name.strip()
                #students = students.filter(hoten__contains=query_name)
                students = students.filter(Q(msv__icontains=query_name) | Q(hoten__icontains=query_name) | Q(lop__ten__icontains=query_name))
                messages.success(request, "Ket qua tim kiem voi: " + query_name)
#                return render(request, 'product-search.html', {"results":results})

    paginator = Paginator(students, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "students": paged_students
    }
    return render(request, "sms/sv_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def sv_lop(request, lop_id):
    #students = Hssv.objects.all()
    tenlop = Lop.objects.get(id = lop_id).ten
    #students = Hssv.objects.filter(lop_id = lop_id).order_by('ten')
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    paginator = Paginator(ans, 50)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    print("LOP ID la ")
    print(lop_id)

    context = {
        "students": paged_students,
        "lop_id": lop_id,
        "tenlop": tenlop
    }
    return render(request, "sms/svlop_list.html", context)

@login_required
def lichhoc_lop(request, lop_id):
    #students = Hssv.objects.all()
    tenlop = Lop.objects.get(id = lop_id).ten
    lichhoc = Lichhoc.objects.filter(lop_id = lop_id)
    paginator = Paginator(lichhoc, 100)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lh": paged_students,
        "tenlop": tenlop
    }
    return render(request, "sms/lichhoc-lop_list.html", context)


@login_required
def lichhoc_lopmh(request, lopmh_id):
    #students = Hssv.objects.all()
    lmh = LopMonhoc.objects.filter(id = lopmh_id).select_related("monhoc", "lop")[0]
    lop_id = LopMonhoc.objects.get(id = lopmh_id).lop_id
    monhoc_id = LopMonhoc.objects.get(id = lopmh_id).monhoc_id
    lichhoc = Lichhoc.objects.filter(lmh_id = lmh.id).order_by('thoigian')
    paginator = Paginator(lichhoc, 100)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)

    #lmh = LopMonhoc.objects.get(id = lopmh_id).select_related("monhoc", "lop")

    context = {
        "lh": paged_students,
        "lmh": lmh,
    }
    return render(request, "sms/lichhoc-lopmh_list.html", context)
@login_required
def hocphi_lop(request, lop_id):
    #students = Hssv.objects.all()
    tenlop = Lop.objects.get(id = lop_id).ten
    hocphi = Hocphi.objects.all().select_related("sv").filter(lop_id = lop_id)
    paginator = Paginator(hocphi, 100)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lh": paged_students,
        "tenlop": tenlop,
        "lop_id": lop_id
    }
    return render(request, "sms/hocphi-lop_list.html", context)
@login_required
def diem_lop(request, lop_id):

    #students = Hssv.objects.all()
    tenlop = Lop.objects.get(id = lop_id).ten
    diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__lop__contains=tenlop)
    if request.method == "POST":
            query_name = request.POST.get('name', None)
            if query_name:
                diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__lop__contains=tenlop, sv__hoten__contains=query_name)
                #students = Hssv.objects.filter(hoten__contains=query_name)
                messages.success(request, "Ket qua tim kiem voi ten co chua: " + query_name)
#                return render(request, 'product-search.html', {"results":results})
    #lichhoc = Diemthanhphan.objects.filter(lop_id = lop_id)
    paginator = Paginator(diems, 40)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "diems": paged_students,
        "tenlop": tenlop,
        "lop_id": lop_id
    }
    return render(request, "sms/diem-lop_list.html", context)

@login_required
def diem_lmh(request, lmh_id):

    #students = Hssv.objects.all()
    lop_id = LopMonhoc.objects.get(id = lmh_id).lop_id
    mh_id = LopMonhoc.objects.get(id = lmh_id).monhoc_id

    tenlop = Lop.objects.get(id = lop_id).ma
    tenmh = Monhoc.objects.get(id = mh_id).ten

    stud_list = Hssv.objects.filter(lop_id = lop_id)
    lds= Loaidiem.objects.all()
    diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, monhoc_id =mh_id).order_by('tp_id', 'sv_id')
    if request.method == "POST":
        for stud in stud_list:
            for ld in lds:
                id = "C"+str(stud.id)+"-"+str(ld.id)
                diem = request.POST[id]
                mark = Diemthanhphan.objects.get(sv_id = stud.id, tp_id = ld.id, monhoc_id=mh_id)
                mark.diem = diem
                mark.status = 1
                mark.save()
        messages.success(request, "Cap nhat diem thanh cong!")
        return redirect("lop_monhoc", lop_id)


    if not diems.first():
        #create mark record
        for stud in stud_list:
            for ld in lds:
                mark = Diemthanhphan(diem =0, sv_id = stud.id, tp_id = ld.id, monhoc_id=mh_id)
                mark.save()
        diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, monhoc_id =mh_id)

    context = {
        "diems": diems,
        "tenlop": tenlop,
        "tenmh": tenmh
    }
    return render(request, "sms/diem-lmh.html", context)

@login_required
def diem_lmh_lst(request, lmh_id):

    #students = Hssv.objects.all()
    lop_id = LopMonhoc.objects.get(id = lmh_id).lop_id
    mh_id = LopMonhoc.objects.get(id = lmh_id).monhoc_id

    tenlop = Lop.objects.get(id = lop_id).ma
    tenmh = Monhoc.objects.get(id = mh_id).ten
    lds= Loaidiem.objects.all()
    lol=[]
    diems = Diemthanhphan.objects.all()
    stud_list = Hssv.objects.filter(lop_id = lop_id)
    lds= Loaidiem.objects.all()
    for ld in lds:
        st = Diemthanhphan.objects.filter(sv__in=stud_list, monhoc_id =mh_id, tp_id=ld.ma, status = 1).first()
        if st:
            lol.append({ "ma":ld.ma,"ten": ld.ten, "st": 1})
        else:
            lol.append({ "ma":ld.ma,"ten": ld.ten, "st": 0})
    context = {
        "tenlop": tenlop,
        "lop_id": lop_id,
        "lds": lds,
        "lmh_id":lmh_id,
        "lol":lol,
        "tenmh": tenmh
    }
    return render(request, "sms/diem-lmh-lst.html", context)

@login_required
def diemtp_lmh_lst(request, lmh_id, opt):

    #students = Hssv.objects.all()
    lop_id = LopMonhoc.objects.get(id = lmh_id).lop_id
    mh_id = LopMonhoc.objects.get(id = lmh_id).monhoc_id

    tenlop = Lop.objects.get(id = lop_id).ma
    tenmh = Monhoc.objects.get(id = mh_id).ten

    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("lop","monhoc", "hk")[0]
    locale.setlocale(locale.LC_ALL, 'vi_VN')

    hls = Hoclai.objects.filter(lmh_id = lmh_id)
    svs = sorted(Hssv.objects.filter(lop_id = lmh.lop.id) | Hssv.objects.filter(id__in = hls.values_list('sv_id', flat=True)), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)

    lds= Loaidiem.objects.all().order_by('id')
    for ld in lds:
        log = LogDiem.objects.filter(id__in=Diemthanhphan.objects.filter(sv__in=svs, monhoc_id =mh_id, tp_id=ld.id, status = 1).values_list('log_id', flat=True)).order_by('id')
        ld.log= log

    #Details section


    if opt:
        for sv in svs:
            tbmhk, tchk = 0,0
            ldl=[]
            tbm1_diem, tbm1_heso, tbm2_diem, tbm2_heso, tbm= 0,0,0,0,0
            kttx1,kttx2,kttx3,ktdk1,ktdk2,ktdk3,ktkt1,ktkt2 = 0,0,0,0,0,0,0,0
            n_kttx1,n_kttx2,n_kttx3,n_ktdk1,n_ktdk2,n_ktdk3,n_ktkt1,n_ktkt2 = 0,0,0,0,0,0,0,0
            for ld in lds:
                dtpl=[]
                dtps = Diemthanhphan.objects.filter(sv = sv, lmh_id = lmh_id, tp_id = ld.id, status=1).order_by('log_id')

                i=1
                for dtp in dtps:
                    if i==1 and ld.ma == 'KTTX':
                        kttx1 = dtp.diem
                        n_kttx1 = 1
                    elif i==2 and ld.ma == 'KTTX':
                        kttx2 = dtp.diem
                        n_kttx2 = 1
                    elif i==3 and ld.ma == 'KTTX':
                        kttx3 = dtp.diem
                        n_kttx3 = 1
                    elif i==1 and ld.ma == 'KTĐK':
                        ktdk1 = dtp.diem
                        n_ktdk1 = 1
                    elif i==2 and ld.ma == 'KTĐK':
                        ktdk2 = dtp.diem
                        n_ktdk2 = 1
                    elif i==3 and ld.ma == 'KTĐK':
                        ktdk3 = dtp.diem
                        n_ktdk3 = 1
                    elif i==1 and ld.ma == 'KTKT' and dtp.att ==1:
                        ktkt1 = dtp.diem
                        n_ktkt1 = 1
                    elif i==2 and ld.ma == 'KTKT' and dtp.att ==1:
                        ktkt2 = dtp.diem
                        n_ktkt2 = 1
                    i=i+1
                    dtpl.append({"id":dtp.log_id,"mark":dtp.diem})
                    if ld.ma == 'KTĐK' or ld.ma == 'KTTX':
                        tbm1_diem = tbm1_diem + dtp.diem * ld.heso
                        tbm1_heso = tbm1_heso + ld.heso
                    elif ld.ma == 'KTKT' and dtp.att ==1:
                        tbm2_diem = dtp.diem * ld.heso
                        # tbm2_heso = ld.heso
                        # print(tbm2_diem)
                        # print(tbm2_heso)
                if ld.ma == 'KTKT':
                    tbm2_heso = ld.heso
                
            tbmkt = round((tbm1_diem/tbm1_heso),1) if tbm1_heso else 0
            tbm = round(((tbm1_diem/tbm1_heso)*(10-tbm2_heso) + tbm2_diem)/10,1) if tbm1_heso else (round((tbm2_diem/tbm2_heso),1) if tbm2_heso else 0)
            if tbm >=8.5 and tbm <=10:
                tbm4 = 4
                tbmc = "A"
            elif tbm >=7 and tbm <=8.4:
                tbm4 = 3
                tbmc = "B"
            elif tbm >=5.5 and tbm <=6.9:
                tbm4 = 2
                tbmc = "C"
            elif tbm >=4 and tbm <=5.4:
                tbm4 = 1
                tbmc = "D"
            elif tbm  < 4:
                tbm4 = 0
                tbmc = "F"

            if DiemTk.objects.filter(sv_id =sv.id, lmh_id = lmh_id).exists():
                dtk = DiemTk.objects.get(sv_id =sv.id, lmh_id = lmh_id)
            else:
                dtk = DiemTk(sv_id =sv.id, lmh_id = lmh_id)
            dtk.ma = sv.msv
            dtk.ten = sv.hoten
            dtk.hk_id = lmh.hk_id
            dtk.monhoc_id = lmh.monhoc_id
            dtk.monhoc = lmh.monhoc.ten
            dtk.mhdk = lmh.mhdk
            dtk.tc = lmh.monhoc.sotinchi
            dtk.tbm = tbm
            dtk.tbmc = tbmc
            dtk.tbm4 = tbm4
            dtk.tbmkt = tbmkt
            dtk.kttx1 = kttx1
            dtk.n_kttx1 = n_kttx1
            dtk.kttx2 = kttx2
            dtk.n_kttx2 = n_kttx2
            dtk.kttx3 = kttx3
            dtk.n_kttx3 = n_kttx3
            dtk.ktdk1 = ktdk1
            dtk.n_ktdk1 = n_ktdk1
            dtk.ktdk2 = ktdk2
            dtk.n_ktdk2 = n_ktdk2
            dtk.ktdk3 = ktdk3
            dtk.n_ktdk3 = n_ktdk3
            dtk.ktkt1 = ktkt1
            dtk.n_ktkt1 = n_ktkt1
            dtk.ktkt2 = ktkt2
            dtk.n_ktkt2 = n_ktkt2
            dtk.save()

    dtks = DiemTk.objects.filter(lmh_id = lmh_id)
    context = {
        "tenlop": tenlop,
        "lop_id": lop_id,
        "lds": lds,
        "lmh_id":lmh_id,
        "tenmh": tenmh,
        "lml": dtks,
        "lmh": lmh
    }
    return render(request, "sms/diemtp-lmh-lst.html", context)

@login_required
def gv_lmh_lst(request, lmh_id):

    #lmh = LopMonhoc.objects.get(id = lmh_id)
    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("monhoc", "lop")[0]

    lh = Lichhoc.objects.filter(lmh_id = lmh_id).select_related("lop", "monhoc")
    gvs = Hsgv.objects.filter(id__in = lh.values_list('giaovien_id', flat=True))
    
    for gv in gvs:
        gv.lhs = Lichhoc.objects.filter(lmh_id = lmh_id, giaovien_id = gv.id)
        gv.sotiet = gv.lhs.aggregate(Sum('sotiet'))['sotiet__sum']
        if Ttgv.objects.filter(gv_id = gv.id, lopmh_id = lmh_id).exists():
            gv.sotien = Ttgv.objects.get(gv_id = gv.id, lopmh_id = lmh_id).sotien2
        else:
            gv.sotien = 0
        #Sale.objects.filter(type='Flour').aggregate(Sum('column'))['column__sum']
        
    context = {
        "lmh": lmh,
        "gvs": gvs,
    }
    return render(request, "sms/gv-lmh-lst.html", context)

@login_required
def details_gv(request, gv_id):

    #lmh = LopMonhoc.objects.get(id = lmh_id)
    lh = Lichhoc.objects.filter(giaovien_id = gv_id).select_related("lmh")
    lmh = LopMonhoc.objects.filter(id__in = lh.values_list('lmh_id', flat=True))

    lops = Lop.objects.filter(id__in = lmh.values_list('lop_id', flat=True))
    #lmh = LopMonhoc.objects.filter(id__in = lh.values_list('monhoc_id', flat=True))
    gv = Hsgv.objects.get(id = gv_id)

    #gvs = Hsgv.objects.filter(id__in = lh.values_list('giaovien_id', flat=True))
    
    for l in lops:
        mhs = LopMonhoc.objects.filter(monhoc_id__in = lmh.values_list('monhoc_id', flat=True), lop_id = l.id).select_related("monhoc")
        for mh in mhs:
            mh.lhs = Lichhoc.objects.filter(giaovien_id = gv_id, lmh_id = mh.id)
            mh.sotiet = mh.lhs.aggregate(Sum('sotiet'))['sotiet__sum']
        l.mhs = mhs

        
    context = {
        "lops": lops,
        "gv": gv,
    }
    return render(request, "sms/gv_details.html", context)

@login_required
def lop81_lst(request, lop_id):

    tenlop = Lop.objects.get(id = lop_id).ma
    hks= Hocky.objects.all()

    lol=[]
    diems = Hs81.objects.all()
    stud_list = Hssv.objects.filter(lop_id = lop_id)
    hks= Hocky.objects.all()
    for hk in hks:
        ft1 = Hs81.objects.filter(lop_id=lop_id, hk= hk.ma,status = 1).first()
        ft2 = Hocphi.objects.filter(lop_id=lop_id, hk= hk.ma,status = 1).first()
        if ft1:
            st1=1
        else:
            st1=0
        if ft2:
            st2=1
        else:
            st2=0
        lol.append({ "ma":hk.ma,"ten": hk.ten, "st1": st1, "st2": st2})

    context = {
        "tenlop": tenlop,
        "lol": lol,
        "lop_id":lop_id
    }
    return render(request, "sms/lop81-lst.html", context)

@login_required
def diem_lmh_dtp(request, lmh_id, dtp_id):

    #students = Hssv.objects.all()
    lop_id = LopMonhoc.objects.get(id = lmh_id).lop_id
    mh_id = LopMonhoc.objects.get(id = lmh_id).monhoc_id

    tenlop = Lop.objects.get(id = lop_id).ma
    tenmh = Monhoc.objects.get(id = mh_id).ten

    stud_list = Hssv.objects.filter(lop_id = lop_id)
    lds= Loaidiem.objects.filter(ma=dtp_id)
    diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, monhoc_id =mh_id, tp_id=dtp_id).order_by('tp_id', 'sv_id')
    if request.method == "POST":
        for stud in stud_list:
            for ld in lds:
                id = "C"+str(stud.id)+"-"+str(ld.ma)
                diem = request.POST[id]
                mark = Diemthanhphan.objects.get(sv_id = stud.id, tp_id = ld.ma, monhoc_id=mh_id)
                mark.diem = diem
                mark.status = 1
                mark.save()
        messages.success(request, "Cap nhat diem thanh cong!")
        return redirect("diemtp-lmh-lst", lmh_id,1)


    if not diems.first():
        #create mark record
        log = LogDiem()
        log.save()
        for stud in stud_list:
            for ld in lds:
                mark = Diemthanhphan(diem =0, sv_id = stud.id, tp_id = ld.ma, monhoc_id=mh_id, log=log) 
                #mark.log = log
                mark.save()
        diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, monhoc_id =mh_id, tp_id=dtp_id).order_by('tp_id', 'sv_id')

    context = {
        "diems": diems,
        "tenlop": tenlop,
        "lop_id": lop_id,
        "lds": lds,
        "lmh_id":lmh_id,
        "tenmh": tenmh
    }
    return render(request, "sms/diem-lmh.html", context)

@login_required
def create_diemtp(request, lop_id, lmh_id, dtp_id):

    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("lop", "monhoc")[0]
    mald = Loaidiem.objects.get(id = dtp_id).ma

    locale.setlocale(locale.LC_ALL, 'vi_VN')

    hls = Hoclai.objects.filter(lmh_id = lmh_id)
    stud_list = sorted(Hssv.objects.filter(lop_id = lmh.lop.id) | Hssv.objects.filter(id__in = hls.values_list('sv_id', flat=True)), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)

    if request.method == "POST":
        for stud in stud_list:
            try:
                id = "C"+str(stud.id)+"-"+str(dtp_id)

                id_att = "Att"+str(stud.id)
                att = request.POST.get(id_att, None)

                diem = request.POST[id]
                log_id = request.POST["log"]
                mark = Diemthanhphan.objects.get(sv_id = stud.id, tp_id = dtp_id, lmh_id=lmh_id,monhoc_id = lmh.monhoc_id,log_id=log_id)
                mark.diem = diem
                mark.att = 1 if mald != 'KTKT' else 1 if att else 0
                mark.status = 1
                mark.save()
                #send notification to Hv
                if stud.user:
                    notify.send(sender=stud.user, recipient= stud.user, verb='Thông tin điểm được cập nhật trên hệ thống', level='info')

            except Exception as e:
                messages.error(request, 'Nhập điểm cho mã: ' + stud.msv  + ' có lỗi:' + str(e))

        messages.success(request, "Cap nhat diem thanh cong!")
        return redirect("diemtp-lmh-lst", lmh_id,1)


    #create mark record
    log = LogDiem(ten = request.user.username)  
    log.save()
    for stud in stud_list:
        mark = Diemthanhphan(sv_id = stud.id, tp_id = dtp_id, lmh_id=lmh_id, monhoc_id = lmh.monhoc_id,log=log) 
        print(stud.hoten)
        mark.save()

    diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, lmh_id=lmh_id, tp_id=dtp_id, log=log).order_by('id')

    context = {
        "diems": diems,
        "mald": mald,
        "log": log,
        "lmh":lmh
    }
    return render(request, "sms/diem-lmh.html", context)

@login_required
def edit_diemtp(request, lop_id, lmh_id, dtp_id, log_id):

    #students = Hssv.objects.all()
    #lop_id = LopMonhoc.objects.get(id = lmh_id).lop_id
    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("lop", "monhoc")[0]
    mald = Loaidiem.objects.get(id = dtp_id).ma

    hls = Hoclai.objects.filter(lmh_id = lmh_id)

    stud_list = Hssv.objects.filter(lop_id = lmh.lop.id) | Hssv.objects.filter(id__in = hls.values_list('sv_id', flat=True))
    #lds= Loaidiem.objects.filter(ma=dtp_id)
    #diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, monhoc_id =mh_id, tp_id=dtp_id).order_by('tp_id', 'sv_id')
    log = LogDiem.objects.get(id = log_id)
    if request.method == "POST":
        log.capnhat_at = datetime.now()
        log.ten = request.user.username
        log.save()
        for stud in stud_list:
            id = "C"+str(stud.id)+"-"+str(dtp_id)

            if request.POST.get(id, False):
                id_att = "Att"+str(stud.id)
                att = request.POST.get(id_att, None)

                diem = request.POST.get(id, False)
                mark = Diemthanhphan.objects.get(sv_id = stud.id, tp_id = dtp_id, lmh_id = lmh_id, log_id=log_id)
                mark.diem = diem
                mark.status = 1
                mark.att = 1 if mald != 'KTKT' else 1 if att else 0
                mark.save()
                #send notification to Hv
                if stud.user:
                    notify.send(sender=stud.user, recipient= stud.user, verb='Thông tin điểm được cập nhật trên hệ thống', level='info')

        messages.success(request, "Cap nhat diem thanh cong!")
        return redirect("diemtp-lmh-lst", lmh_id, 1)

    diems = Diemthanhphan.objects.all().select_related("sv", "tp", "monhoc").filter(sv__in=stud_list, lmh_id = lmh_id, tp_id=dtp_id, log=log).order_by('id')

    for diem in diems:
        print(diem.sv.hoten)
        print(diem.diem) 

    context = {
        "diems": diems,
        "mald": mald,
        "log": log,
        "lmh":lmh
    }
    return render(request, "sms/diem-lmh.html", context)

@login_required
def delete_diemtp(request, lmh_id, dtp_id, log_id):

    lmh = LopMonhoc.objects.get(id = lmh_id)
    hls = Hoclai.objects.filter(lmh_id = lmh_id)

    stud_list = Hssv.objects.filter(lop_id = lmh.lop.id) | Hssv.objects.filter(id__in = hls.values_list('sv_id', flat=True))
    try:
        Diemthanhphan.objects.filter(sv__in=stud_list
                                    , monhoc_id =lmh.monhoc.id
                                    , lmh_id = lmh.id
                                    , tp_id=dtp_id
                                    , log_id=log_id).delete()
        
        messages.success(request, "Xóa điểm thành công!")
        return redirect("diemtp-lmh-lst", lmh_id, 1)
    except Exception as e:
        messages.error(request, 'Lỗi khi xóa điểm: ' + str(e))
        return redirect("diemtp-lmh-lst", lmh_id,1)

@login_required
def lop81_hk(request, lop_id, hk_ma):

    tenlop = Lop.objects.get(id = lop_id).ma
    hks= Hocky.objects.filter(ma=hk_ma)
    stud_list = Hssv.objects.filter(lop_id = lop_id)
    hss = Hs81.objects.all().select_related("sv").filter(sv__in=stud_list, hk=hk_ma).order_by('sv_id')
    if request.method == "POST":
        for stud in stud_list:
            for hk in hks:
                hs = Hs81.objects.get(sv_id = stud.id,hk=hk.ma)
                hs.dondn = request.POST["dondn"+str(stud.id)]
                hs.cntn = request.POST["cntn"+str(stud.id)]
                hs.bangtn = request.POST["bangtn"+str(stud.id)]
                hs.xnct = request.POST["xnct"+str(stud.id)]
                hs.cccd = request.POST["cccd"+str(stud.id)]
                hs.cccdbo = request.POST["cccdbo"+str(stud.id)]
                hs.cccdme = request.POST["cccdme"+str(stud.id)]
                hs.gks = request.POST["gks"+str(stud.id)]
                hs.ghichu = request.POST["ghichu"+str(stud.id)]
                hs.status = 1

                hs.save()
        messages.success(request, "Cap nhat ho so 81 thanh cong!")
        return redirect("lop81-lst", lop_id)


    if not hss.first():
        #create hs81 record
        for stud in stud_list:
            for hk in hks:
                hs = Hs81(sv_id = stud.id, lop_id = lop_id, hk=hk.ma)
                hs.save()
        hss = Hs81.objects.all().select_related("sv").filter(sv__in=stud_list, hk=hk_ma).order_by('sv_id')

    context = {
        "hss": hss,
        "tenlop": tenlop,
        "hk_ma": hk_ma
    }
    return render(request, "sms/lop81-hk.html", context)

@login_required
def lophp_hk(request, lop_id, hk_ma):

    tenlop = Lop.objects.get(id = lop_id).ma
    hks= Hocky.objects.filter(ma=hk_ma)
    st= HocphiStatus.objects.all()
    stud_list = Hssv.objects.filter(lop_id = lop_id)
    hps = Hocphi.objects.all().select_related("sv").filter(sv__in=stud_list, hk=hk_ma).order_by('sv_id')
    if request.method == "POST":
        for stud in stud_list:
            for hk in hks:
                hs = Hocphi.objects.get(sv_id = stud.id,hk=hk.ma,lop_id = lop_id)
                hs.hpstatus = request.POST["status"+str(stud.id)]
                hs.ghichu = request.POST["ghichu"+str(stud.id)]
                hs.status = 1
                hs.save()
        messages.success(request, "Cap nhat thong tin hoc phi thanh cong!")
        return redirect("lop81-lst", lop_id)


    if not hps.first():
        #create hoc phi record
        for stud in stud_list:
            for hk in hks:
                hp = Hocphi(sv_id = stud.id, lop_id = lop_id, hk=hk.ma)
                hp.save()
        hps = Hocphi.objects.all().select_related("sv").filter(sv__in=stud_list, hk=hk_ma).order_by('sv_id')

    context = {
        "hps": hps,
        "st": st,
        "tenlop": tenlop,
        "hk_ma": hk_ma
    }
    return render(request, "sms/lophp-hk.html", context)

@login_required
def ctdt_list(request):
    ctdt = Ctdt.objects.all()

    paginator = Paginator(ctdt, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "ctdt": paged_students
    }
    return render(request, "sms/ctdt_list.html", context)

@login_required
def import_monhoc_dm(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        if 'mh-lst' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin mon hoc")
            #return redirect("ctdt_list")
        else:
            sheet = wb["mh-lst"]
            for r in range(4, sheet.max_row):
            #for r in range(sheet.max_row-1, sheet.max_row):
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                chuongtrinh = sheet.cell(r,3).value
                sotinchi = sheet.cell(r,4).value
                sogio_lt = sheet.cell(r,5).value
                sogio_th = sheet.cell(r,6).value
                sogio_kt = sheet.cell(r,7).value
                if not ma or not ten or not chuongtrinh or not sotinchi:
                    messages.error(request, 'Ma, ten, chuong trinh không đủ có thông tin bắt buộc')
                    continue
                if Monhoc.objects.filter(ma=ma, ten=ten, chuongtrinh=chuongtrinh).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    mh = Monhoc(ma=ma, ten=ten, chuongtrinh=chuongtrinh, sotinchi=sotinchi, sogio_lt=sogio_lt,sogio_th=sogio_th,sogio_kt=sogio_kt)
                    try:
                        mh.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))

        if 'hk-lst' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin hoc ky")
            #return redirect("ctdt_list")
        else:

            sheet = wb["hk-lst"]
            for r in range(2, sheet.max_row+1):
            #for r in range(sheet.max_row-1, sheet.max_row):
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                if not ma or not ten:
                    messages.error(request, 'Ma, ten không có thông tin')
                    continue
                if Hocky.objects.filter(ma=ma).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    hk = Hocky(ma=ma, ten=ten)
                    try:
                        hk.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))

        if 'ld-lst' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin loai diem")
            #return redirect("ctdt_list")
        else:

            sheet = wb["ld-lst"]
            for r in range(2, sheet.max_row+1):
            #for r in range(sheet.max_row-1, sheet.max_row):
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                trunglap = sheet.cell(r,3).value
                heso = sheet.cell(r,4).value
                if not ma or not ten or not trunglap or not heso:
                    messages.error(request, 'Ma, ten không có thông tin')
                    continue
                if Loaidiem.objects.filter(ma=ma).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    ld = Loaidiem(ma=ma, ten=ten,trunglap=trunglap, heso=heso)
                    try:
                        ld.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
                    #ld.save()

        if 'hp-st' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin tinh trang hoc phi")
            #return redirect("ctdt_list")

        else:
            sheet = wb["hp-st"]
            for r in range(2, sheet.max_row+1):
            #for r in range(sheet.max_row-1, sheet.max_row):
                if not ma or not ten:
                    messages.error(request, 'Ma, ten không có thông tin')
                    continue
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                if HocphiStatus.objects.filter(ma=ma).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    hp = HocphiStatus(ma=ma, ten=ten)
                    try:
                        hp.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
                    #hp.save()

        if 'sv-st' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin tinh trang sinh vien")
            #return redirect("ctdt_list")

        else:
            sheet = wb["sv-st"]
            for r in range(2, sheet.max_row+1):
            #for r in range(sheet.max_row-1, sheet.max_row):
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                if not ma or not ten:
                    messages.error(request, 'Ma, ten không có thông tin')
                    continue
                if SvStatus.objects.filter(ma=ma).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    hp = SvStatus(ma=ma, ten=ten)
                    try:
                        hp.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
                    #hp.save()

        if 'tt-lst' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin danh sách trung tâm")
            #return redirect("ctdt_list")

        else:
            sheet = wb["tt-lst"]
            for r in range(2, sheet.max_row+1):
            #for r in range(sheet.max_row-1, sheet.max_row):
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                if not ma or not ten:
                    messages.error(request, 'Ma, ten không có thông tin')
                    continue
                if Trungtam.objects.filter(ma=ma).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    hp = Trungtam(ma=ma, ten=ten)
                    try:
                        hp.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
#                    hp.save()

        if 'phong-lst' not in wb.sheetnames:
            messages.error(request, "File excel khong co thong tin danh sách phòng")
            #return redirect("ctdt_list")

        else:
            sheet = wb["phong-lst"]
            for r in range(2, sheet.max_row+1):
            #for r in range(sheet.max_row-1, sheet.max_row):
                ma = sheet.cell(r,1).value
                ten = sheet.cell(r,2).value
                if not ma or not ten:
                    messages.error(request, 'Ma, ten không có thông tin')
                    continue
                if Phong.objects.filter(ma=ma).exists():
                    messages.error(request, 'Ma: ' + str(ma) + ' already exists')
                else:
                    hp = Phong(ma=ma, ten=ten)
                    try:
                        hp.save()
                    except Exception as e:
                        messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
#                    hp.save()
        messages.success(request, "Import done!")
        return redirect("ctdt_list")
@login_required
def import_lopsv(request, lop_id):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        if 'lopsv' not in wb.sheetnames:
            messages.error(request, "File excel khong dung format")
            return redirect("svlop_list", lop_id)

        sheet = wb["lopsv"]
        #for r in range(3, sheet.max_row+1):
        #maxid=500
        #for r in range(3, 44):
        for r in range(3, sheet.max_row+1):
            #maxid = maxid+1
            v1 = sheet.cell(r,1).value
            v2 = sheet.cell(r,2).value
            v3 = sheet.cell(r,3).value
            v4 = sheet.cell(r,4).value
            v5 = sheet.cell(r,5).value
            v6 = sheet.cell(r,6).value
            v7 = sheet.cell(r,7).value
            v8 = sheet.cell(r,8).value
            v9 = sheet.cell(r,9).value
            v10 = sheet.cell(r,10).value
            v11 = sheet.cell(r,11).value
            v12 = sheet.cell(r,12).value
            v13 = sheet.cell(r,13).value
            v14 = sheet.cell(r,14).value
            v15 = sheet.cell(r,15).value
            v16 = sheet.cell(r,16).value
            v17 = sheet.cell(r,17).value
            v18 = sheet.cell(r,18).value
            v19 = sheet.cell(r,19).value
            v20 = sheet.cell(r,20).value
            v21 = sheet.cell(r,21).value
            v22 = sheet.cell(r,22).value
            v23 = sheet.cell(r,23).value
            v24 = sheet.cell(r,24).value
            v25 = sheet.cell(r,25).value
            v26 = sheet.cell(r,26).value
            v27 = sheet.cell(r,27).value
            v28 = sheet.cell(r,28).value
            v29 = sheet.cell(r,29).value
            v30 = sheet.cell(r,30).value
            v31 = sheet.cell(r,31).value
            v32 = sheet.cell(r,32).value
            #print(type(v4))
            if not v1 or not v2:
                messages.error(request, 'Dòng: '+str(r)+'có Ma, ten không có thông tin')
                continue
            if Hssv.objects.filter(msv=v1.strip()).exists():
                messages.error(request, 'Dòng: '+str(r)+' có Mã ' + v1+ ' already exists')
            # elif (v4 and type(v4) is not datetime) or (v14 and type(v14) is not datetime):
            #     messages.error(request, 'Dòng: '+str(r)+' có dữ liệu ngày không đúng format')
            else:
                sv = Hssv(
                    msv=v1, 
                    hoten = v2, 
#                    lop = v3, 
                    namsinh=v4, 
                    gioitinh=v5, 
                    dantoc=v6, 
                    noisinh=v7, 
                    quequan=v8, 
                    diachi=v9, 
                    xa=v10, 
                    huyen=v11, 
                    tinh=v12, 
                    cccd=v13, 
                    ngaycap=v14, 
                    noicap=v15,
                    stk=v16, 
                    nh=v17, 
                    hotenbo=v18, 
                    hotenme=v19,
                    sdths=v20, 
                    sdtph=v21, 
                    hs_syll= True if v22 == 1 else False,
                    hs_pxt= True if v23 == 1 else False,
                    hs_btn= True if v24 == 1 else False,
                    hs_gcntttt= True if v25 == 1 else False,
                    hs_hbthcs= True if v26 == 1 else False,
                    hs_cccd= True if v27 == 1 else False,
                    hs_gks= True if v28 == 1 else False,
                    hs_shk= True if v29 == 1 else False,
                    hs_a34= True if v30 == 1 else False,
                    hs_status= "Đủ" if v31 == 1 else "Thiếu",
                    ghichu=v32, 
                    lop_id= lop_id if lop_id else None)
                try:
                    sv.save()
                except Exception as e:
                    messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))

        messages.success(request, "Import thanh cong!")
        return redirect("svlop_list", lop_id) if lop_id else redirect("sv_list")

@login_required
def import_gv(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        print(sheets)
        if 'hsgv' not in wb.sheetnames:
            messages.error(request, "File excel khong dung format")
            return redirect("gv_list")

        sheet = wb["hsgv"]
        for r in range(2, sheet.max_row+1):
            v1 = sheet.cell(r,1).value
            v2 = sheet.cell(r,2).value
            v3 = sheet.cell(r,3).value
            v4 = sheet.cell(r,4).value
            v5 = sheet.cell(r,5).value
            v6 = sheet.cell(r,6).value
            v7 = sheet.cell(r,7).value
            v8 = sheet.cell(r,8).value
            v9 = sheet.cell(r,9).value
            v10 = sheet.cell(r,10).value
            v11 = sheet.cell(r,11).value
            v12 = sheet.cell(r,12).value
            v13 = sheet.cell(r,13).value
            v14 = sheet.cell(r,14).value
            v15 = sheet.cell(r,15).value
            v16 = sheet.cell(r,16).value
            v17 = sheet.cell(r,17).value
            v18 = sheet.cell(r,18).value
            v19 = sheet.cell(r,19).value
            v20 = sheet.cell(r,20).value
            v21 = sheet.cell(r,21).value
            v22 = sheet.cell(r,22).value
            v23 = sheet.cell(r,23).value
            v24 = sheet.cell(r,24).value
            v25 = sheet.cell(r,25).value
            v26 = sheet.cell(r,26).value
            v27 = sheet.cell(r,27).value
            v28 = sheet.cell(r,28).value
            v29 = sheet.cell(r,29).value
            if not v1 or not v2 or not v4 or not v7 or not v9:
                messages.error(request, 'Dòng ' + str(r+1) + ' Trường bắt buộc thiếu thông tin')
                continue
            if Hsgv.objects.filter(ma=v1).exists():
                messages.error(request, 'Ma: ' + v1 + ' already exists')
            # elif (v3 and type(v3) is not datetime) or (v5 and type(v5) is not datetime) or (v20 and type(v20) is not datetime) or (v21 and type(v21) is not datetime):
            #     messages.error(request, 'Dòng ' + str(r+1) + ' có trường date không đúng định dạng')
            else:
                gv = Hsgv(
                    ma = v1,
                    hoten = v2,
                    namsinh = v3,
                    cccd = v4,
                    ngaycap = v5,
                    noicap = v6,
                    diachi = v7,
                    sdt = v8,
                    email = v9,
                    mst = v10,
                    bhxh = v11,
                    dongbhxh = v12,
                    stk = v13,
                    nh = v14,
                    cn = v15,
                    trinhdo = v16,
                    truongtn = v17,
                    nganhtn = v18,
                    shdtg = v19,
                    ngayky = v20,
                    ngayhh = v21,
                    hs_btn = True if v22 == 1 else False,
                    hs_bd = True if v23 == 1 else False,
                    hs_cc = True if v24 == 1 else False,
                    hs_syll = True if v25 == 1 else False,
                    hs_ccta = True if v26 == 1 else False,
                    hs_ccth = True if v27 == 1 else False,
                    hs_status = "Đủ" if v28 == 1 else "Thiếu",
                    ghichu = v29
                )
                try:
                    gv.save()
                except Exception as e:
                    messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))

        messages.success(request, "Import thanh cong!")
        return redirect("gv_list")

def import_renter(request):
    if request.method == 'POST':
        file = request.FILES['file']
        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f'Lỗi đọc file: {e}')
            return redirect('renter_list')

        imported, errors = 0, 0
        for idx, row in df.iterrows():
            try:
                renter, created = Renter.objects.update_or_create(
                    hoten=row.get('hoten'),
                    chu_id=request.user.id,
                    defaults={
                        'email': row.get('email') if pd.notna(row.get('email')) else '',
                        'sdt': row.get('sdt') if pd.notna(row.get('sdt')) else '',
                        'cccd': row.get('cccd') if pd.notna(row.get('cccd')) else '',
                        'ngaycap': parse_date_auto(row['ngaycap']),
                        'noicap': row.get('noicap') if pd.notna(row.get('noicap')) else '',
                        'mst': row.get('mst') if pd.notna(row.get('mst')) else '',
                        'ghichu': row.get('ghichu') if pd.notna(row.get('ghichu')) else '',
                    }
                )
                if created:
                    renter.ma = "NT"+str(request.user.id)+str(renter.id).zfill(3)
                    renter.save()

                imported += 1
            except IntegrityError as e:
                print (e)
                errors += 1
            except Exception as e:
                print (e)
                errors += 1
        messages.success(request, f'Đã import {imported} người thuê. Lỗi: {errors}')
        return redirect('renter_list')


@login_required
def import_ns(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        print(sheets)
        #if 'hsgv' not in wb.sheetnames:
        #    messages.error(request, "File excel khong dung format hsgv")
        #    return redirect("gv_list")
        sheet = wb["hsns"]
        for r in range(3, sheet.max_row+1):
            v1 = sheet.cell(r,1).value
            v2 = sheet.cell(r,2).value
            v3 = sheet.cell(r,3).value
            v4 = sheet.cell(r,4).value
            v5 = sheet.cell(r,5).value
            v6 = sheet.cell(r,6).value
            v7 = sheet.cell(r,7).value
            v8 = sheet.cell(r,8).value
            v9 = sheet.cell(r,9).value
            v10 = sheet.cell(r,10).value
            v11 = sheet.cell(r,11).value
            v12 = sheet.cell(r,12).value
            v13 = sheet.cell(r,13).value
            v14 = sheet.cell(r,14).value
            v15 = sheet.cell(r,15).value
            v16 = sheet.cell(r,16).value
            v17 = sheet.cell(r,17).value
            v18 = sheet.cell(r,18).value
            v19 = sheet.cell(r,19).value
            v20 = sheet.cell(r,20).value

            v21 = sheet.cell(r,21).value
            v22 = sheet.cell(r,22).value
            v23 = sheet.cell(r,23).value
            v24 = sheet.cell(r,24).value
            v25 = sheet.cell(r,25).value
            v26 = sheet.cell(r,26).value
            v27 = sheet.cell(r,27).value
            v28 = sheet.cell(r,28).value
            v29 = sheet.cell(r,29).value
            v30 = sheet.cell(r,30).value

            v31 = sheet.cell(r,31).value
            v32 = sheet.cell(r,32).value
            v33 = sheet.cell(r,33).value
            v34 = sheet.cell(r,34).value
            v35 = sheet.cell(r,35).value
            v36 = sheet.cell(r,36).value
            v37 = sheet.cell(r,37).value
            v38 = sheet.cell(r,38).value
            v39 = sheet.cell(r,39).value
            v40 = sheet.cell(r,40).value

            v41 = sheet.cell(r,41).value
            v42 = sheet.cell(r,42).value
            v43 = sheet.cell(r,43).value
            v44 = sheet.cell(r,44).value
            v45 = sheet.cell(r,45).value
            v46 = sheet.cell(r,46).value
            v47 = sheet.cell(r,47).value
            v48 = sheet.cell(r,48).value

            if not v1 or not v2 or not v3:
                messages.error(request, 'Dòng: '+str(r)+' có Ma, email, ten không có thông tin')
                continue
            if Hsns.objects.filter(ma=v1.strip()).exists():
                messages.error(request, 'Dòng: '+str(r) + 'có Mã: ' + v1 + ' already exists')
            if Hsns.objects.filter(email=v3.strip()).exists():
                messages.error(request, 'Dòng: '+str(r)+ ' có Email: ' + v3 + ' already exists')
            # elif (v5 and type(v5) is not datetime) or (v14 and type(v14) is not datetime):
            #     messages.error(request, 'Dòng: '+str(r) + ' có dữ liệu ngày không đúng format')
            # elif (v25 and type(v25) is not datetime) or (v26 and type(v26) is not datetime):
            #     messages.error(request, 'Dòng: '+str(r) + ' có dữ liệu ngày không đúng format')
            # elif v31 and type(v31) is not datetime:
            #     messages.error(request, 'Dòng: '+str(r) + ' có dữ liệu ngày không đúng format')
            else:
                ns = Hsns(
                    ma = v1,
                    hoten = v2,
                    email = v3,
                    gioitinh = v4,
                    namsinh = v5,
                    dantoc = v6,
                    tongiao = v7,
                    quoctich = v8,
                    quequan = v9,
                    diachi1 = v10,
                    diachi2 = v11,
                    sdt = v12,
                    cccd = v13,
                    ngaycap = v14,
                    noicap = v15,
                    mst = v16,
                    tddt = v17,
                    noidt = v18,
                    kdt = v19,
                    cndt = v20,
                    namtn = v21,
                    xldt = v22,
                    cdcv = v23,
                    vtcv = v24,
                    shd = v25,
                    ngayky = v26,
                    ngayhh = v27,
                    trangthaihd = v28,
                    tcld = v29,
                    loaihd = v30,
                    ngaylv = v31,
                    tgcd = v32,
                    tgbhxh = v33,
                    ssbhxh = v34,

                    tongluong = v35,
                    luongcb = v36,
                    stk = v37,
                    nh = v38,
                    chinhanh = v39,

                    hs_btn = True if v40 == 1 else False,
                    hs_bd = True if v41 == 1 else False,
                    hs_cc = True if v42 == 1 else False,
                    hs_syll = True if v43 == 1 else False,
                    hs_ccta = True if v44 == 1 else False,
                    hs_ccth = True if v45 == 1 else False,
                    hs_khac = v46,
                    hs_status = "Đủ" if v47 == 1 else "Thiếu",
                    ghichu = v48
                        )
                try:
                    ns.save()
                except Exception as e:
                    messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
 
        messages.success(request, "Import thanh cong!")
        return redirect("ns_list")

@login_required
def export_gv(request):
    # Query the Person model to get all records
    gvs = Hsgv.objects.all().values()
    # Convert the QuerySet to a DataFrame
    df = pd.DataFrame(list(gvs))

    # Define the Excel file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=giaovien.xlsx'

    # Use Pandas to write the DataFrame to an Excel file
    df.to_excel(response, index=False, engine='openpyxl')

    return response

@login_required
def export_lopsv(request, lop_id):
    # Query the Person model to get all records
    #svs = Hssv.objects.all().filter(lop_id=lop_id).values().order_by("ten")
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
    exp=[]
    for sv in svs:
        exp.append({"Mã học tên": sv.msv,"Họ tên": sv.hoten})

    # Convert the QuerySet to a DataFrame
    df = pd.DataFrame(list(exp))

    # Define the Excel file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hocvien.xlsx'

    # Use Pandas to write the DataFrame to an Excel file
    df.to_excel(response, index=False, engine='openpyxl')

    return response

@login_required
def export_lh(request):
    # Query the Person model to get all records
    gvs = Lichhoc.objects.all().values()
    # Convert the QuerySet to a DataFrame
    df = pd.DataFrame(list(gvs))

    # Define the Excel file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lichhoc.xlsx'

    # Use Pandas to write the DataFrame to an Excel file
    df.to_excel(response, index=False, engine='openpyxl')

    return response

@login_required
@permission_required('sms.view_lop',raise_exception=True)
def lop_list(request):

    if request.user.is_superuser or request.user.is_supervisor:
        lop = Lop.objects.all().select_related("ctdt").order_by('id')
    elif request.user.is_gv:

        gv = Hsgv.objects.get(user_id = request.user.id)

        nl = GvLmh.objects.filter(gv_id = gv.id, status = 1).select_related("lopmh")
        print("lop_id: ")
        print(nl[0].lopmh.lop_id)
        #gvs = Hsgv.objects.filter(id__in = lh.values_list('giaovien_id', flat=True))
        lop = Lop.objects.filter(id__in=[ns.lopmh.lop_id for ns in nl]).select_related("ctdt").order_by('id')

        print("lop_id: ")
        print(lop[0].id)


    elif request.user.is_internalstaff:

        ns = Hsns.objects.get(user_id = request.user.id)

        nl = NsLop.objects.filter(ns_id = ns.id, status = 1)
        #gvs = Hsgv.objects.filter(id__in = lh.values_list('giaovien_id', flat=True))
        lop = Lop.objects.filter(id__in=[ns.lop_id for ns in nl]).select_related("ctdt").order_by('id')
    else:
        return HttpResponseForbidden("Bạn không có quyền truy cập trang này")

    print(Lop.history.all())
    #print(request.user.username)
    paginator = Paginator(lop, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lop": paged_students
    }
    return render(request, "sms/lop_list.html", context)

def index1(request):
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth
    from datetime import datetime
    
    # Thống kê cho landing page
    stats = {
        'total_locations': Location.objects.count(),
        'total_houses': House.objects.count(),
        'total_renters': HouseRenter.objects.filter(active=True).values('renter').distinct().count(),
        'revenue_month': 0,
    }
    
    # Doanh thu tháng hiện tại
    current_month = datetime.now().month
    current_year = datetime.now().year
    revenue = Hoadon.objects.filter(
        ngay_tao__month=current_month,
        ngay_tao__year=current_year
    ).aggregate(total=Sum('SO_TIEN_DA_TRA'))['total'] or 0
    stats['revenue_month'] = revenue
    
    context = {
        "stats": stats
    }
    return render(request, "sms/index.html", context)

def renter_landing(request):
    """Landing page riêng cho renter với thiết kế đẹp"""
    from django.db.models import Sum
    from datetime import datetime
    
    # Thống kê tổng quan cho landing page
    stats = {
        'total_locations': Location.objects.count(),
        'total_houses': House.objects.count(),
        'total_renters': HouseRenter.objects.filter(active=True).values('renter').distinct().count(),
        'revenue_month': 0,
    }
    
    # Doanh thu tháng hiện tại (chỉ hiển thị cho superuser hoặc landlord)
    if request.user.is_authenticated:
        current_month = datetime.now().month
        current_year = datetime.now().year
        if request.user.is_superuser:
            revenue = Hoadon.objects.filter(
                ngay_tao__month=current_month,
                ngay_tao__year=current_year
            ).aggregate(total=Sum('SO_TIEN_DA_TRA'))['total'] or 0
        else:
            # Chỉ tính doanh thu của chủ nhà hiện tại
            revenue = Hoadon.objects.filter(
                house__loc__chu=request.user,
                ngay_tao__month=current_month,
                ngay_tao__year=current_year
            ).aggregate(total=Sum('SO_TIEN_DA_TRA'))['total'] or 0
        stats['revenue_month'] = int(revenue) if revenue else 0
    
    context = {
        'stats': stats
    }
    
    return render(request, 'sms/renter_landing.html', context)

@login_required
def renter_dashboard(request):
    """Dashboard riêng cho renter - hiển thị hợp đồng và hóa đơn"""
    try:
        # Lấy thông tin renter từ user hiện tại
        renter = Renter.objects.get(user=request.user)
    except Renter.DoesNotExist:
        messages.error(request, "Bạn chưa được liên kết với tài khoản khách thuê.")
        return redirect('renter_landing')
    
    # DEBUG: Log renter info
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Renter Dashboard - User: {request.user.username}, Renter ID: {renter.id}, Renter Name: {renter.hoten}")
    
    # Lấy tất cả hợp đồng của renter (chỉ lấy hợp đồng có house)
    contracts = HouseRenter.objects.filter(
        renter=renter,
        house__isnull=False
    ).select_related('house', 'house__loc', 'house__loc__chu').order_by('-active', '-rent_from')
    
    # DEBUG: Log contracts
    logger.info(f"Contracts count: {contracts.count()}")
    for c in contracts:
        logger.info(f"  - Contract: {c.house.ten}, Renter: {c.renter.hoten} (ID: {c.renter.id})")
    
    # Lấy tất cả hóa đơn liên quan đến renter (chỉ lấy hóa đơn có house)
    bills = Hoadon.objects.filter(
        renter=renter,
        house__isnull=False
    ).select_related('house', 'house__loc').order_by('-ngay_tao')
    
    # DEBUG: Log bills
    logger.info(f"Bills count: {bills.count()}")
    for b in bills[:5]:  # Log first 5 bills
        logger.info(f"  - Bill ID: {b.id}, House: {b.house.ten}, Renter: {b.renter.hoten if b.renter else 'None'} (ID: {b.renter.id if b.renter else 'None'})")
    
    # Thống kê
    active_contracts = contracts.filter(active=True).count()
    total_contracts = contracts.count()
    
    # Thống kê hóa đơn
    from django.db.models import Q, Count
    bill_stats = {
        'total': bills.count(),
        'unpaid': bills.filter(status='ChuaTT').count(),
        'paying': bills.filter(status='DangTT').count(),
        'paid': bills.filter(status='DaTT').count(),
        'overdue': bills.filter(status='QuaHan').count(),
    }
    
    # Tổng công nợ
    total_debt = bills.filter(
        Q(status='ChuaTT') | Q(status='DangTT') | Q(status='QuaHan')
    ).aggregate(
        total=Sum('CONG_NO')
    )['total'] or 0
    
    context = {
        'renter': renter,
        'contracts': contracts,
        'bills': bills[:10],  # 10 hóa đơn gần nhất
        'active_contracts': active_contracts,
        'total_contracts': total_contracts,
        'bill_stats': bill_stats,
        'total_debt': total_debt,
    }
    
    return render(request, 'sms/renter_dashboard.html', context)

@login_required
def renter_bills(request):
    """Trang xem tất cả hóa đơn của renter"""
    try:
        renter = Renter.objects.get(user=request.user)
    except Renter.DoesNotExist:
        messages.error(request, "Bạn chưa được liên kết với tài khoản khách thuê.")
        return redirect('renter_landing')
    
    # Lấy tất cả hóa đơn
    bills = Hoadon.objects.filter(
        renter=renter,
        house__isnull=False
    ).select_related('house', 'house__loc').order_by('-ngay_tao')
    
    # Lọc theo trạng thái nếu có
    status_filter = request.GET.get('status', '')
    if status_filter:
        bills = bills.filter(status=status_filter)
    
    # Thống kê
    from django.db.models import Q, Sum
    bill_stats = {
        'total': bills.count(),
        'unpaid': bills.filter(status='ChuaTT').count(),
        'paying': bills.filter(status='DangTT').count(),
        'paid': bills.filter(status='DaTT').count(),
        'overdue': bills.filter(status='QuaHan').count(),
    }
    
    total_debt = bills.filter(
        Q(status='ChuaTT') | Q(status='DangTT') | Q(status='QuaHan')
    ).aggregate(total=Sum('CONG_NO'))['total'] or 0
    
    context = {
        'renter': renter,
        'bills': bills,
        'bill_stats': bill_stats,
        'total_debt': total_debt,
        'status_filter': status_filter,
    }
    
    return render(request, 'sms/renter_bills.html', context)

@login_required
def renter_locations(request):
    """Trang xem các địa điểm mà renter có hợp đồng"""
    try:
        renter = Renter.objects.get(user=request.user)
    except Renter.DoesNotExist:
        messages.error(request, "Bạn chưa được liên kết với tài khoản khách thuê.")
        return redirect('renter_landing')
    
    # Lấy tất cả location mà renter có hợp đồng
    location_ids = HouseRenter.objects.filter(
        renter=renter,
        house__isnull=False
    ).values_list('house__loc_id', flat=True).distinct()
    
    locations = Location.objects.filter(
        id__in=location_ids
    ).select_related('chu', 'xp').prefetch_related('house_set')
    
    # Thêm thông tin hợp đồng và files cho mỗi location
    for loc in locations:
        houses = House.objects.filter(
            loc=loc,
            houserenter__renter=renter
        ).distinct()
        
        # Thêm thông tin hợp đồng cho mỗi house
        for house in houses:
            house.renter_contract = HouseRenter.objects.filter(
                house=house,
                renter=renter
            ).order_by('-rent_from').first()
        
        loc.renter_houses = houses
        
        # Lấy files đính kèm của location
        files = UploadedFile.objects.filter(link_id=loc.id, type=1)
        for file in files:
            file.ten = file.file.name.split("/")[-1]
        loc.files = files
        # Địa chỉ đầy đủ (dùng cho Google Maps)
        try:
            # Địa chỉ theo định dạng giống trang chủ nhà: "diachi, xa-phuong, tinh-thanh"
            full_parts = [p for p in [loc.diachi, getattr(loc.xp, 'ten', ''), getattr(loc.xp, 'tp', '')] if p]
            loc.full_address = ', '.join(full_parts)
        except Exception:
            loc.full_address = loc.diachi or ''
    
    context = {
        'renter': renter,
        'locations': locations,
    }
    
    return render(request, 'sms/renter_locations.html', context)

@login_required
def renter_contracts(request):
    """Trang xem tất cả hợp đồng của renter"""
    try:
        renter = Renter.objects.get(user=request.user)
    except Renter.DoesNotExist:
        messages.error(request, "Bạn chưa được liên kết với tài khoản khách thuê.")
        return redirect('renter_landing')
    
    # Lấy tất cả hợp đồng của renter
    contracts = HouseRenter.objects.filter(
        renter=renter,
        house__isnull=False
    ).select_related(
        'house',
        'house__loc',
        'house__loc__chu'
    ).order_by('-active', '-rent_from')
    
    # Thống kê
    total_contracts = contracts.count()
    active_contracts = contracts.filter(active=True).count()
    expired_contracts = contracts.filter(active=False).count()
    
    context = {
        'renter': renter,
        'contracts': contracts,
        'total_contracts': total_contracts,
        'active_contracts': active_contracts,
        'expired_contracts': expired_contracts,
    }
    
    return render(request, 'sms/renter_contracts.html', context)

@login_required
@permission_required('sms.view_lop',raise_exception=True)
def lop_list_guardian(request):

    list_of_ids = []
    for l in Lop.objects.all():
        if request.user.has_perm('assign_lop', l):
            list_of_ids.append(l.id)
            print("lop_id: ")
            print(l.id)

    lop = Lop.objects.filter(id__in=list_of_ids).select_related("ctdt").order_by('id')

#    lop = Lop.objects.all().select_related("ctdt").order_by('id')
    paginator = Paginator(lop, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lop": paged_students
    }
    return render(request, "sms/lop_list.html", context)

@login_required
def loc_list(request):
    # Chặn renter truy cập - chuyển về trang locations của họ
    try:
        if request.user.is_renter:
            messages.info(request, "Vui lòng xem địa điểm của bạn tại đây.")
            return redirect('renter_locations')
    except (AttributeError, Renter.DoesNotExist):
        pass

    loca = Location.objects.filter(chu=request.user).order_by('id')
    paginator = Paginator(loca, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "loca": paged_students
    }
    return render(request, "sms/loca_list.html", context)

@login_required
def house_list(request, loc_id):
    # Chặn renter truy cập - chuyển về trang locations của họ
    try:
        if request.user.is_renter:
            messages.info(request, "Vui lòng xem địa điểm của bạn tại đây.")
            return redirect('renter_locations')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Chỉ xem được nhà thuộc location của chính chủ
    loc = get_object_or_404(Location, pk=loc_id, chu=request.user) if not request.user.is_superuser else get_object_or_404(Location, pk=loc_id)
    house = House.objects.filter(loc_id=loc.id).order_by('id')
    paginator = Paginator(house, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "house": paged_students,
        "loc_id": loc_id
    }
    return render(request, "sms/house_list.html", context)

@login_required
def houses(request):
    """Consolidated Houses page with location filter and quick actions.
    - Shows a location dropdown scoped to the current landlord (or all for superuser)
    - Lists houses for the selected location with actions: Edit, Manage contracts
    """
    # Block renters; redirect them to their own locations page
    try:
        if request.user.is_renter:
            messages.info(request, "Vui lòng xem địa điểm của bạn tại đây.")
            return redirect('renter_locations')
    except (AttributeError, Renter.DoesNotExist):
        pass

    # Locations available to the current user
    if request.user.is_superuser:
        locations = Location.objects.all().order_by('id')
    else:
        locations = Location.objects.filter(chu=request.user).order_by('id')

    # Selected location from querystring
    selected_loc_id = request.GET.get('loc')
    selected_loc = None
    houses_qs = House.objects.none()

    if selected_loc_id:
        # Validate access to selected location
        selected_loc, error_response = validate_location_access(request, selected_loc_id)
        if error_response:
            return error_response
    elif locations.exists():
        # Default to the first location for convenience
        selected_loc = locations.first()

    if selected_loc:
        houses_qs = House.objects.filter(loc=selected_loc).order_by('id')

    # Optional pagination for long lists
    paginator = Paginator(houses_qs, 20)
    page = request.GET.get('page')
    paged_houses = paginator.get_page(page)

    context = {
        'locations': locations,
        'selected_loc': selected_loc,
        'houses': paged_houses,
    }
    return render(request, 'sms/houses.html', context)

@login_required
def houses_partial(request):
    """HTMX fragment for the houses table filtered by location (query param: ?loc=<id>)."""
    # Block renters
    try:
        if request.user.is_renter:
            messages.info(request, "Vui lòng xem địa điểm của bạn tại đây.")
            return redirect('renter_locations')
    except (AttributeError, Renter.DoesNotExist):
        pass

    # Determine available locations
    locations_qs = Location.objects.all().order_by('id') if request.user.is_superuser else Location.objects.filter(chu=request.user).order_by('id')

    loc_id = request.GET.get('loc')
    selected_loc = None
    houses_qs = House.objects.none()

    if loc_id:
        selected_loc, error_response = validate_location_access(request, loc_id)
        if error_response:
            return error_response
    elif locations_qs.exists():
        selected_loc = locations_qs.first()

    if selected_loc:
        houses_qs = House.objects.filter(loc=selected_loc).order_by('id')

    paginator = Paginator(houses_qs, 20)
    page = request.GET.get('page')
    paged_houses = paginator.get_page(page)

    return render(request, 'sms/fragments/_houses_table.html', {
        'houses': paged_houses,
        'selected_loc': selected_loc,
    })

@login_required
@permission_required('sms.view_lop',raise_exception=True)
def xlop_list_guardian(request):

    list_of_ids = []
    for l in Lop.objects.all():
        if request.user.has_perm('assign_lop', l):
            list_of_ids.append(l.id)
            print("lop_id: ")
            print(l.id)

    lop = Lop.objects.filter(id__in=list_of_ids).select_related("ctdt").order_by('id')

#    lop = Lop.objects.all().select_related("ctdt").order_by('id')
    paginator = Paginator(lop, 20)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lop": paged_students
    }
    return render(request, "sms/xlop_list.html", context)

@login_required
def lichhoc_list(request):
    # phân quyền xem lịch học
    if request.user.is_gv:
        list_of_ids = []
        for l in LopMonhoc.objects.all():
            if request.user.has_perm('assign_lopmonhoc', l):
                list_of_ids.append(l.id)
                print(l)
        lmh = LopMonhoc.objects.filter(id__in=list_of_ids)

        gv= Hsgv.objects.get(user = request.user)
        lhgv = Lichhoc.objects.filter(giaovien=gv)

        lh = Lichhoc.objects.filter(lmh__in=lmh)
        lh = (lh | lhgv).distinct()

    elif request.user.is_hv:
        sv= Hssv.objects.get(user = request.user)
        lmh = LopMonhoc.objects.filter(lop_id=sv.lop_id)
        lh = Lichhoc.objects.filter(lmh__in=lmh)
    else:
        list_of_ids = []
        for l in Lop.objects.all():
            if request.user.has_perm('assign_lop', l):
                list_of_ids.append(l.id)
        lmh = LopMonhoc.objects.filter(lop_id__in=list_of_ids)
        lh = Lichhoc.objects.filter(lmh__in=lmh)


    lh = lh.select_related("lmh").order_by('thoigian')
    

    if request.method == "POST":
            query_lop = request.POST.get('lop', None)
            query_mh = request.POST.get('monhoc', None)
            query_tgian1 = request.POST.get('tgian1', None)
            query_tgian2 = request.POST.get('tgian2', None)
            if query_lop:
                lh = lh.filter(lmh__lop__ten__contains=query_lop)     
            if query_mh:
                lh = lh.filter(lmh__monhoc__ten__contains=query_mh)     
            if query_tgian1:
                lh = lh.filter(thoigian__gte=query_tgian1)     
            if query_tgian2:
                lh = lh.filter(thoigian__lte=query_tgian2)     
            messages.success(request, "Tìm kiếm thành công!")
#    else:
#        lh = lh.filter(thoigian__gte=datetime.now()).order_by('thoigian') if lh else lh

    if lh.exists():
        ngay_first=lh[0].thoigian
        count=0
        for l in lh:
        
            if l.thoigian.date() == ngay_first.date() and count == 0:
                l.ngay=l.thoigian
                count=count+1
            elif l.thoigian.date() == ngay_first.date() and count > 0:
                l.ngay=None
                count=count+1
            elif l.thoigian.date() != ngay_first.date():    
                l.ngay=l.thoigian
                ngay_first = l.thoigian
                count=count+1
            

    paginator = Paginator(lh, 30)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lh": paged_students
    }
    return render(request, "sms/lichhoc_list.html", context)

@login_required
def diemdanh_list(request, lh_id):
    lh = Diemdanh.objects.all().select_related("sv", "lichhoc").filter(lichhoc_id=lh_id)
    ttlh = Lichhoc.objects.all().select_related("lop", "monhoc").get(id = lh_id)
    paginator = Paginator(lh, 50)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "lh": paged_students,
        "ttlh": ttlh

    }
    return render(request, "sms/lichhoc-diemdanh_list.html", context)

@login_required
def diemdanh_lop(request, lh_id):
    ttlh = Lichhoc.objects.get(id = lh_id)
    lmh = LopMonhoc.objects.get(id =ttlh.lmh_id)
    
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    svlop = sorted(Hssv.objects.filter(lop_id = lmh.lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
    #svlop = Hssv.objects.filter(lop_id = lmh.lop_id)
    mh = Monhoc.objects.get(id = lmh.monhoc_id)
    dds = Diemdanh.objects.filter(lichhoc_id = lh_id).order_by('id')
    #sv = Hssv.objects.filter(lop_id = ttlh.lop_id)
    if request.method == "POST":
        for stud in svlop:
            id = "C"+str(stud.id)
            status = request.POST.get(id, None)
            dd = Diemdanh.objects.get(lichhoc_id = lh_id, sv_id=stud.id)
            dd.status= 1 if status else 0
            dd.save()
        #ttlh.status=1
        #ttlh.save()
        messages.success(request, "Cap nhat diem danh thanh cong!")
        return redirect("lichhoclopmh_list", lmh.id)


    for stud in svlop:
        if not Diemdanh.objects.filter(sv_id = stud.id, lichhoc_id = lh_id).first():
            dd = Diemdanh(sv_id = stud.id, lichhoc_id = lh_id, status=1)
            dd.save()

    dds = Diemdanh.objects.select_related("sv").filter(lichhoc_id = lh_id).order_by('id')
    context = {
        "dds": dds,
        "lmh": lmh,
        "mh": mh,
        "ttlh": ttlh
    }
    return render(request, "sms/diemdanh.html", context)

@login_required
def ctdt_monhoc(request, ctdt_id):
    mhs = Monhoc.objects.all()
    cms = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id).order_by('monhoc_id')
    if request.method == "POST":
        list_of_ids = []

        for mh in cms:
            id = "C"+str(mh.id)
            status = request.POST.get(id, None)
            
            print(id)
            print(status)
            if status:
                list_of_ids.append(mh.id)
        cms.update(status = 0)
        cms.filter(id__in = list_of_ids).update(status=1)
        #     id = "C"+str(stud.id)
        #     status = request.POST[id]
        #     dd = Diemdanh.objects.get(lichhoc_id = lh_id, sv_id=stud.id)
        #     dd.status=status
        #     dd.save() 
        # ttlh.status=1
        # ttlh.save()
        messages.success(request, "Cập nhật môn học thành công!")
        #return redirect("ctdt_list")


    for mh in mhs:
        if not CtdtMonhoc.objects.filter(ctdt_id = ctdt_id, monhoc_id = mh.id).first():
            dd = CtdtMonhoc(ctdt_id = ctdt_id, monhoc_id = mh.id)
            dd.save()

    cms = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id).select_related("monhoc").order_by('-status','monhoc_id')
    context = {
        "ctdt_id": ctdt_id,
        "cms": cms
    }
    return render(request, "sms/monhoc-ctdt-htmx.html", context)

@login_required
def ns_lop(request, ns_id):

    ns = Hsns.objects.get(id = ns_id)
    ls = Lop.objects.all()
    #nls = NsLop.objects.filter(ns_id = ns_id).order_by('lop_id')
    if request.method == "POST":
        for l in ls:
            id = "C"+str(l.id)
            status = request.POST.get(id, None)
            nl = NsLop.objects.get(ns_id = ns_id, lop_id = l.id)
            print(status)
            nl.status = 1 if status else 0
            nl.save()
            #Add Nhân sự to assign_lop permission
            if ns.user:
                if status:
                    if not ns.user.has_perm('assign_lop', l):
                        assign_perm('assign_lop', ns.user, l)
                else:
                    if ns.user.has_perm('assign_lop', l):
                        remove_perm('assign_lop', ns.user, l)

        messages.success(request, "Cập nhật lớp thành công!")
        return redirect("ns_list")

    for l in ls:
        if not NsLop.objects.filter(ns_id = ns_id, lop_id = l.id).first():
            dd = NsLop(ns_id = ns_id, lop_id = l.id)
            dd.save()

    nls = NsLop.objects.filter(ns_id = ns_id).select_related("lop").order_by('lop_id')
    context = {
        "ns": ns,
        "nls": nls
    }
    return render(request, "sms/ns_lop.html", context)

@login_required
@permission_required('sms.add_gvlop',raise_exception=True)
def gv_lop(request, gv_id):

    gv = Hsgv.objects.get(id = gv_id)
    ls = Lop.objects.all()
    #nls = GvLop.objects.filter(gv_id = gv_id).order_by('lop_id')
    if request.method == "POST":
        for l in ls:
            id = "C"+str(l.id)
            status = request.POST[id]
            print(id)
            print(status)
            nl = GvLop.objects.get(gv_id = gv_id, lop_id = l.id)
            nl.status=status
            nl.save() 

        #     id = "C"+str(stud.id)
        #     status = request.POST[id]
        #     dd = Diemdanh.objects.get(lichhoc_id = lh_id, sv_id=stud.id)
        #     dd.status=status
        #     dd.save() 
        # ttlh.status=1
        # ttlh.save()
        messages.success(request, "Cập nhật lớp thành công!")
        return redirect("gv_list")

    for l in ls:
        if not GvLop.objects.filter(gv_id = gv_id, lop_id = l.id).first():
            dd = GvLop(gv_id = gv_id, lop_id = l.id)
            dd.save()

    gls = GvLop.objects.filter(gv_id = gv_id).select_related("lop").order_by('lop_id')
    context = {
        "gv_id": gv_id,
        "gls": gls
    }
    return render(request, "sms/gv_lop.html", context)

@login_required
@permission_required('sms.add_gvmonhoc',raise_exception=True)
def gv_monhoc(request, gv_id):
    mhs = Monhoc.objects.all()
    nls = GvMonhoc.objects.filter(gv_id = gv_id).order_by('monhoc_id')
    if request.method == "POST":
        for mh in mhs:
            id = "C"+str(mh.id)
            status = request.POST[id]
            print(id)
            print(status)
            gmh = GvMonhoc.objects.get(gv_id = gv_id, monhoc_id = mh.id)
            gmh.status=status
            gmh.save() 
        #     id = "C"+str(stud.id)
        #     status = request.POST[id]
        #     dd = Diemdanh.objects.get(lichhoc_id = lh_id, sv_id=stud.id)
        #     dd.status=status
        #     dd.save() 
        # ttlh.status=1
        # ttlh.save()
        messages.success(request, "Cập nhật môn học thành công!")
        return redirect("gv_list")

    for mh in mhs:
        if not GvMonhoc.objects.filter(gv_id = gv_id, monhoc_id = mh.id).first():
            dd = GvMonhoc(gv_id = gv_id, monhoc_id = mh.id)
            dd.save()

    gms = GvMonhoc.objects.filter(gv_id = gv_id).select_related("monhoc").order_by('monhoc_id')
    context = {
        "gv_id": gv_id,
        "gms": gms
    }
    return render(request, "sms/gv_monhoc.html", context)

@login_required
@permission_required('sms.add_gvlmh',raise_exception=True)
def gv_lmh(request, gv_id):
    lmhs = LopMonhoc.objects.all()
    glmhs = GvLmh.objects.filter(gv_id = gv_id)
    gv = Hsgv.objects.get(id = gv_id)
    if request.method == "POST":
        for mh in glmhs:
            id = "C"+str(mh.lopmh_id)
            status = request.POST[id]
            gmh = GvLmh.objects.get(gv_id = gv_id, lopmh_id = mh.lopmh_id)
            gmh.status=status
            gmh.save() 

            #Add Giao viên to assign_lopmonhoc permission
            lmh = LopMonhoc.objects.get(id = mh.lopmh_id)
            if status == "1":
                if not gv.user.has_perm('assign_lopmonhoc', lmh):
                    assign_perm('assign_lopmonhoc', gv.user, lmh)
            else:
                if gv.user.has_perm('assign_lopmonhoc', lmh):
                    remove_perm('assign_lopmonhoc', gv.user, lmh)

        lop = Lop.objects.all()
        for l in lop:
            lmhs = LopMonhoc.objects.filter(lop_id = l.id)

            # group_name = "Lop_permisison_"+str(l.id)
            # if Group.objects.filter(name=group_name).exists():
            #     group = Group.objects.get(name=group_name)
            #     if GvLmh.objects.filter(gv_id = gv_id, status =1, lopmh__in = lmhs).first():
            #         if not gv.user.groups.filter(name=group_name).exists():
            #             gv.user.groups.add(group)
            #     else:
            #         if gv.user.groups.filter(name=group_name).exists():
            #             gv.user.groups.remove(group)

            #Add Gv to assign_lop permission
            if GvLmh.objects.filter(gv_id = gv_id, status =1, lopmh__in = lmhs).first():
                if not gv.user.has_perm('assign_lop', l):
                    assign_perm('assign_lop', gv.user, l)
            else:
                if gv.user.has_perm('assign_lop', l):
                    remove_perm('assign_lop', gv.user, l)

        messages.success(request, "Cập nhật môn học thành công!")
        return redirect("gv_list")

    for lmh in lmhs:
        if not GvLmh.objects.filter(gv_id = gv_id, lopmh_id = lmh.id).first():
            dd = GvLmh(gv_id = gv_id, lopmh_id = lmh.id)
            dd.save()

    glmhs = GvLmh.objects.filter(gv_id = gv_id).select_related("lopmh")
    context = {
        "gv": gv,
        "glmhs": glmhs
    }
    return render(request, "sms/gv_lopmh.html", context)

@login_required
def single_ctdtmonhoc_old(request, ctdt_id):
        ctdtmonhoc1 = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id, hocky = 1).select_related("monhoc")
        ctdtmonhoc2 = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id, hocky = 2).select_related("monhoc")
        ctdtmonhoc3 = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id, hocky = 3).select_related("monhoc")
        ctdtmonhoc4 = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id, hocky = 4).select_related("monhoc")
        ten = Ctdt.objects.get(id = ctdt_id).ten
        context = {
            "ctdtmonhoc1": ctdtmonhoc1,
            "ctdtmonhoc2": ctdtmonhoc2,
            "ctdtmonhoc3": ctdtmonhoc3,
            "ctdtmonhoc4": ctdtmonhoc4,
            "ten": ten
        }
        return render(request, "sms/ctdt-monhoc_list.html", context)

@login_required
def single_ctdtmonhoc(request, ctdt_id):
        ctdt = CtdtMonhoc.objects.filter(ctdt_id = ctdt_id).select_related("monhoc").order_by('hocky')
        ten = Ctdt.objects.get(id = ctdt_id).ten
        context = {
            "ctdt": ctdt,
            "ten": ten
        }
        return render(request, "sms/ctdt-monhoc_list.html", context)
@login_required
def lop_monhoc(request, lop_id):
        if not request.user.has_perm('assign_lop',Lop.objects.get(id=lop_id)):
            return redirect("lop_list")
        if request.user.is_gv:
            gv = Hsgv.objects.get(user_id = request.user.id)
            glm = GvLmh.objects.filter(gv_id = gv.id, status = 1)
            lm = LopMonhoc.objects.filter(lop_id = lop_id, id__in=[ns.lopmh_id for ns in glm]).select_related("lop", "monhoc").order_by('lop_id')
        else:    
            lm = LopMonhoc.objects.filter(lop_id = lop_id).select_related("lop", "monhoc")
        ten = Lop.objects.get(id = lop_id).ten
        context = {
            "lm": lm,
            "ten": ten,
            "lop_id": lop_id
        }
        return render(request, "sms/lop-monhoc_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def lop_monhoc_testwithGuardian(request, lop_id):
        if request.user.is_gv:
            list_of_ids = []
            for lm in LopMonhoc.objects.filter(lop_id = lop_id):
                if request.user.has_perm('assign_lopmonhoc', lm):
                    list_of_ids.append(lm.id)
            lm = LopMonhoc.objects.filter(id__in=list_of_ids).select_related("lop", "monhoc").order_by('lop_id')
            # gv = Hsgv.objects.get(user_id = request.user.id)
            # glm = GvLmh.objects.filter(gv_id = gv.id, status = 1)
            # lm = LopMonhoc.objects.filter(lop_id = lop_id, id__in=[ns.lopmh_id for ns in glm]).select_related("lop", "monhoc").order_by('lop_id')
        else:    
            lm = LopMonhoc.objects.filter(lop_id = lop_id).select_related("lop", "monhoc")
        ten = Lop.objects.get(id = lop_id).ten
        context = {
            "lm": lm,
            "ten": ten,
            "lop_id": lop_id
        }
        return render(request, "sms/lop-monhoc_list.html", context)

@login_required
def lop_monhoc_gv(request):
        gv = Hsgv.objects.get(user_id = request.user.id)
        glm = GvLmh.objects.filter(gv_id = gv.id)
        lm = LopMonhoc.objects.filter(id__in=[ns.lopmh_id for ns in glm]).select_related("lop", "monhoc").order_by('lop_id')
        context = {
            "lm": lm
        }
        return render(request, "sms/lop-monhoc_list.html", context)

@login_required
@permission_required('sms.view_hp81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def hv_hp81_list(request, sv_id, lop_id):
        hp81s = Hp81.objects.filter(sv_id = sv_id).select_related("sv", "hk")
        sv = Hssv.objects.get(id = sv_id)
        lop_id = sv.lop_id
        context = {
            "hp81s": hp81s,
            "sv": sv
        }
        return render(request, "sms/hv-hp81_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def hv_hp81_new_list(request, lop_id):
    #svs = Hssv.objects.filter(lop_id = lop_id)
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
    lop = Lop.objects.get(id = lop_id)

    for sv in svs:
        sv.hp81 = Hp81.objects.filter(sv_id = sv.id).select_related("sv", "hk")

    context = {
        "lop": lop,
        "svs": svs
    }
    return render(request, "sms/hv-hp81-new_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def hv_hp81_hk_list(request, lop_id):
    #svs = Hssv.objects.filter(lop_id = lop_id)
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    hks = Hocky.objects.all()
    #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
    lop = Lop.objects.get(id = lop_id)

    if request.method == "POST":
        hk_id = request.POST.get('hk', None)
        hk = Hocky.objects.get(id = hk_id)
    else:
        hk = hks[0]

    for sv in svs:
        if Hp81.objects.filter(sv = sv, hk =  hk).exists():
            sv.hp81 = Hp81.objects.filter(sv = sv, hk =  hk)[0]
    
    context = {
        "svs": svs,
        "hks": hks,
        "hk": hk,
        "lop": lop
    }

    return render(request, "sms/hv-hp81-hk_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def hv_hs81_list(request, sv_id, lop_id):
        hs81s = Hs81.objects.filter(sv_id = sv_id).select_related("sv", "hk")
        sv = Hssv.objects.get(id = sv_id)
        lop_id = sv.lop_id
        context = {
            "hs81s": hs81s,
            "sv": sv
        }
        return render(request, "sms/hv-hs81_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def hv_hs81_new_list(request, lop_id):
    #hs81s = Hs81.objects..select_related("sv", "hk")
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
    #svs = Hssv.objects.filter(lop_id = lop_id)
    lop = Lop.objects.get(id = lop_id)

    for sv in svs:
        print(sv.hoten)
        print(Hs81.objects.filter(sv_id = sv.id).select_related("sv", "hk").count())
        sv.hs81 = Hs81.objects.filter(sv_id = sv.id).select_related("sv", "hk")
    
    context = {
        "svs": svs,
        "lop": lop
    }
    return render(request, "sms/hv-hs81-new_list.html", context)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def hv_hs81_hk_list(request, lop_id):
    #hs81s = Hs81.objects..select_related("sv", "hk")
    locale.setlocale(locale.LC_ALL, 'vi_VN')
    hks = Hocky.objects.all()
    #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
    svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
    #svs = Hssv.objects.filter(lop_id = lop_id)
    lop = Lop.objects.get(id = lop_id)

    if request.method == "POST":
        hk_id = request.POST.get('hk', None)
        hk = Hocky.objects.get(id = hk_id)
    else:
        hk = hks[0]

    for sv in svs:
        if Hs81.objects.filter(sv = sv, hk =  hk).exists():
            sv.hs81 = Hs81.objects.filter(sv = sv, hk =  hk)[0]
    
    context = {
        "svs": svs,
        "hks": hks,
        "hk": hk,
        "lop": lop
    }
    return render(request, "sms/hv-hs81-hk_list.html", context)

@login_required
def single_hs81lop(request, lop_id):
        ctdtmonhs811 = Hs81.objects.filter(lop_id = lop_id, hk = 1).select_related("sv", "lop")
        ctdtmonhs812 = Hs81.objects.filter(lop_id = lop_id, hk = 2).select_related("sv", "lop")
        ctdtmonhs813 = Hs81.objects.filter(lop_id = lop_id, hk = 3).select_related("sv", "lop")
        ctdtmonhs814 = Hs81.objects.filter(lop_id = lop_id, hk = 4).select_related("sv", "lop")
        ten = Lop.objects.get(id = lop_id).ten
        context = {
            "ctdtmonhoc1": ctdtmonhs811,
            "ctdtmonhoc2": ctdtmonhs812,
            "ctdtmonhoc3": ctdtmonhs813,
            "ctdtmonhoc4": ctdtmonhs814,
            "ten": ten
        }
        return render(request, "sms/lop-hs81_list.html", context)
   
def single_teacher(request, teacher_id):
    single_teacher = get_object_or_404(TeacherInfo, pk=teacher_id)
    context = {
        "single_teacher": single_teacher
    }
    return render(request, "sms/single_teacher.html", context)

@login_required
def create_lichhoc(request):
    if request.method == "POST":
        forms = CreateLichhoc(request.POST, request.FILES or None)
        lop = request.POST.get('lop', None)
        if forms.is_valid():
            forms.save()
        messages.success(request, "Lich hoc duoc tao thanh cong!")
        return redirect("lichhoclop_list", lop)
    else:
        forms = CreateLichhoc()
    context = {
        "forms": forms
    }
    return render(request, "sms/create_lichhoc.html", context)

@login_required
@permission_required('sms.add_hp81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def create_hp81(request, lop_id, sv_id):
    sv = Hssv.objects.get(id = sv_id) 
    if request.method == "POST":
        hk_id = request.POST.get('hk', None)
        forms = CreateHp81(request.POST, request.FILES or None)
        if Hp81.objects.filter(hk_id = hk_id, sv_id=sv_id).first():
            #messages.success(request, "Mon hoc da ton taij!")
            messages.error(request, "Bản ghi đã tồn tại!")
        else:
            lop = request.POST.get('lop', None)
            if forms.is_valid():
                forms.save()
                messages.success(request, "Bản ghi duoc tao thanh cong!")
                return redirect("hv_hp81_hk_list", sv.lop_id)
    else:
        forms = CreateHp81()
    context = {
        "sv": sv,
        "forms": forms
    }
    return render(request, "sms/create_hp81.html", context)

@login_required
@permission_required('sms.add_hs81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def create_hs81(request, lop_id, sv_id):
    sv = Hssv.objects.get(id = sv_id) 
    if request.method == "POST":
        hk_id = request.POST.get('hk', None)
        forms = CreateHs81(request.POST, request.FILES or None)
        if Hs81.objects.filter(hk_id = hk_id, sv_id=sv_id).first():
            #messages.success(request, "Mon hoc da ton taij!")
            messages.error(request, "Bản ghi đã tồn tại!")
        else:
            lop = request.POST.get('lop', None)
            if forms.is_valid():
                forms.save()
                messages.success(request, "Bản ghi duoc tao thanh cong!")
                return redirect("hv_hs81_hk_list", sv.lop_id)
    else:
        forms = CreateHs81()
    context = {
        "sv": sv,
        "forms": forms
    }
    return render(request, "sms/create_hs81.html", context)

@login_required
def create_lichhoclm(request, lopmh_id):
    if request.method == "POST":
        forms = CreateLichhoc(request.POST, request.FILES or None)
        lop = request.POST.get('lop', None)
        if forms.is_valid():
            forms.save()
        messages.success(request, "Lich hoc duoc tao thanh cong!")
        return redirect("lichhoclopmh_list", lopmh_id)
    else:
        forms = CreateLichhoc()

    lmh = LopMonhoc.objects.get(id = lopmh_id)
    forms = CreateLichhoc()
    context = {
        "forms": forms,
        "lmh": lmh
    }
    return render(request, "sms/create_lichhoclm.html", context)

@login_required
def create_ctdtmonhoc(request):
    if request.method == "POST":
        forms = CreateCtdtMonhoc(request.POST, request.FILES or None)
        ctdt = request.POST.get('ctdt', None)
        if forms.is_valid():
            forms.save()
        messages.success(request, "Mon hoc duoc gan cho CTDT thanh cong!")
        return redirect("single_ctdtmonhoc", ctdt)
    else:
        forms = CreateCtdtMonhoc()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_ctdtmonhoc.html", context)

@login_required
#@permission_required('sms.add_lopmonhoc', raise_exception=True)
def create_lopmonhoc(request, lop_id):
#    if not request.user.has_perm('sms.add_lopmonhoc'):
#        return HttpResponseForbidden("Bạn không có quyền thực hiện thao tác thêm môn học!")
    if request.method == "POST":
        monhoc_id = request.POST.get('monhoc', None)
        if LopMonhoc.objects.filter(monhoc_id = monhoc_id, lop_id=lop_id).first():
            #messages.success(request, "Mon hoc da ton taij!")
            messages.error(request, "Môn học đã tồn tại!")
        else:
            forms = CreateLopMonhoc(request.POST, request.FILES or None)
            if forms.is_valid():
                forms.save()
                messages.success(request, "Môn học đã được thêm thành công!")
        #messages.success(request, "Mon hoc duoc them thanh cong!")
        return redirect("lop_monhoc", lop_id)
    forms = CreateLopMonhoc()

    query = '''select mh.id, mh.ma, mh.ten 
                            from sms_monhoc mh
                            inner join sms_ctdtmonhoc ctdtmh
								on mh.id = ctdtmh.monhoc_id and ctdtmh.status = 1
                            inner join sms_ctdt ctdt
								on ctdt.id = ctdtmh.ctdt_id
                            inner join sms_lop l
								on l.ctdt_id = ctdt.id
                            where l.id = %s and mh.id not in (select monhoc_id from sms_lopmonhoc where lop_id = %s)    
                            order by mh.id'''

    #sv = DiemdanhAll.objects.raw(query, [ttlh.lop_id, lh_id])
    #lmhs = LopMonhoc.objects.filter(lop_id = lop_id)
    mhs = Monhoc.objects.raw(query, [lop_id, lop_id])
    context = {
        "forms": forms,
        "mhs": mhs,
        "lop_id": lop_id
    }
    return render(request, "sms/create_lopmonhoc.html", context)

@login_required
def create_ctdt(request):
    if request.method == "POST":
        forms = CreateCtdt(request.POST, request.FILES or None)
        if Ctdt.objects.filter(ten = forms['ten'].value()).first():
            messages.error(request, "CTĐT đã tồn tại!")
            return redirect("ctdt_list")
        if forms.is_valid():
            forms.save()
            messages.success(request, "Tao moi CTDT thanh cong!")
            return redirect("ctdt_list")
    else:
        forms = CreateCtdt()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_ctdt.html", context)


@login_required
def create_lop(request):
    if request.method == "POST":
        forms = CreateLop(request.POST, request.FILES or None)
        if Lop.objects.filter(ma = forms['ma'].value()).first():
            messages.error(request, "Mã lớp đã tồn tại!")
            return redirect("lop_list")
        if forms.is_valid():
            lop = forms.save()
            #create Group and assign permission to Lop
            # group = Group.objects.create(name="Lop_permisison_"+str(lop.id))
            # assign_perm('assign_lop', group, lop)
            hks= Hocky.objects.all()
            for hk in hks:
                lhk = LopHk.objects.create(hk_id = hk.id, lop_id = lop.id)
                lhk.save()        
            messages.success(request, "Tạo mới lớp thành công!")
        return redirect("lop_list")
    else:
        forms = CreateLop()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_lop.html", context)

@login_required
def create_loc(request):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền tạo địa điểm.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if request.method == "POST":
        forms = CreateLoc(request.POST, request.FILES or None)
        if forms.is_valid():
            loc = forms.save()
            loc.chu_id = request.user.id
            loc.save()
            messages.success(request, "Tạo mới địa chỉ thành công!")
        else:
            for error in forms.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))
        return redirect("loc_list")
    else:
        forms = CreateLoc()

    context = {
        "forms": forms,
        "chu_id": request.user.id
    }
    return render(request, "sms/create_loc.html", context)

@login_required
def create_house(request,loc_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền tạo nhà/phòng!!!")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if request.method == "POST":
        forms = CreateHouse(request.POST, request.FILES or None)

        if forms.is_valid():
            # Đảm bảo Location thuộc chủ hiện tại
            posted_loc_id = request.POST.get('loc', None) or loc_id
            if not request.user.is_superuser:
                get_object_or_404(Location, pk=posted_loc_id, chu=request.user)
            house = forms.save(commit=False)
            house.loc_id = posted_loc_id
            
            house.save()
            messages.success(request, "Tạo mới nhà trọ thành công!")
        else:
            messages.error(request, "Lỗi Forms: " + str(forms.errors.as_text()))    
            # print(error)
            # messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

        return redirect("view_loc_secure", loc_id)
    else:
        forms = CreateHouse()

    context = {
        "forms": forms,
        "loc_id": loc_id
    }
    return render(request, "sms/create_house.html", context)

@login_required
#@require_http_methods(['POST'])
def create_xlop(request):
    if request.method == "POST":
        forms = CreateLop(request.POST, request.FILES or None)
        if Lop.objects.filter(ma = forms['ma'].value()).first():
            messages.error(request, "Mã lớp đã tồn tại!")
            return redirect("lop_list")
        if forms.is_valid():
            lop = forms.save()
            #create Group and assign permission to Lop
            # group = Group.objects.create(name="Lop_permisison_"+str(lop.id))
            # assign_perm('assign_lop', group, lop)
            hks= Hocky.objects.all()
            for hk in hks:
                lhk = LopHk.objects.create(hk_id = hk.id, lop_id = lop.id)
                lhk.save()        
            messages.success(request, "Tạo mới lớp thành công!")
        return redirect("xlop_list")
    else:
        forms = CreateLop()

    context = {
        "forms": forms
    }
    return render(request, "includes/create_xlop.html", context)

@login_required
def create_sv(request):
    if request.method == "POST":
        forms = CreateSv(request.POST, request.FILES or None)
        #lh = request.POST.get('lichhoc', None)
        if Hssv.objects.filter(msv = forms['msv'].value()).first():
            messages.error(request, "Mã học viên đã tồn tại!")
        else:
            #return redirect("sv_list")
            if forms.is_valid():
                #uploaded_img = forms.save(commit=False)
                #uploaded_img.image_data = forms.cleaned_data['image'].file.read()
                #uploaded_img.save()
                forms.save()
                messages.success(request, "Tạo mới học viên thành công!")
                if forms.instance.lop_id:
                    return redirect("svlop_list", forms.instance.lop_id)
                else:
                    return redirect("sv_list")
    else:
        forms = CreateSv()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_sv.html", context)


@login_required
def create_gv(request):
    if request.method == "POST":
        forms = CreateGv(request.POST, request.FILES or None)
        if Hsgv.objects.filter(ma = forms['ma'].value()).first():
            messages.error(request, "Mã giáo viên đã tồn tại!")
            return redirect("gv_list")
        if forms.is_valid():
            forms.save()
        messages.success(request, "Tạo mới giáo viên thành công!")
        return redirect("gv_list")
    else:
        forms = CreateGv()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_gv.html", context)

@login_required
def create_ns(request):
    if request.method == "POST":
        forms = CreateNs(request.POST, request.FILES or None)
        if Hsns.objects.filter(ma = forms['ma'].value().strip()).first():
            messages.error(request, "Mã nhân sự đã tồn tại!")
            return redirect("ns_list")
        if Hsns.objects.filter(email = forms['email'].value().strip()).first():
            messages.error(request, "Email nhân sự đã tồn tại!")
            return redirect("ns_list")
        if forms.is_valid():
            forms.save()
        messages.success(request, "Tạo mới nhân sự thành công!")
        return redirect("ns_list")
    else:
        forms = CreateNs()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_ns.html", context)

@login_required
def create_renter(request):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền tạo khách thuê.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if request.method == "POST":
        forms = CreateRenter(request.POST, request.FILES or None)
        # if Renter.objects.filter(ma = forms['ma'].value().strip()).first():
        #     messages.error(request, "Mã người thuê đã tồn tại!")
        #     return redirect("renter_list")
        # if Renter.objects.filter(email = forms['email'].value().strip()).first():
        #     messages.error(request, "Email người thuê đã tồn tại!")
        #     return redirect("renter_list")
        if forms.is_valid():
            # Save renter but ensure optional fields default safely to avoid DB NOT NULL issues
            renter = forms.save(commit=False)
            # Defensive defaults: if form omits optional fields, set to empty string instead of None
            renter.sdt = renter.sdt or ""
            renter.cccd = renter.cccd or ""
            renter.mst = renter.mst or ""
            renter.ghichu = renter.ghichu or ""
            renter.chu_id = request.user.id
            renter.save()
            # Generate renter code after we have an ID
            renter.ma = "NT"+str(request.user.id)+str(renter.id).zfill(3)
            renter.save()
            messages.success(request, "Tạo mới người thuê thành công!")
            return redirect("renter_list")

        else:
            messages.error(request, "Lỗi Forms: " + str(forms.errors.as_text()))

    else:
        forms = CreateRenter()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_renter.html", context)

@login_required
def create_diemdanh(request, lh_id):
    print('hello')
    if request.method == "POST":
        forms = CreateDiemdanh(request.POST, request.FILES or None)
        #forms.initial['lichhoc']=7
        lh = request.POST.get('lichhoc', None)
        print('hello')
        if forms.is_valid():
            forms.save()
        messages.success(request, "Them sinh vien vang mat thanh cong!")
        return redirect("diemdanh_list", lh)
    else:
        forms = CreateDiemdanh()
    ttlh = Lichhoc.objects.all().select_related("lop", "monhoc").get(id = lh_id)
    #sv = Hssv.objects.all().filter(lop = ttlh.lop.ten)
    ds = Hssv.objects.all().filter(lop = ttlh.lop.ten)
    context = {
        "forms": forms,
        "lh_id": lh_id,
        "ttlh": ttlh,
        "dssv": ds
    }
    return render(request, "sms/create_diemdanh-new.html", context)

@login_required
def create_hocphi(request, lh_id):
    print('hello')
    if request.method == "POST":
        forms = CreateHocphi(request.POST, request.FILES or None)
        #forms.initial['lichhoc']=7
        lh = request.POST.get('lichhoc', None)
        #print('hello')
        if forms.is_valid():
            forms.save()
        messages.success(request, "Them thong tin hoc phi thanh cong!")
        return redirect("hocphi_list", lh_id)
    else:
        forms = CreateHocphi()
    ttlh = Lop.objects.get(id = lh_id)
    #sv = Hssv.objects.all().filter(lop = ttlh.lop.ten)
    ds = Hssv.objects.all().filter(lop = ttlh.ten)
    context = {
        "forms": forms,
        "lh_id": lh_id,
        "ttlh": ttlh,
        "dssv": ds
    }
    return render(request, "sms/create_hocphi.html", context)


@login_required
def create_diem(request, lh_id):
    print('hello')
    if request.method == "POST":
        forms = CreateDiem(request.POST, request.FILES or None)
        #forms.initial['lichhoc']=7
        #lh = request.POST.get('lichhoc', None)
        #print('hello')
        if forms.is_valid():
            forms.save()
        messages.success(request, "Them thong tin diem thanh cong!")
        return redirect("diem-lop_list", lh_id)
    else:
        forms = CreateDiem()
    ttlh = Lop.objects.get(id = lh_id)
    #sv = Hssv.objects.all().filter(lop = ttlh.lop.ten)
    ds = Hssv.objects.all().filter(lop = ttlh.ten)
    context = {
        "forms": forms,
        "lh_id": lh_id,
        "ttlh": ttlh,
        "dssv": ds
    }
    return render(request, "sms/create_diemtp.html", context)

@login_required
def create_diemdanh1(request):
    if request.method == "POST":
        forms = CreateDiemdanh(request.POST, request.FILES or None)
        lh = request.POST.get('lichhoc', None)
        if forms.is_valid():
            forms.save()
        messages.success(request, "Them sinh vien vang mat thanh cong!")
        return redirect("diemdanh_list", lh)
    else:
        forms = CreateDiemdanh()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_diemdanh.html", context)

def create_teacher(request):
    if request.method == "POST":
        forms = CreateTeacher(request.POST, request.FILES or None)

        if forms.is_valid():
            forms.save()
        messages.success(request, "Teacher Registration Successfully!")
        return redirect("teacher_list")
    else:
        forms = CreateTeacher()

    context = {
        "forms": forms
    }
    return render(request, "sms/create_teacher.html", context)

@login_required
def edit_lopmonhoc(request, lmh_id):
    lmh = LopMonhoc.objects.filter(id=lmh_id).select_related("lop", "monhoc")[0]
    lop_id = lmh.lop_id
    mh = Monhoc.objects.get(id=lmh.monhoc_id)

    lmh_forms = CreateLopMonhoc(instance=lmh)

    if request.method == "POST":
        edit_forms = CreateLopMonhoc(request.POST, request.FILES or None, instance=lmh)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit Mon hoc Info Successfully!")
            return redirect("lop_monhoc", lop_id)

    context = {
        "forms": lmh_forms,
        "lmh": lmh
    }
    return render(request, "sms/edit_lopmonhoc.html", context)

@login_required
def edit_loc(request, loc_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền sửa địa điểm.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Validate location access
    loc, error_response = validate_location_access(request, loc_id)
    if error_response:
        return error_response
    
    # Reload with select_related for optimization
    loc = Location.objects.select_related("xp").get(id=loc_id)
    loc_forms = CreateLoc(instance=loc)

    if request.method == "POST":
        edit_forms = CreateLoc(request.POST, request.FILES or None, instance=loc)
        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Cập nhật thông tin địa điểm thành công!")
            return redirect("loc_list")

    context = {
        "forms": loc_forms,
        "loc": loc
    }
    return render(request, "sms/edit_loc.html", context)

@login_required
def edit_house(request, loc_id, house_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền sửa nhà/phòng.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Validate location access first
    loc, error_response = validate_location_access(request, loc_id)
    if error_response:
        return error_response
    
    # Validate house access and ensure it belongs to the location
    house, error_response = validate_house_access(request, house_id)
    if error_response:
        return error_response
    
    # Ensure house belongs to the specified location
    if house.loc_id != loc_id:
        messages.error(request, "Nhà này không thuộc vị trí được chỉ định")
        return redirect('view_loc_secure', loc_id)
    
    #house_forms = CreateHouse(instance=house)

    if request.method == "POST":
        edit_forms = CreateHouse(request.POST, request.FILES or None, instance=house)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit House Info Successfully!")
            return redirect("view_loc_secure", loc_id)
        else:
            messages.error(request, "Lỗi Forms: " + str(edit_forms.errors.as_text()))

# --- GET REQUEST (HOẶC KHI FORM KHÔNG VALID) ---
    # ĐỊNH DẠNG SỐ TIỀN TRƯỚC KHI HIỂN THỊ
    
    # 1. Định dạng permonth: Chuyển số nguyên thành chuỗi có dấu chấm
    # Python định dạng hàng ngàn bằng dấu phẩy, sau đó chúng ta thay thế phẩy thành chấm (chuẩn VN).
    formatted_permonth = f"{house.permonth:,}"
    
    # 2. Định dạng deposit
    formatted_deposit = f"{house.deposit:,}"

    # 3. Khởi tạo Form với dữ liệu đã định dạng (Initial Data)
    # Tham số 'initial' ghi đè giá trị mặc định của trường (permonth và deposit) bằng chuỗi đã định dạng.
    house_forms = CreateHouse(
        instance=house, 
        initial={
            'permonth': formatted_permonth,
            'deposit': formatted_deposit,
        }
    )

    context = {
        "forms": house_forms,
        "loc_id": loc_id,
        "house": house
    }
    return render(request, "sms/edit_house.html", context)

# Nhân bản house
def save_as_house(request, house_id):
    house = get_object_or_404(House, id=house_id)
    # Tạo bản sao, chỉ khác tên
    new_house = House.objects.get(id=house_id)
    new_house.pk = None
    new_house.ten = f"{house.ten}_1"
    new_house.save()
    messages.success(request, f"Đã nhân bản nhà trọ thành công: {new_house.ten}")
    return redirect("view_loc_secure", new_house.loc_id)

@login_required
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
@permission_required('sms.change_ttgv',raise_exception=True)
def edit_ttgv(request, lop_id, lopmh_id, gv_id):
    if Ttgv.objects.filter(lopmh_id=lopmh_id, gv_id=gv_id).exists():
        ttgv = Ttgv.objects.get(lopmh_id=lopmh_id, gv_id=gv_id)
    else:
        lhs = Lichhoc.objects.filter(lmh_id = lopmh_id, giaovien_id = gv_id)
        sotiet = lhs.aggregate(Sum('sotiet'))['sotiet__sum']
        ttgv = Ttgv(lopmh_id=lopmh_id, gv_id=gv_id,sotiet = sotiet)
        ttgv.save()

    lmh = LopMonhoc.objects.get(id=lopmh_id)
    gv=Hsgv.objects.get(id=gv_id)
    mh=Monhoc.objects.get(id=lmh.monhoc_id)

    ttgv_forms = CreateTtgv(instance=ttgv)

    if request.method == "POST":
        edit_forms = CreateTtgv(request.POST, request.FILES or None, instance=ttgv)

        if edit_forms.is_valid():
            edit_forms.save()
            #send notification to gv
            if gv.user:
                notify.send(sender=gv.user, recipient= gv.user, verb='Thông tin thanh toán được cập nhật trên hệ thống', level='info')

            messages.success(request, "Cập nhật thông tin thanh toán thành công!")
            return redirect("gv-lmh-lst", lopmh_id)

    context = {
        "forms": ttgv_forms,
        "lmh": lmh,
        "mh": mh,
        "gv": gv,
        "ttgv": ttgv
    }
    return render(request, "sms/edit_ttgv.html", context)

@login_required
def history_lopmonhoc(request, lmh_id):

    lmh = LopMonhoc.objects.select_related("lop", "monhoc").get(id=lmh_id)
    
    #poll = Poll.objects.get(pk=poll_id)
    p = lmh.history.all()
    changes = []
    if p is not None and lmh_id:
        last = p.first()
        for all_changes in range(p.count()):
            new_record, old_record = last, last.prev_record
            if old_record is not None:
                delta = new_record.diff_against(old_record)
                changes.append(delta)
            last = old_record
            # create first time
            if all_changes == p.count()-1:
                delta = new_record.diff_against(new_record)
                changes.append(delta)

    context = {
        "lmh": lmh,
        "changes": changes
    }
    return render(request, "sms/history_lopmonhoc.html", context)

@login_required
def edit_lichhoc(request, lh_id):
    lh = Lichhoc.objects.get(id=lh_id)
    lmh = LopMonhoc.objects.filter(id=lh.lmh_id).select_related("lop", "monhoc")[0]
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lh_forms = CreateLichhoc(instance=lh)

    if request.method == "POST":
        edit_forms = CreateLichhoc(request.POST, request.FILES or None, instance=lh)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit Lịch học Info Successfully!")
            return redirect("lichhoclopmh_list", lmh.id)

    context = {
        "forms": lh_forms,
        "lmh": lmh,
        "lh": lh
    }
    return render(request, "sms/edit_lichhoc.html", context)
@login_required
def edit_sv(request, sv_id):
    sv = Hssv.objects.get(id=sv_id)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lh_forms = CreateSv(instance=sv)
    #base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    #if sv.image:
    #    mroot = os.path.join(base_dir, 'media',sv.image.name)
    #    troot = os.path.join(base_dir, 'static')
        #print(MEDIA_ROOT)
        #print(STATIC_ROOT)
    #    shutil.copy(mroot, troot + "\\uploads\\" + sv.image.name[8:]) # file will be renamed

    if request.method == "POST":
        edit_forms = CreateSv(request.POST, request.FILES or None, instance=sv)
        #lop_id = request.POST.get('lop', None)
        #uploaded_file = UploadedFile.objects.get(pk=file_id)
        #response = HttpResponse(sv.image, content_type='application/force-download')
        #response['Content-Disposition'] = f'attachment; filename="{sv.image.name}"'
        #return response

        if edit_forms.is_valid():
            #uploaded_img = edit_forms.save(commit=False)
            #uploaded_img.image_data = edit_forms.cleaned_data['image'].file.read()
            sv = edit_forms.save()
            messages.success(request, "Edit Học viên Info Successfully!")
            if sv.lop:
                return redirect("svlop_list", sv.lop.id)
            return redirect("sv_list")
        else:
            for error in edit_forms.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))
                # for data in error:
                #     print(data)

    context = {
        "forms": lh_forms,
#        "img": sv.image,
        "sv": sv
    }
    return render(request, "sms/edit_sv.html", context)

@login_required
def details_sv(request, sv_id, opt = None):
    
    sv = Hssv.objects.get(id=sv_id)

    lhk = LopHk.objects.filter(lop_id = sv.lop_id).select_related('hk').order_by('hk_id')
    next_hk=0

    for l in lhk:
        if l.end_hk and date.today() > l.end_hk:
            next_hk = l.hk.ma
            #break
    if next_hk == 0:
        next_hk = lhk.last().hk.ma + 1
    else:
        next_hk = next_hk + 1


    lds = Loaidiem.objects.all()
    hks = Hocky.objects.filter(ma__lt = next_hk).order_by('ma')

    # for ar in args:

    hps = Hp81.objects.filter(sv_id = sv_id)
    mhl=[]
    ldl=[]
    dtpl=[]



    # for mh in lmhs:
    #     ldl=[]
    #     for l in ld:
    #         dtpl=[]
    #         for dtp in Diemthanhphan.objects.filter(sv_id = sv_id, monhoc_id = mh.monhoc_id, tp_id = l.id, status=1).order_by('log_id'):
    #             dtpl.append({"id":dtp.log_id,"mark":dtp.diem})

    #         ldl.append({"ma":l.ma,"dtplst": dtpl})

    #     mhl.append({ "ma":mh.monhoc.ten,"ttdiem0": ldl[0], "ttdiem": ldl})

    tctl, diem4 = 0,0
    for hk in hks:
        lml=[]
        tbmhk, tbmhk4, tchk = 0,0,0
        dtks = DiemTk.objects.filter(sv_id = sv.id, hk_id = hk.id)
        for dtk in dtks:
                        
            if not dtk.mhdk:    
                tbmhk= tbmhk+ dtk.tbm*dtk.tc
                tbmhk4= tbmhk4+ dtk.tbm4*dtk.tc
                tchk=tchk+dtk.tc

        hk.lml = dtks
        hk.tchk = tchk
        hk.tbmhk = round(tbmhk/tchk,1) if tchk else 0
        hk.tbmhk4 = round(tbmhk4/tchk,1) if tchk else 0

        tctl = tctl + tchk
        diem4 = diem4 + hk.tbmhk4*tchk
        hk.tbctl = round(diem4/tctl,1) if tctl else 0

        if hk.tbctl >=3.5 and hk.tbctl <=4:
            hk.xl = "Xuất xắc"
        elif hk.tbctl >=3 and hk.tbctl <3.5:
            hk.xl = "Giỏi"
        elif hk.tbctl >=2.5 and hk.tbctl <3:
            hk.xl = "Khá"
        elif hk.tbctl >=2 and hk.tbctl <2.5:
            hk.xl = "Trung bình"
        elif hk.tbctl <2:
            hk.xl = "Yếu"


    #export to excel
    if opt == 3:
        import pandas as pd
        #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)
        exp=[]
        exp.append({"Học viên": sv.hoten
                    })
        for hk in hks:
            exp.append({"Học kỳ|Môn học": hk.ten, 
                        "Tín chỉ": hk.tchk, 
                        "kttx1": "",
                        "kttx2": "",
                        "kttx3": "",
                        "ktdk1": "",
                        "ktdk2": "",
                        "ktdk3": "",
                        "TBM KT": "",
                        "ktkt1": "",
                        "ktkt2": "",
                        "TBM 10": hk.tbmhk,
                        "TBM CHỮ": "",
                        "TBM 4": hk.tbmhk4,
                        "TBM TL": hk.tbctl,
                        "Xếp loại": hk.xl
                    })
            for mh in hk.lml:
                exp.append({"Học kỳ|Môn học": mh.monhoc, 
                            "Tín chỉ": mh.tc, 
                            "kttx1": mh.kttx1 if mh.n_kttx1 else "",
                            "kttx2": mh.kttx2 if mh.n_kttx2 else "",
                            "kttx3": mh.kttx3 if mh.n_kttx3 else "",
                            "ktdk1": mh.ktdk1 if mh.n_ktdk1 else "",
                            "ktdk2": mh.ktdk2 if mh.n_ktdk2 else "",
                            "ktdk3": mh.ktdk3 if mh.n_ktdk3 else "",
                            "TBM KT": mh.tbmkt, 
                            "ktkt1": mh.ktkt1 if mh.n_ktkt1 else "",
                            "ktkt2": mh.ktkt2 if mh.n_ktkt2 else "",
                            "TBM 10": mh.tbm,
                            "TBM CHỮ": mh.tbmc,
                            "TBM 4": mh.tbm4
                        })

        # Convert the QuerySet to a DataFrame
        df = pd.DataFrame(list(exp))

        tenf = "kqht_"+ sv.msv + ".xlsx"
        # Define the Excel file response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=kqht-lop.xlsx'
        response['Content-Disposition'] = 'attachment; filename=' + tenf


        # Use Pandas to write the DataFrame to an Excel file
        df.to_excel(response, index=False, engine='openpyxl')

        return response

    #export to excel template
    if opt == 5:
        import pandas as pd
        # import pythoncom
        # import win32com.client
        # pythoncom.CoInitialize()
        import random
        import string


        temp_path = "template_kqht.xlsx"
        out_path = ''.join(random.choices(string.ascii_lowercase, k=5)) + "sv_" + sv.msv + "_kqht.xlsx"
        pdf_path = "sv_kqht.pdf"
        temp_file_path = os.path.join(settings.MEDIA_ROOT, temp_path)
        out_file_path = os.path.join(settings.MEDIA_ROOT, out_path)
        pdf_file_path = os.path.join(settings.MEDIA_ROOT, pdf_path)
        shutil.copy(temp_file_path, out_file_path)
        #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)
        exptt=[]
        exptt.append({"Học viên": sv.hoten, "Mã": sv.msv, "Lớp": sv.lop.ten})
        # use `with` to avoid other exceptions
        dtftt = pd.DataFrame(list(exptt))
        with pd.ExcelWriter(out_file_path, mode="a",engine="openpyxl", if_sheet_exists="overlay",) as writer:
            #writer.book = template

            dtftt.to_excel(writer, sheet_name='hks', index=False, header=False, startrow=1, startcol=0)

        for hk in hks:
            exp=[]
            expxl=[]
            for mh in hk.lml:
                exp.append({"Học kỳ|Môn học": mh['ten'], 
                            "kttx1": mh['kttx1'],
                            "kttx2": mh['kttx2'],
                            "kttx3": mh['kttx3'],
                            "ktdk1": mh['ktdk1'],
                            "ktdk2": mh['ktdk2'],
                            "ktdk3": mh['ktdk3'],
                            "TBM KT": mh['tbmkt'], 
                            "ktkt1": mh['ktkt1'],
                            "ktkt2": mh['ktkt2'],
                            "TBM 10": mh['tbm']
                        })
            # Convert the QuerySet to a DataFrame
            expxl.append({"XL": hk.xl})
            dtf = pd.DataFrame(list(exp))
            dtfxl = pd.DataFrame(list(expxl))

            # use `with` to avoid other exceptions
            with pd.ExcelWriter(out_file_path, mode="a",engine="openpyxl", if_sheet_exists="overlay",) as writer:
                #writer.book = template

                dtf.to_excel(writer, sheet_name='hks', index=False, header=False, startrow=9+(hk.ma-1)*27, startcol=0)
                dtfxl.to_excel(writer, sheet_name='hks', index=False, header=False, startrow=24+(hk.ma-1)*27, startcol=1)

        #if os.path.exists(out_file_path):
        with open(out_file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(out_file_path)
        
        os.remove(out_file_path)

        return response
        #raise Http404
        
    #export to pdf template
    if opt == 4:
        import random
        import string
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak,Spacer
        from reportlab.graphics.shapes import Line, LineShape, Drawing

        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER, TA_JUSTIFY
        from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
        from reportlab.platypus import Paragraph
        import io


        from reportlab.pdfbase.pdfmetrics import registerFont
        from reportlab.pdfbase.ttfonts import TTFont
        registerFont(TTFont('Arial','ARIAL.ttf'))


        temp_path = "template_kqht.xlsx"
        out_path = ''.join(random.choices(string.ascii_lowercase, k=5)) + "sv_" + sv.msv + "_kqht.xlsx"
        pdf_path = "sv_kqht.pdf"
        temp_file_path = os.path.join(settings.MEDIA_ROOT, temp_path)
        out_file_path = os.path.join(settings.MEDIA_ROOT, out_path)
        pdf_file_path = os.path.join(settings.MEDIA_ROOT, pdf_path)
        shutil.copy(temp_file_path, out_file_path)
        #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)

        out_pdf = ''.join(random.choices(string.ascii_lowercase, k=5)) + "sv_" + sv.msv + "_kqht.pdf"
        pdf_file_path = os.path.join(settings.MEDIA_ROOT, out_pdf)
        doc = SimpleDocTemplate(pdf_file_path, pagesize=landscape(A4))
        story = []
        buffer = io.BytesIO()

        for hk in hks:
            title = "Học kỳ: " + str(hk.ma)
            data=[]
            data = [
                [title, '', '', '', '', '', '', '', '', '', ''],
                ['Môn học/Mô-đun', 'Kết quả học tập Môn học/Mô đun', '', '', '', '', '', '', '', '', ''],
                ['', 'Điểm kiểm tra thường xuyên', '','','Điểm kiểm tra định kỳ', '', '', 'Điểm TB\n điểm\n KT', 'Điểm kiểm tra\n hết MH/MĐ', '', 'Điểm\n tổng\n kết'],
                ['', '', '', '', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', '', 'Lần 1', 'Lần 2', ''],
                ]


            for mh in hk.lml:
                data.append([mh.monhoc,
                            mh.kttx1 if mh.n_kttx1 else "",
                            mh.kttx2 if mh.n_kttx2 else "",
                            mh.kttx3 if mh.n_kttx3 else "",
                            mh.ktdk1 if mh.n_ktdk1 else "",
                            mh.ktdk2 if mh.n_ktdk2 else "",
                            mh.ktdk3 if mh.n_ktdk3 else "",
                            mh.tbmkt,
                            mh.ktkt1 if mh.n_ktkt1 else "",
                            mh.ktkt2 if mh.n_ktkt2 else "",
                            mh.tbm
                ])
                
            ptext = "<font name='%s'><b>%s</b></font>" % ("Arial", "Xếp loại học tập:")
            titlesTable = Paragraph(ptext)
            data.append([titlesTable, hk.xl, '', '', '', '', '', '', '', '', ''])

            tblstyle = TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                                ('FONTNAME', (0,0), (-1,-1), 'Arial'),
                                ('SPAN', (0,0), (-1,0)), # dòng 1
                                ('SPAN', (1,1), (-1,1)), # dòng 2
                                ('SPAN', (0,1), (0,4)), # cột 1, dòng 5
                                ('SPAN', (1,2), (3,4)), # cột 2, dòng 3 - cột 4, dòng 5
                                ('SPAN', (4,2), (6,4)), # cột 5, dòng 3 - cột 5, dòng 5
                                ('SPAN', (7,2), (7,4)), # cột 5, dòng 3 - cột 5, dòng 5
                                ('SPAN', (8,2), (9,3)), # cột 5, dòng 3 - cột 5, dòng 5
                                ('SPAN', (10,2), (10,4)), # cột 5, dòng 3 - cột 5, dòng 5

                                ('SPAN', (1,-1), (10,-1)), # cột 5, dòng 3 - cột 5, dòng 5

                                ('VALIGN', (0, 1), (0, 1), 'MIDDLE'),  # second column
                                ('ALIGN', (0,1), (0, 1), 'CENTER'),  # second column

                                ('ALIGN', (0, 0), (0, 0), 'LEFT'),   # first column

                                ('ALIGN', (0, 3), (0, -1), 'LEFT'),   # first column
                                ('ALIGN', (1, 1), (1, 1), 'CENTER'),   # first column

                                ('VALIGN', (1, 2), (10, 4), 'MIDDLE'),  # second column
                                ('ALIGN', (1, 2), (10, 4), 'CENTER'),  # second column
                                ])

            tbl = Table(data, colWidths=[250, 45, 45],               
                        )
            tbl.setStyle(tblstyle)

            psHeaderText = ParagraphStyle('Arial', fontSize=12, alignment=TA_LEFT, borderWidth=3)
            #text = 'OTRAS ACTIVIDADES Y DOCUMENTACIÓN'
            text = "Học viên: " + sv.hoten + " - Mã: " + sv.msv + " - Lớp: " + sv.lop.ten
            paragraphReportHeader = Paragraph('<font name="Arial">' + text + '</font>', psHeaderText)
            story.append(paragraphReportHeader)

            spacer = Spacer(5, 5)
            story.append(spacer)

            d = Drawing(500, 1)
            line = Line(-15, 0, 705, 0)
            line.strokeWidth = 1
            d.add(line)
            story.append(d)

            spacer = Spacer(10, 1)
            story.append(spacer)

            d = Drawing(500, 1)
            line = Line(-15, 0, 705, 0)
            line.strokeWidth = 0.5
            d.add(line)
            story.append(d)

            spacer = Spacer(10, 22)
            story.append(spacer)
            story.append(tbl)
            story.append(PageBreak())

        doc.build(story)

        # response = HttpResponse(pdf_file_path, content_type='application/force-download')
        # response['Content-Disposition'] = f'attachment; filename="{os.path.basename(pdf_file_path)}"'
        # return response

        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = f'attachment; filename="{os.path.basename(pdf_file_path)}"'
        # response.write(buffer.getvalue())
        # buffer.close()

        # return response
        
        #file_path = os.path.join(settings.MEDIA_ROOT, path)
        if os.path.exists(pdf_file_path):
            with open(pdf_file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/pdf")
                response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(pdf_file_path)

            os.remove(pdf_file_path)
            return response

        response = FileResponse(open(pdf_file_path, 'rb'), content_type='application/pdf')
        #os.remove(pdf_file_path)
        #os.remove(pdf_file_path)
        return response



    
    for hk in hks:
        hk.hs81 = Hs81.objects.get(sv_id = sv_id, hk_id = hk.id) if Hs81.objects.filter(sv_id = sv_id, hk_id = hk.id).exists() else None
        hk.hp81 = Hp81.objects.get(sv_id = sv_id, hk_id = hk.id) if Hp81.objects.filter(sv_id = sv_id, hk_id = hk.id).exists() else None

    if SvTn.objects.filter(sv_id = sv_id).exists():
        svtn = SvTn.objects.get(sv_id = sv_id)
        edit_svtn = 1
        forms = CreateSvTn(instance=svtn)
    else:
        edit_svtn = 0
        forms = CreateSvTn()    

    if request.method == "POST":
        if edit_svtn:
            forms = CreateSvTn(request.POST, request.FILES or None, instance=svtn)
        else:
            forms = CreateSvTn(request.POST, request.FILES or None)
        if forms.is_valid():
            svtn = forms.save(commit=False)
            svtn.sv_id = sv_id
            svtn.save()
            messages.success(request, "Cập nhật xét tốt nghiệp học viên thành công!")
            return redirect("sv_list")


    context = {
        "opt": opt,
        "hks": hks,
        "forms": forms,
        "hps": hps,
        "sv": sv
    }
    return render(request, "sms/sv_details.html", context)

@login_required
def details_diemtp(request, lop_id, lmh_id):
    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("lop","monhoc")[0]
    svs = Hssv.objects.filter(lop_id=lmh.lop_id)
    dtp = Diemthanhphan.objects.filter(sv__in = svs)
    lds = Loaidiem.objects.all()
    svl=[]

    for sv in svs:
        ldl=[]
        for ld in lds:
            dtpl=[]
            dtps = Diemthanhphan.objects.filter(sv = sv, monhoc_id = lmh.monhoc_id, tp_id = ld.id, status=1).order_by('log_id')
#            logs = dtps.log
            for dtp in dtps:
                dtpl.append({"id":dtp.log_id,"mark":dtp.diem})

            ldl.append({"ma":ld.ma,"dtplst": dtpl})

        svl.append({ "ma":sv.msv,"hoten":sv.hoten,"ttdiem": ldl})
#        mh.ttdiem = ld
 
    # print('after')
    # for mh in mhl:
    #     for ld in mh['ttdiem']:
    #         for dtp in ld['dtplst']:
    #             print(mh['ma'])
    #             print(ld['ma'])
    #             print(dtp['id'])
    #             print(dtp['mark'])

    context = {
        "svl": svl,
        "sv0": svl[0],
        "lmh": lmh
    }
    return render(request, "sms/details_diemtp.html", context)

@login_required
@permission_required('sms.change_hp81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def edit_hp81(request, lop_id, hp81_id):
    hp = Hp81.objects.get(id=hp81_id)
    sv_id = hp.sv_id
    hk_id = hp.hk_id
    sv = Hssv.objects.get(id = hp.sv_id)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lh_forms = CreateHp81(instance=hp)

    if request.method == "POST":
        edit_forms = CreateHp81(request.POST, request.FILES or None, instance=hp)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit Info Successfully!")
            return redirect("hv_hp81_hk_list", sv.lop_id)

    context = {
        "forms": lh_forms,
        "sv": sv,
        "hp": hp,
        "hk_id": hk_id
    }
    return render(request, "sms/edit_hp81.html", context)

@login_required
@permission_required('sms.change_hs81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def edit_hs81(request, lop_id, hs81_id):
    hs = Hs81.objects.get(id=hs81_id)
    sv_id = hs.sv_id
    hk_id = hs.hk_id
    sv = Hssv.objects.get(id = hs.sv_id)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lh_forms = CreateHs81(instance=hs)

    if request.method == "POST":
        edit_forms = CreateHs81(request.POST, request.FILES or None, instance=hs)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit Info Successfully!")
            return redirect("hv_hs81_hk_list", sv.lop_id)

    context = {
        "forms": lh_forms,
        "sv": sv,
        "hs": hs,
        "hk_id": hk_id
    }
    return render(request, "sms/edit_hs81.html", context)

@login_required
def edit_gv(request, gv_id):
    gv = Hsgv.objects.get(id=gv_id)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lh_forms = CreateGv(instance=gv)

    if request.method == "POST":
        edit_forms = CreateGv(request.POST, request.FILES or None, instance=gv)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit Giao viên Info Successfully!")
            return redirect("gv_list")

    context = {
        "forms": lh_forms,
        "gv": gv
    }
    return render(request, "sms/edit_gv.html", context)

@login_required
def edit_ns(request, ns_id):
    ns = Hsns.objects.get(id=ns_id)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    ns_forms = CreateNs(instance=ns)

    if request.method == "POST":
        edit_forms = CreateNs(request.POST, request.FILES or None, instance=ns)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit nhân sự Info Successfully!")
            return redirect("ns_list")

    context = {
        "forms": ns_forms,
        "ns": ns
    }
    return render(request, "sms/edit_ns.html", context)

@login_required
def edit_renter(request, renter_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền sửa thông tin khách thuê.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Validate renter access
    renter, error_response = validate_renter_access(request, renter_id)
    if error_response:
        return error_response
    
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    renter_forms = CreateRenter(instance=renter)

    if request.method == "POST":
        edit_forms = CreateRenter(request.POST, request.FILES or None, instance=renter)

        if edit_forms.is_valid():
            edit_forms.save()
            messages.success(request, "Edit Info Successfully!")
            return redirect("renter_list")

    context = {
        "forms": renter_forms,
        "renter": renter
    }
    return render(request, "sms/edit_renter.html", context)
@login_required
def edit_lop(request, lop_id):
    lop = Lop.objects.get(id=lop_id)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lop_forms = CreateLop(instance=lop)
#    lhk = LopHk.objects.select_related("hk").filter(lop_id=lop_id)

    if request.method == "POST":
        edit_forms = CreateLop(request.POST, request.FILES or None, instance=lop)

        if edit_forms.is_valid():
            edit_forms.save()
            # for hk in lhk:
            #     start_hk = "start"+str(hk.hk.ma)
            #     end_hk = "end"+str(hk.hk.ma)
            #     elhk = LopHk.objects.get(lop_id = lop_id, hk_id = hk.hk.ma)
            #     elhk.start_hk = request.POST[start_hk] if request.POST[start_hk] else None
            #     elhk.end_hk = request.POST[end_hk] if request.POST[end_hk] else None

            #     elhk.save()
            messages.success(request, "Cập nhật thông tin lớp thành công!")
            return redirect("lop_list")

    context = {
        "forms": lop_forms,
        "ma": lop.ma
#        "lhk": lhk
    }
    return render(request, "sms/edit_lop.html", context)

@login_required
def edit_lop_new(request, lop_id):
    lop = Lop.objects.get(id=lop_id)
    hk1 = LopHk.objects.get(lop_id=lop_id, hk_id=1)
    hk2 = LopHk.objects.get(lop_id=lop_id, hk_id=2)
    hk3 = LopHk.objects.get(lop_id=lop_id, hk_id=3)
    hk4 = LopHk.objects.get(lop_id=lop_id, hk_id=4)
    #lop_id, monhoc_id = lmh.lop_id, lmh.monhoc_id
    lop_forms = CreateLop(instance=lop,prefix='formlop')
    hk1_forms = CreateLopHk(instance=hk1,prefix='formhk1')
    hk2_forms = CreateLopHk(instance=hk2,prefix="formhk2")
    hk3_forms = CreateLopHk(instance=hk3,prefix="formhk3")
    hk4_forms = CreateLopHk(instance=hk4,prefix="formhk4")

    #lhk = LopHk.objects.select_related("hk").filter(lop_id=lop_id)

    if request.method == "POST":
        edit_flop = CreateLop(request.POST,request.FILES or None, instance=lop, prefix='formlop')
        edit_fhk1 = CreateLopHk(request.POST,request.FILES or None, instance=hk1, prefix='formhk1')
        edit_fhk2 = CreateLopHk(request.POST,request.FILES or None, instance=hk2, prefix='formhk2')
        edit_fhk3 = CreateLopHk(request.POST,request.FILES or None, instance=hk3, prefix='formhk3')
        edit_fhk4 = CreateLopHk(request.POST,request.FILES or None, instance=hk4, prefix='formhk4')

        if edit_flop.is_valid():
            edit_flop.save()
            messages.success(request, "Cập nhật thông tin lớp thành công!")
        if edit_fhk1.is_valid() and edit_fhk2.is_valid() and edit_fhk3.is_valid() and edit_fhk4.is_valid():
            edit_fhk1.save()
            edit_fhk2.save()
            edit_fhk3.save()
            edit_fhk4.save()
            messages.success(request, "Cập nhật thông tin HK thành công!")
        return redirect("lop_list")
    context = {
        "lop_forms": lop_forms,
        "lop": lop,
        "hk1": hk1,
        "hk2": hk2,
        "hk3": hk3,
        "hk4": hk4,
        "hk1_forms": hk1_forms,
        "hk2_forms": hk2_forms,
        "hk3_forms": hk3_forms,
        "hk4_forms": hk4_forms,
    }
    return render(request, "sms/edit_lop-new.html", context)

def edit_teacher(request, pk):
    teacher_edit = TeacherInfo.objects.get(id=pk)
    edit_teacher_forms = CreateTeacher(instance=teacher_edit)

    if request.method == "POST":
        edit_teacher_forms = CreateTeacher(request.POST, request.FILES or None, instance=teacher_edit)

        if edit_teacher_forms.is_valid():
            edit_teacher_forms.save()
            messages.success(request, "Edit Teacher Info Successfully!")
            return redirect("teacher_list")

    context = {
        "edit_teacher_forms": edit_teacher_forms
    }
    return render(request, "sms/edit_teacher.html", context)

def delete_teacher(request, teacher_id):
    teacher_delete = TeacherInfo.objects.get(id=teacher_id)
    teacher_delete.delete()
    messages.success(request, "Delete Teacher Info Successfully")
    return redirect("teacher_list")

#@login_required
def react(request):

    return render(request, 'sms/product-index.html')

@login_required
def download_file1(request):
    path = "uploads/monhoc.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404

@login_required
def download_temp(request, opt):
    if opt == 1:
        path = "uploads/tpl_monhoc.xlsx"
    elif opt == 2:
        path = "uploads/tpl_hssv.xlsx"
    elif opt == 3:
        path = "uploads/tpl_hs81.xlsx"
    elif opt == 4:
        path = "uploads/tpl_hp81.xlsx"
    elif opt == 5:
        path = "uploads/tpl_hsgv.xlsx"
    elif opt == 6:
        path = "uploads/tpl_hsns.xlsx"
    elif opt == 7:
        path = "uploads/tpl_renter.xlsx"
    else:
        path = "uploads/monhoc.xlsx"

    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404

@login_required
def download_temp_data(request, lop_id, opt):
    if opt == 1:
        temp_path = "uploads/tpl_hs81.xlsx"
        out_path = "uploads/tpl_" + str(lop_id) + "_hs81.xlsx"
        sheet_name = 'hs81'
    elif opt == 2:
        temp_path = "uploads/tpl_hp81.xlsx"
        out_path = "uploads/tpl_" + str(lop_id) + "_hp81.xlsx"
        sheet_name = 'hp81'
    else:
        temp_path = "uploads/tpl_hs81.xlsx"
        out_path = "uploads/tpl_" + str(lop_id) + "_hs81.xlsx"
        sheet_name = 'hs81'


    temp_file_path = os.path.join(settings.MEDIA_ROOT, temp_path)
    out_file_path = os.path.join(settings.MEDIA_ROOT, out_path)

    shutil.copy(temp_file_path, out_file_path)
    #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)

    locale.setlocale(locale.LC_ALL, 'vi_VN')
    svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)

    #export to excel
    exp=[]
    for sv in svs:
        exp.append({"Mã học tên": sv.msv,
                    "Họ tên": sv.hoten, 
                    })

    # Convert the QuerySet to a DataFrame
    df = pd.DataFrame(list(exp))
    with pd.ExcelWriter(out_file_path, mode="a",engine="openpyxl", if_sheet_exists="overlay",) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=1, startcol=0)

    #if os.path.exists(out_file_path):
    with open(out_file_path, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename(out_file_path)
    
    os.remove(out_file_path)

    return response
    #raise Http404

@login_required
@permission_required('sms.add_uploadedfile',raise_exception=True)
def upload_file(request):
    if request.method == 'POST':
        form = CreateUploadFile(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "File mẫu được tải lên thành công!")
            return redirect('upload_file')
    else:
        form = CreateUploadFile()

    list_of_ids = []
    for f in UploadedFile.objects.all():
        if not (f.lopmh_id or f.user_id):
            list_of_ids.append(f.id)

    files = UploadedFile.objects.filter(id__in=list_of_ids).order_by('-id')
    context = {
        'form': form,
        'files': files
    }
    return render(request, "sms/file_list.html", context)

@login_required
def hoclai_list(request,lmh_id):
    if request.method == "POST":
        sv_id = request.POST.get('sv', None)
        #svs = Hssv.objects.filter(lop_id=lop_id)
        hl = Hoclai(lmh_id=lmh_id, sv_id=sv_id)
        hl.save()
        messages.success(request, "Thêm học viên học lại thành công!")

    lmh = LopMonhoc.objects.get(id = lmh_id )

    hls = Hoclai.objects.filter(lmh_id = lmh_id).select_related("sv")
    svs = Hssv.objects.exclude(lop_id = lmh.lop_id).exclude(id__in = hls.values_list('sv_id', flat=True)).order_by('-id')


    svl = Hssv.objects.filter(id__in= hls.values_list('sv_id', flat=True)).order_by('-id')
    print('ok')
    context = {
        'svl': svl,
        'lmh': lmh,
        'svs': svs
    }
    return render(request, "sms/hoclai_list.html", context)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def upload_file_gv(request, gv_id):

    gv = Hsgv.objects.get(id = gv_id)
    if request.method == 'POST':
        if not(gv.user == request.user):
            return HttpResponseForbidden()
        #test if file upload is image        
        image = request.FILES['file']

        pre, ext = os.path.splitext(image.name)
        print(ext)
        if ext =='.pdf':
            try:
                PdfReader(image)
            except PdfReadError:
                #print("invalid PDF file")
                messages.error(request, "invalid PDF file")
                return redirect('upload_file_gv', gv_id)
        # check file ảnh
        else:
            try:
                img = Image.open(image)
                img = img.convert('RGB')
            except Exception as e:
                messages.error(request, 'Chỉ chấp nhận file ảnh')
                return redirect('upload_file_gv', gv_id)
        #img.save("uploads/temp.pdf", format="PDF")

        limit = 5 * 1024 * 1024
        if image.size > limit:
            #raise ValidationError('File too large. Size should not exceed 2 MiB.')
            messages.error(request, 'File too large. Size should not exceed 5 MiB.')
            return redirect('upload_file_gv', gv_id)

        form = CreateUploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = form.save()
            file.user = request.user
            file.uploaded_by = request.user.username
            file.save()
            messages.success(request, "File được tải thành công!")
            return redirect('upload_file_gv', gv_id)
        else:
            for error in form.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

    else:
        form = CreateUploadFile()

    if gv.user:
        files = UploadedFile.objects.filter(user = gv.user)
        for file in files:
            file.ten = file.file.name.split("/")[-1]
    else:
        files = None

    tl=[{"ma":"Bằng Tốt nghiệp"}, {"ma":"Bảng điểm"}, {"ma":"Chứng chỉ NVSP/dạy nghề"}, {"ma":"Sơ yếu lý lịch"}, {"ma":"Chứng chỉ tiếng Anh"}, {"ma":"Chứng chỉ tin học"}]

    context = {
        'form': form,
        'gv': gv,
        'tl': tl,
        'files': files
    }
    return render(request, "sms/file_list_gv.html", context)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def upload_file_hv(request, hv_id):

    sv = Hssv.objects.get(id = hv_id)
    if request.method == 'POST':
        if not(sv.user == request.user):
            return HttpResponseForbidden()
        #test if file upload is image        
        image = request.FILES['file']

        pre, ext = os.path.splitext(image.name)
        print(ext)
        if ext =='.pdf':
            try:
                PdfReader(image)
            except PdfReadError:
                #print("invalid PDF file")
                messages.error(request, "invalid PDF file")
                return redirect('upload_file_hv', hv_id)
        # check file ảnh
        else:
            try:
                img = Image.open(image)
                img = img.convert('RGB')
            except Exception as e:
                messages.error(request, 'Chỉ chấp nhận file ảnh')
                return redirect('upload_file_hv', hv_id)
        #img.save("uploads/temp.pdf", format="PDF")

        limit = 5 * 1024 * 1024
        if image.size > limit:
            #raise ValidationError('File too large. Size should not exceed 2 MiB.')
            messages.error(request, 'File too large. Size should not exceed 5 MiB.')
            return redirect('upload_file_hv', hv_id)

        form = CreateUploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = form.save()
            file.user = request.user
            file.uploaded_by = request.user.username
            file.save()
            messages.success(request, "File được tải thành công!")
            return redirect('upload_file_hv', hv_id)
        else:
            for error in form.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

    else:
        form = CreateUploadFile()

    if sv.user:
        files = UploadedFile.objects.filter(user = sv.user)
        for file in files:
            file.ten = file.file.name.split("/")[-1]
    else:
        files = None


    tl=[{"ma":"Phiếu xét tuyển"}, 
        {"ma":"Bằng TN"}, 
        {"ma":"Giay CNTTTT"}, 
        {"ma":"Học bạ THCS"}, 
        {"ma":"CCCD"}, 
        {"ma":"Giấy khai sinh"}, 
        {"ma":"Sổ HK"}, 
        {"ma":"Ảnh 3x4"}
        ]

    context = {
        'form': form,
        'sv': sv,
        'tl': tl,
        'files': files
    }
    return render(request, "sms/file_list_hv.html", context)

@login_required
def view_house(request, house_id):
    """Xem chi tiết nhà trọ (read-only)"""
    house = get_object_or_404(House, id=house_id)
    
    # Kiểm tra quyền: chỉ chủ nhà của location hoặc người thuê hiện tại
    is_owner = house.loc.chu == request.user
    is_current_renter = False
    
    try:
        renter = request.user.is_renter
        active_contract = HouseRenter.objects.filter(
            house=house, 
            renter=renter,
            active=True
        ).first()
        is_current_renter = active_contract is not None
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if not (is_owner or is_current_renter):
        messages.error(request, "Bạn không có quyền xem nhà trọ này")
        return redirect('home')
    
    # Lấy hợp đồng hiện tại
    active_contract = HouseRenter.objects.filter(
        house=house, 
        active=True
    ).select_related('renter').first()
    
    # Lấy lịch sử hợp đồng
    contract_history = HouseRenter.objects.filter(
        house=house
    ).select_related('renter').order_by('-active', '-rent_from')
    
    # Lấy hóa đơn gần nhất
    recent_bills = Hoadon.objects.filter(house=house).order_by('-ngay_tao')[:5]
    
    context = {
        'house': house,
        'active_contract': active_contract,
        'contract_history': contract_history,
        'recent_bills': recent_bills,
        'is_owner': is_owner,
        'is_current_renter': is_current_renter,
    }
    return render(request, "sms/view_house.html", context)

@login_required
def renter_view_house(request, house_id: int):
    """Trang chi tiết nhà cho renter (read-only, giao diện renter). 
    Chỉ cho phép khi người dùng là renter có hợp đồng với căn nhà này (bất kỳ trạng thái).
    """
    house = get_object_or_404(House.objects.select_related('loc', 'loc__xp'), id=house_id)

    # Lấy renter từ user
    try:
        renter = Renter.objects.get(user=request.user)
    except Renter.DoesNotExist:
        messages.error(request, "Bạn chưa được liên kết với tài khoản khách thuê.")
        return redirect('renter_landing')

    # Kiểm tra quan hệ hợp đồng giữa renter và house
    any_contract = HouseRenter.objects.filter(house=house, renter=renter).exists()
    if not any_contract:
        messages.error(request, "Bạn không có quyền xem phòng này.")
        return redirect('renter_locations')

    # Hợp đồng hiện tại của renter với căn nhà này
    active_contract = HouseRenter.objects.filter(
        house=house, renter=renter, active=True
    ).select_related('renter').first()

    # Lịch sử hợp đồng của CHÍNH renter này với căn nhà
    contract_history = HouseRenter.objects.filter(
        house=house, renter=renter
    ).select_related('renter').order_by('-active', '-rent_from')

    context = {
        'house': house,
        'active_contract': active_contract,
        'contract_history': contract_history,
        'is_owner': False,
        'is_current_renter': active_contract is not None,
    }
    return render(request, "sms/renter_view_house.html", context)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def view_loc(request, loc_id):
    # Validate location access
    loc, error_response = validate_location_access(request, loc_id)
    if error_response:
        return error_response

# ----------------------------------------------------
    # BỔ SUNG LOGIC LẤY DỮ LIỆU NHÀ VÀ NGƯỜI THUÊ
    # ----------------------------------------------------
    
    # 1. Lấy tất cả các House thuộc Location này
    houses = House.objects.filter(loc_id=loc_id)

    # 2. Lấy thông tin người thuê hiện tại (Active Renter) cho từng nhà
    for house in houses:
        # Lấy hợp đồng đang active cho căn nhà này (chỉ 1 bản ghi)
        active_renter_record = HouseRenter.objects.filter(
            house=house, 
            active=True
        ).select_related('renter').first() # Tối ưu hóa truy vấn
        
        # Gán bản ghi active và thông tin người thuê vào đối tượng house
        house.active_contract = active_renter_record
        house.current_renter = active_renter_record.renter if active_renter_record else None
        # Đếm số hợp đồng của nhà này
        try:
            house.contract_count = HouseRenter.objects.filter(house=house).count()
        except Exception:
            house.contract_count = 0

    # ----------------------------------------------------


    files = UploadedFile.objects.filter(link_id = loc_id, type=1)
    for file in files:
        file.ten = file.file.name.split("/")[-1]

    context = {
    'loc': loc,
    'files': files,
    'houses': houses, # <--- BIẾN MỚI
    'is_owner': loc.chu == request.user,
    }
    return render(request, "sms/view_loc.html", context)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def upload_file_loc(request, loc_id):
    # Validate location access
    loc, error_response = validate_location_access(request, loc_id)
    if error_response:
        return error_response

    if request.method == 'POST':
#        if not(sv.user == request.user):
 #           return HttpResponseForbidden()
        #test if file upload is image        
        image = request.FILES['file']

        pre, ext = os.path.splitext(image.name)
        print(ext)
        if ext =='.pdf':
            try:
                PdfReader(image)
            except PdfReadError:
                #print("invalid PDF file")
                messages.error(request, "invalid PDF file")
                return redirect('upload_file_loc', loc_id)
        # check file ảnh
        else:
            try:
                img = Image.open(image)
                img = img.convert('RGB')
            except Exception as e:
                messages.error(request, 'Chỉ chấp nhận file ảnh')
                return redirect('upload_file_loc', loc_id)
        #img.save("uploads/temp.pdf", format="PDF")

        limit = 5 * 1024 * 1024
        if image.size > limit:
            #raise ValidationError('File too large. Size should not exceed 2 MiB.')
            messages.error(request, 'File too large. Size should not exceed 5 MiB.')
            return redirect('upload_file_loc', loc_id)

        form = CreateUploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = form.save()
            file.user = request.user
            file.uploaded_by = request.user.username
            file.type = 1
            file.link_id = loc_id
            file.save()
            messages.success(request, "File được tải thành công!")
            return redirect('upload_file_loc', loc_id)
        else:
            for error in form.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

    else:
        form = CreateUploadFile()


    files = UploadedFile.objects.filter(link_id = loc_id, type=1)
    for file in files:
        file.ten = file.file.name.split("/")[-1]

    context = {
    'form': form,
    'loc': loc,
    'files': files,
    }
    return render(request, "sms/file_list_loc.html", context)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def upload_bill_file(request, bill_id):

    bill = Hoadon.objects.get(id=bill_id)
    if request.method == 'POST':
#        if not(sv.user == request.user):
 #           return HttpResponseForbidden()
        #test if file upload is image        
        image = request.FILES['file']

        pre, ext = os.path.splitext(image.name)
        print(ext)
        if ext =='.pdf':
            try:
                PdfReader(image)
            except PdfReadError:
                #print("invalid PDF file")
                messages.error(request, "invalid PDF file")
                return redirect('bill_detail', bill_id)
        # check file ảnh
        else:
            try:
                img = Image.open(image)
                img = img.convert('RGB')
            except Exception as e:
                messages.error(request, 'Chỉ chấp nhận file ảnh')
                return redirect('bill_detail', bill_id)
        #img.save("uploads/temp.pdf", format="PDF")

        limit = 5 * 1024 * 1024
        if image.size > limit:
            #raise ValidationError('File too large. Size should not exceed 2 MiB.')
            messages.error(request, 'File too large. Size should not exceed 5 MiB.')
            return redirect('bill_detail', bill_id)

        form = CreateUploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = form.save()
            file.user = request.user
            file.uploaded_by = request.user.username
            file.type = 2
            file.link_id = bill_id
            file.save()
            messages.success(request, "File được tải thành công!")
            return redirect('bill_detail', bill_id)
        else:
            for error in form.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

    else:
        form = CreateUploadFile()


    files = UploadedFile.objects.filter(link_id = bill_id, type=2)
    for file in files:
        file.ten = file.file.name.split("/")[-1]

    context = {
    'form': form,
    'bill': bill,
    'files': files,
    }
    return redirect('bill_detail', bill_id)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def create_hr(request, id):
    # Validate house access
    house, error_response = validate_house_access(request, id)
    if error_response:
        return error_response

    if request.method == 'POST':
        # Sử dụng HouseRenterForm bạn đã tạo
        forms = CreateHouseRenter(request.POST) 
        
        if forms.is_valid():
            new_renter_contract = forms.save(commit=False)
            # Gắn house trước khi xử lý trạng thái các hợp đồng khác
            new_renter_contract.house_id = id

            # --- Logic nghiệp vụ: Cập nhật trạng thái active của các hợp đồng cũ ---
            if new_renter_contract.active:
                # Đặt tất cả hợp đồng đang active của nhà trọ này về False
                HouseRenter.objects.filter(house_id=id, active=True).update(active=False)

            # Lưu hợp đồng mới
            new_renter_contract.save()
            
            messages.success(request, 'Thêm hợp đồng thuê mới thành công!')
            
            # AJAX/HTMX request: trả về timeline mới
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('HX-Request'):
                hrs = HouseRenter.objects.filter(house_id=id).order_by('-rent_from','-id')
                return render(request, 'sms/fragments/_hr_timeline_list.html', {'hrs': hrs, 'id': id})
            
            # Normal request: chuyển hướng về trang hiện tại
            return redirect('hr_list_secure', id)
        else:
            for field, errors in forms.errors.items():
                for error in errors:
                    messages.error(request, f"Lỗi ở trường '{field}': {error}")

    # -----------------------------------
    # Logic Xử lý Form GET (Hiển thị)
    # -----------------------------------
    else:
        # Nếu là GET request, tạo một Form trống để hiển thị
        forms = CreateHouseRenter() 
    #hrs = HouseRenter.objects.filter(house_id = id).order_by('-id')
    hrs = HouseRenter.objects.filter(house_id = id).order_by('-rent_from','-id')

    context = {
    'forms': forms,
    'id': id,
    'hrs': hrs
    }
    return render(request, "sms/create_hr.html", context)

# app_quanly/views.py (Thay thế hàm cũ bằng hàm này)

# Đảm bảo bạn đã import các thư viện cần thiết ở đầu file:
# from .forms import * # from django.shortcuts import render, redirect 
# from django.contrib import messages
# from django.contrib.auth.decorators import login_required
# from .models import HouseRenter 

@login_required
def hr_list(request, id):
    # Block renter access
    renter_block = block_renter_access(request)
    if renter_block:
        return renter_block
    
    # Validate house access
    house, error_response = validate_house_access(request, id)
    if error_response:
        return error_response
    
    hrs = HouseRenter.objects.filter(house_id=id).order_by('-rent_from','-id')
    house = get_object_or_404(House, pk=id)
    location = house.loc if house else None
    owner = location.chu if location else None
    
    # -----------------------------------
    # Logic Xử lý Form POST (Thêm mới)
    # -----------------------------------
    if request.method == 'POST':
        # Sử dụng HouseRenterForm bạn đã tạo
        forms = CreateHouseRenter(request.POST, owner=owner)
        
        if forms.is_valid():
            new_renter_contract = forms.save(commit=False)
            # Gắn house trước khi xử lý trạng thái các hợp đồng khác
            new_renter_contract.house_id = id

            # --- Logic nghiệp vụ: Cập nhật trạng thái active của các hợp đồng cũ ---
            if new_renter_contract.active:
                # Đặt tất cả hợp đồng đang active của nhà trọ này về False
                HouseRenter.objects.filter(house_id=id, active=True).update(active=False)

            # Lưu hợp đồng mới
            new_renter_contract.save()
            
            messages.success(request, 'Thêm hợp đồng thuê mới thành công!')
            # Quay lại danh sách hợp đồng của nhà hiện tại
            return redirect('hr_list_secure', id)
        else:
            # Nếu Form không hợp lệ, giữ nguyên forms để hiển thị lỗi và dữ liệu cũ
            messages.error(request, 'Lỗi: Dữ liệu nhập không hợp lệ. Vui lòng kiểm tra lại Form.')
            # Tiếp tục logic bên dưới để render lại trang

    # -----------------------------------
    # Logic Xử lý Form GET (Hiển thị)
    # -----------------------------------
    else:
        # Nếu là GET request, tạo một Form trống để hiển thị
        forms = CreateHouseRenter(owner=owner)
    
    # Chuẩn bị Context để render template
    context = {
        'segment': 'createhouserenter',
        'forms': forms, # Form để hiển thị (có thể chứa lỗi nếu POST thất bại)
        'hrs': hrs,     # Dữ liệu để hiển thị Timeline
        'id': id,
        'house': house,
        'location': location
    }
    return render(request, 'sms/house_renter_list_gemini.html', context)

@login_required
def hr_list_partial(request, id):
    # Block renter access
    renter_block = block_renter_access(request)
    if renter_block:
        return renter_block
    
    # Validate house access
    house, error_response = validate_house_access(request, id)
    if error_response:
        return error_response

    hrs = HouseRenter.objects.filter(house_id=id).order_by('-rent_from','-id')
    return render(request, 'sms/fragments/_hr_timeline_list.html', { 'hrs': hrs, 'id': id })

@login_required
def hr_toggle_active(request, hr_id: int):
    """Quick action to toggle a HouseRenter contract's active status from the timeline.
    If turning active on, ensure all other contracts of the same house are set to inactive.
    """
    # Validate access to the contract
    contract, error_response = validate_contract_access(request, hr_id)
    if error_response:
        return error_response

    house_id = contract.house_id

    if request.method != 'POST':
        messages.error(request, 'Phương thức không hợp lệ.')
        return redirect('hr_list_secure', house_id)

    if not contract.active:
        # Activate this contract and deactivate any other active contracts for the same house
        HouseRenter.objects.filter(house_id=house_id, active=True).exclude(id=contract.id).update(active=False)
        contract.active = True
        contract.save(update_fields=['active'])
        messages.success(request, 'Đã kích hoạt hợp đồng và vô hiệu các hợp đồng khác của nhà này.')
    else:
        # Deactivate this contract
        contract.active = False
        contract.save(update_fields=['active'])
        messages.success(request, 'Đã hủy kích hoạt hợp đồng.')
    # If HTMX request, return the updated fragment for this item
    if request.headers.get('HX-Request') == 'true':
        # For HTMX, refresh the entire list so other items reflect the new active state
        hrs = HouseRenter.objects.filter(house_id=house_id).order_by('-rent_from','-id')
        return render(request, 'sms/fragments/_hr_timeline_list.html', { 'hrs': hrs, 'id': house_id })
    # Fallback full page redirect
    return redirect('hr_list_secure', house_id)

@login_required
def hr_delete(request, hr_id: int):
    """Delete a HouseRenter contract only if there is no invoice associated (same house & renter).
    Supports HTMX inline deletion (removes item) or full redirect with message.
    """
    contract, error_response = validate_contract_access(request, hr_id)
    if error_response:
        return error_response

    house_id = contract.house_id

    if request.method != 'POST':
        messages.error(request, 'Phương thức không hợp lệ.')
        return redirect('hr_list_secure', house_id)

    # Delete without invoice restriction per updated requirement
    contract.delete()
    msg = 'Đã xóa hợp đồng thành công.'
    messages.success(request, msg)
    if request.headers.get('HX-Request') == 'true':
        # Return refreshed full timeline list so UI updates immediately
        hrs = HouseRenter.objects.filter(house_id=house_id).order_by('-rent_from','-id')
        response = render(request, 'sms/fragments/_hr_timeline_list.html', { 'hrs': hrs, 'id': house_id })
        response["HX-Trigger"] = json.dumps({"showToast": {"message": msg, "level": "success"}})
        return response
    return redirect('hr_list_secure', house_id)

# ----------------------------
# Secure Contract-related views
# ----------------------------
from django.shortcuts import get_object_or_404

def contract_detail(request, contract_id: int):
    """Display details of a single HouseRenter contract (hashid-based route).

    Uses validate_contract_access for authorization.
    """
    contract, error_response = validate_contract_access(request, contract_id)
    if error_response:
        return error_response
    return render(request, 'sms/contract_detail.html', {
        'contract': contract,
        'house': contract.house,
        'renter': contract.renter,
        'segment': 'contract_detail'
    })


def house_contracts(request, house_id: int):
    """List contracts for a given house (hashid-based route)."""
    # Lazy import to avoid top-of-file changes
    from .models import House, HouseRenter as HR
    house = get_object_or_404(House, pk=house_id)

    if not request.user.is_authenticated:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Authentication required')

    # Basic ownership check: owner of location or staff
    owner_id = getattr(getattr(house, 'loc', None), 'chu_id', None)
    if not (request.user.is_staff or owner_id == request.user.id):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Not allowed to view contracts for this house')

    contracts = HR.objects.filter(house=house).order_by('-rent_from', '-id')
    active_contract = contracts.filter(active=True).first()
    return render(request, 'sms/house_contracts.html', {
        'house': house,
        'contracts': contracts,
        'active_contract': active_contract,
        'segment': 'house_contracts'
    })

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def lophk_list(request, lop_id, lhk_id):
    if lhk_id > 0:
        lhk = LopHk.objects.filter(id = lhk_id).select_related('lop',"hk").order_by('hk__ma')[0]
    else:
        #lhk = LopHk.objects.filter(lop_id = lop_id).select_related('lop',"hk").order_by('hk__ma')[0]
        if not LopHk.objects.filter(lop_id = lop_id).exists():
            hks= Hocky.objects.all()
            for hk in hks:
                lhk = LopHk.objects.create(hk_id = hk.id, lop_id = lop_id)
                lhk.save()        
        lhk = LopHk.objects.filter(lop_id = lop_id).select_related('lop',"hk").order_by('hk__ma')[0]
    
    lhks = LopHk.objects.filter(lop_id = lop_id).select_related('lop',"hk").order_by('hk__ma')
    
    if request.method == 'POST':
        form = CreateLopHk(request.POST, request.FILES or None, instance=lhk)

        if form.is_valid():
            file = form.save()
            messages.success(request, "Cập nhật thành công!")
        # else:
        #     for error in form.errors:
        #         print(error)
        #         messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

    else:
        form = CreateLopHk(instance=lhk)
    context = {
        'forms': form,
        'lhks': lhks,
        'lhk': lhk
    }
    return render(request, "sms/lophk_list.html", context)

@login_required
#@permission_required('sms.add_uploadedfile',raise_exception=True)
def upload_file_lmh(request, lmh_id):

    lmh = LopMonhoc.objects.get(id = lmh_id)
    lop = Lop.objects.get(id = lmh.lop_id)
    if request.user.is_gv:
        if not request.user.has_perm('assign_lopmonhoc', lmh):
            return HttpResponseForbidden()
    else:
        if not request.user.has_perm('assign_lop', lop):
            return HttpResponseForbidden()
    
    if request.method == 'POST':
        # if not(gv.user == request.user):
        #     return HttpResponseForbidden()
        #test if file upload is image        
        image = request.FILES['file']

        pre, ext = os.path.splitext(image.name)
        print(ext)
        if ext =='.pdf':
            try:
                PdfReader(image)
            except PdfReadError:
                #print("invalid PDF file")
                messages.error(request, "invalid PDF file")
                return redirect('upload_file_lmh', lmh_id)
        # check file ảnh
        else:
            try:
                img = Image.open(image)
                img = img.convert('RGB')
            except Exception as e:
                messages.error(request, 'Chỉ chấp nhận file ảnh')
                return redirect('upload_file_lmh', lmh_id)
        #img.save("uploads/temp.pdf", format="PDF")

        limit = 5 * 1024 * 1024
        if image.size > limit:
            #raise ValidationError('File too large. Size should not exceed 2 MiB.')
            messages.error(request, 'File too large. Size should not exceed 5 MiB.')
            return redirect('upload_file_lmh', lmh_id)

        form = CreateUploadFile(request.POST, request.FILES)
        if form.is_valid():
            file = form.save()
            file.uploaded_by = request.user.username
            file.user = request.user
            file.lopmh_id = lmh_id
            file.save()
            messages.success(request, "File được tải thành công!")
            return redirect('upload_file_lmh', lmh_id)
        else:
            for error in form.errors:
                print(error)
                messages.error(request, "Lỗi khi lưu dữ liệu: " + str(error))

    else:
        form = CreateUploadFile()

    files = UploadedFile.objects.filter(lopmh_id = lmh_id)
    for file in files:
        file.ten = file.file.name.split("/")[-1]
        
    tl=[{"ma":"Tiến độ, kế hoạch, CTĐT"}, 
        {"ma":"Phiếu báo giảng"}, 
        {"ma":"Giáo án"}, 
        {"ma":"Sổ tay giảng viên"}, 
        {"ma":"Danh sách học sinh ký thi"}, 
        {"ma":"Bài thi"},
        {"ma":"Đề thi/đáp án"}, 
        ]

    context = {
        'form': form,
        'tl': tl,
        'lmh': lmh,
        'files': files
    }
    return render(request, "sms/file_list_lmh.html", context)

@login_required
@permission_required('sms.view_uploadedfile',raise_exception=True)
def download_file(request, file_id):
    uploaded_file = UploadedFile.objects.get(pk=file_id)
    response = HttpResponse(uploaded_file.file, content_type='application/force-download')
    response['Content-Disposition'] = f'attachment; filename="{uploaded_file.file.name}"'
    return response

@login_required
@permission_required('sms.view_uploadedfile',raise_exception=True)
def download_temp_diem(request, lmh_id):

    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("monhoc", "lop")[0]
    locale.setlocale(locale.LC_ALL, 'vi_VN')

    hls = Hoclai.objects.filter(lmh_id = lmh_id)
    svs = sorted(Hssv.objects.filter(lop_id = lmh.lop.id) | Hssv.objects.filter(id__in = hls.values_list('sv_id', flat=True)), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)

    #export to excel
    exp=[]
    for sv in svs:
        exp.append({"Mã học tên": sv.msv,
                    "Họ tên": sv.hoten, 
                    "Điểm": "", 
                    })

    # Convert the QuerySet to a DataFrame
    df = pd.DataFrame(list(exp))
    tenf = lmh.lop.ten + "_"+ lmh.monhoc.ma + ".xlsx"
    print(tenf)
    # Define the Excel file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + tenf

    # Use Pandas to write the DataFrame to an Excel file
    df.to_excel(response, sheet_name='diemtp', index=False, engine='openpyxl')

    return response

@login_required
#@permission_required('sms.view_uploadedfile',raise_exception=True)
def view_file(request, file_id):
    #import magic
    from PIL import Image
    uploaded_file = UploadedFile.objects.get(pk=file_id)
    if request.user.is_gv:
        if uploaded_file.lopmh is not None:
            if not request.user.has_perm('assign_lopmonhoc', uploaded_file.lopmh):
                return HttpResponseForbidden()
        else:
            if not(uploaded_file.user == request.user):
                return HttpResponseForbidden()
    elif request.user.is_hv:
        if not(uploaded_file.user == request.user):
            return HttpResponseForbidden()
    elif not request.user.has_perm('sms.view_uploadedfile'): 
        return HttpResponseForbidden()

    try:
        file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)

        pre, ext = os.path.splitext(file_path)
        if ext !='.pdf':
            filepdf = pre + '.pdf'
        #    os.rename(renamee, pre + new_extension)
            img = Image.open(file_path)
            img = img.convert('RGB')
            img.save(filepdf, format="PDF")
        else:
            filepdf = file_path
        head, tail = os.path.split(filepdf)
        print(head)
        print(tail)
        return FileResponse(open(filepdf, 'rb'), content_type='application/pdf')
    except Exception as e:
        messages.error(request, 'Có lỗi khi view file')

    #response = HttpResponse(uploaded_file.file, content_type='application/force-download')
    #response['Content-Disposition'] = f'attachment; filename="{uploaded_file.file.name}"'
    #return response

@login_required
@permission_required('sms.delete_uploadedfile',raise_exception=True)
def delete_file(request, file_id):

    uploaded_file = UploadedFile.objects.get(pk=file_id)

    uploaded_file.delete()
    file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)
    if os.path.exists(file_path):
        os.remove(file_path)
        messages.success(request, "File được xóa thành công!")
    return redirect('upload_file')


@login_required
def delete_file_loc(request, loc_id, file_id):
    # Validate location access
    loc, error_response = validate_location_access(request, loc_id)
    if error_response:
        return error_response

    uploaded_file = UploadedFile.objects.get(pk=file_id)

    # ensure the file belongs to this location
    try:
        if int(uploaded_file.link_id) != int(loc_id):
            return HttpResponseForbidden()
    except Exception:
        return HttpResponseForbidden()

    # allow owner or users with delete permission
    if hasattr(uploaded_file, 'user') and uploaded_file.user == request.user:
        pass
    elif not request.user.has_perm('sms.delete_uploadedfile'):
        return HttpResponseForbidden()

    uploaded_file.delete()
    file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)
    if os.path.exists(file_path):
        os.remove(file_path)
        messages.success(request, "File được xóa thành công!")
    return redirect('view_loc', loc_id)

@login_required
def delete_hoclai(request, lmh_id, sv_id):

    hl = Hoclai.objects.get(sv_id=sv_id)
    hl.delete()

    messages.success(request, "Học viên học lại được xóa thành công!")
    return redirect('hoclai_list', lmh_id)

@login_required
#@permission_required('sms.delete_uploadedfile',raise_exception=True)
def delete_file_gv(request, gv_id, file_id):

    uploaded_file = UploadedFile.objects.get(pk=file_id)

    if request.user.is_gv:
        if not(uploaded_file.user == request.user):
            return HttpResponseForbidden()
    elif not request.user.has_perm('sms.delete_uploadedfile'): 
        return HttpResponseForbidden()


    uploaded_file.delete()
    file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)
    if os.path.exists(file_path):
        os.remove(file_path)
        messages.success(request, "File được xóa thành công!")
    return redirect('upload_file_gv', gv_id)

@login_required
#@permission_required('sms.delete_uploadedfile',raise_exception=True)
def delete_file_hv(request, hv_id, file_id):

    uploaded_file = UploadedFile.objects.get(pk=file_id)

    if request.user.is_hv:
        if not(uploaded_file.user == request.user):
            return HttpResponseForbidden()
    elif not request.user.has_perm('sms.delete_uploadedfile'): 
        return HttpResponseForbidden()

    uploaded_file.delete()
    file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)
    if os.path.exists(file_path):
        os.remove(file_path)
        messages.success(request, "File được xóa thành công!")
    return redirect('upload_file_hv', hv_id)

@login_required
#@permission_required('sms.delete_uploadedfile',raise_exception=True)
def delete_file_lmh(request, lmh_id, file_id):
    uploaded_file = UploadedFile.objects.get(pk=file_id)

    if request.user.is_gv:
        if not(uploaded_file.user == request.user):
            return HttpResponseForbidden()
    elif not request.user.has_perm('sms.delete_uploadedfile'): 
        return HttpResponseForbidden()

    uploaded_file.delete()
    file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)
    if os.path.exists(file_path):
        os.remove(file_path)
        messages.success(request, "File được xóa thành công!")
    return redirect('upload_file_lmh', lmh_id)
@login_required
def download_file2(request):
    import pandas as pd

    path = "uploads/pandas_simple.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, path)

    # Create a Pandas dataframe from the data.
    df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name='Sheet1')

    # Get the xlsxwriter workbook and worksheet objects.
    workbook = writer.book
    worksheet = writer.sheets["Sheet1"]

    # Create a chart object.
    chart = workbook.add_chart({"type": "column"})

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df.shape

    # Configure the series of the chart from the dataframe data.
    chart.add_series({"values": ["Sheet1", 1, 1, max_row, 1]})

    # Insert the chart into the worksheet.
    worksheet.insert_chart(1, 3, chart)

    writer.close()    # Get the xlsxwriter objects from the dataframe writer object.

    #path = "uploads/monhoc.xlsx"
    #file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404

    # Save the workbook to response object
    #workbook.save(response) # Chỉ dùng với openpyxl
#    writer.close()    # Get the xlsxwriter objects from the dataframe writer object.

    # Return the response object
 #   return response

@login_required
def report_hs81(request):
    hps = Hp81.objects.all().select_related("sv")
    if request.method == "POST":
            query_lop = request.POST.get('lop', None)
            query_tt = request.POST.get('trungtam', None)
            lh = Lop.objects.all().select_related("trungtam")
            if query_tt:
                lh = lh.filter(trungtam__ten__contains=query_tt)
            if query_lop:
                lh = lh.filter(lop__ten__contains=query_lop)     

            hps = hps.filter(sv__malop__in=lh.values_list('id', flat=True))
            messages.success(request, "Tìm kiếm thành công!")
    paginator = Paginator(hps, 50)
    page = request.GET.get('page')
    paged_students = paginator.get_page(page)
    context = {
        "hps": paged_students
    }
    return render(request, "sms/report_hs81.html", context)

@login_required()
def reset_pwd(request, ns_id):
    if not request.user.is_superuser:
        return redirect("/")
    ns = Hsns.objects.get(id = ns_id)
    # Check if username already exists
    if User.objects.filter(username=ns.ma).exists():
        user = User.objects.get(username=ns.ma)
        user.set_password(ns.ma + '@123654')
        user.save()    
        messages.success(request, "Reset password thành công!")
        return redirect('ns_list')    
    return redirect("ns_list")

@login_required()
def add_nsuser(request, id):
    if not request.user.is_superuser:
        return redirect("/")
    ns = Hsns.objects.get(id = id)
    # Check if username already exists
    if User.objects.filter(username=ns.ma).exists():
        messages.error(request, 'Username already exists')
        return redirect('ns_list')    
    
    user = User.objects.create_user(
        username=ns.ma,
        password=ns.ma + '@123654'
    )
    # Set họ tên để hiển thị thay vì username
    name_parts = ns.hoten.strip().split(maxsplit=1)
    user.last_name = name_parts[0] if name_parts else ''
    user.first_name = name_parts[1] if len(name_parts) > 1 else ''
    user.save()
    ns.user = user
    ns.save()
    messages.success(request, "Tạo tài khoản cho " + ns.hoten + " thành công")
    return redirect("ns_list")

@login_required
@login_required
def user_changepwd(request):
    if request.method == "POST":
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('password', '')
        
        # Kiểm tra mật khẩu hiện tại
        if not request.user.check_password(current_password):
            messages.error(request, "Mật khẩu hiện tại không đúng!")
            return render(request, "sms/changepwd.html")
        
        # Kiểm tra mật khẩu mới không được trống
        if not new_password or len(new_password) < 6:
            messages.error(request, "Mật khẩu mới phải có ít nhất 6 ký tự!")
            return render(request, "sms/changepwd.html")
        
        # Đổi mật khẩu
        user = request.user
        user.set_password(new_password)
        user.save()
        
        # Update session để không bị logout
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, user)
        
        messages.success(request, "Đổi mật khẩu thành công!")
        return redirect("loc_list")
    
    return render(request, "sms/changepwd.html")

@login_required
def bill_list_view(request, loc_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.info(request, "Vui lòng xem hóa đơn của bạn tại đây.")
            return redirect('renter_bills')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Chỉ xem hóa đơn của location thuộc chủ hiện tại
    location = get_object_or_404(Location, pk=loc_id, chu=request.user) if not request.user.is_superuser else get_object_or_404(Location, pk=loc_id)
    houses_in_loc = House.objects.filter(loc_id=location.id)
    bills = Hoadon.objects.filter(house__in=houses_in_loc).select_related('house', 'renter').order_by('-duedate', '-id')
    
    paginator = Paginator(bills, 20)
    page = request.GET.get('page')
    paged_bills = paginator.get_page(page)
    
    context = {
        'bills': paged_bills,
        'location': location,
        'loc_id': loc_id,
        # Nếu muốn hiển thị danh sách các House để tạo hóa đơn mới
    'houses': House.objects.filter(loc_id=location.id).order_by('ten')
    }
    return render(request, 'sms/bill_list.html', context)

@login_required
def create_bill_view(request, house_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền tạo hóa đơn.")
            return redirect('renter_dashboard')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    house = get_object_or_404(House, pk=house_id)
    # Ủy quyền: chỉ chủ của location chứa house mới được tạo hóa đơn
    if not request.user.is_superuser and house.loc.chu_id != request.user.id:
        return HttpResponseForbidden("Bạn không có quyền thao tác với nhà trọ này")
    loc_id = house.loc.id 
    
    # Lấy hóa đơn ngay trước hóa đơn đang tạo (Lấy Max ID)
    last_bill = Hoadon.objects.filter(house=house).order_by('-id').first()
    
    

    # Lấy renter đang thuê (active contract) để làm default
    active_contract = HouseRenter.objects.filter(house=house, active=True).select_related('renter').first()
    default_renter = active_contract.renter if active_contract else None
    
    # Lấy danh sách tất cả renter từng thuê nhà này (để hiển thị info)
    all_renters = HouseRenter.objects.filter(house=house).select_related('renter').values_list('renter__hoten', flat=True).distinct()

    if request.method == 'POST':
        form = CreateHoaDonForm(request.POST, request.FILES, house=house)
        if form.is_valid():
            
            # LƯU HÓA ĐƠN
            new_bill = form.save(commit=False)
            new_bill.house = house
            # Renter đã được chọn trong form (có thể là active hoặc renter khác từng thuê)
            
            new_bill.save() # Kích hoạt logic tự động tính TONG_CONG, CONG_NO

            # Thông báo cho renter (nếu có)
            try:
                if new_bill.renter and new_bill.renter.user:
                    from .models import Notification
                    hashid_converter = HashidConverter()
                    Notification.objects.create(
                        user=new_bill.renter.user,
                        title="Bạn có hóa đơn mới",
                        message=f"Hóa đơn {new_bill.ten} cho {house.ten} đã được tạo.",
                        type='bill_created',
                        target_url=f"/sms/bill/{hashid_converter.to_url(new_bill.id)}/detail/"
                    )
            except Exception:
                pass

            # Gửi email PDF hóa đơn cho renter (nếu có email)
            if new_bill.renter and new_bill.renter.email:
                try:
                    email_sent = send_bill_email_with_pdf(new_bill)
                    if email_sent:
                        messages.info(request, f"Đã gửi hóa đơn PDF qua email tới {new_bill.renter.email}")
                    else:
                        messages.warning(request, "Không thể gửi email hóa đơn. Vui lòng kiểm tra cấu hình email.")
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Lỗi khi gửi email hóa đơn: {e}")
                    messages.warning(request, "Có lỗi khi gửi email hóa đơn.")

            renter_info = f" cho {new_bill.renter.hoten}" if new_bill.renter else ""
            messages.success(request, f"Đã tạo Hóa đơn mới cho {house.ten}{renter_info} thành công!")
            
            return redirect('bill_list', loc_id=loc_id) 
        else:
            # Nếu Form không hợp lệ, giữ nguyên form để hiển thị lỗi
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Lỗi ở trường '{field}': {error}")

    else:
        # GET request (Điền sẵn thông tin)
        initial_data = {
            'house': house.id,
            'ten': f"Hóa đơn {house.ten} - {timezone.now().strftime('%m/%Y')}",
        }
        # Tự động chọn renter đang thuê (active) làm default
        if default_renter:
            initial_data['renter'] = default_renter.id
            
        form = CreateHoaDonForm(initial=initial_data, house=house)

    context = {
        'forms': form,
        'house': house,
        'loc_id': loc_id,
        'bill': None,  # Để template biết đang ở chế độ create
    }
    return render(request, 'sms/create_bill.html', context)


@login_required
def invoice_search(request):
    """Trang tìm kiếm Hóa đơn theo Location, House, ngày tạo, ngày đến hạn, trạng thái"""
    from django.db.models import Sum
    from django.utils.dateparse import parse_date
    
    # Chặn renter truy cập - chuyển hướng về trang hóa đơn của họ
    try:
        if request.user.is_renter:
            messages.info(request, "Vui lòng xem hóa đơn của bạn tại đây.")
            return redirect('renter_bills')
    except (AttributeError, Renter.DoesNotExist):
        pass

    qs = Hoadon.objects.select_related('house', 'house__loc', 'renter').all().order_by('-ngay_tao', '-id')
    if not request.user.is_superuser:
        qs = qs.filter(house__loc__chu=request.user)

    # --- Đọc tham số lọc ---
    loc_id = request.GET.get('loc') or ''
    house_id = request.GET.get('house') or ''
    status = request.GET.get('status') or ''
    created_from = request.GET.get('created_from') or ''
    created_to = request.GET.get('created_to') or ''
    due_from = request.GET.get('due_from') or ''
    due_to = request.GET.get('due_to') or ''

    # --- Áp dụng bộ lọc ---
    if loc_id:
        qs = qs.filter(house__loc_id=loc_id)
    if house_id:
        qs = qs.filter(house_id=house_id)
    if status:
        qs = qs.filter(status=status)

    # Ngày tạo là DateTimeField -> dùng __date để so sánh theo ngày
    if created_from:
        d = parse_date(created_from)
        if d:
            qs = qs.filter(ngay_tao__date__gte=d)
    if created_to:
        d = parse_date(created_to)
        if d:
            qs = qs.filter(ngay_tao__date__lte=d)

    # Ngày đến hạn là DateField
    if due_from:
        d = parse_date(due_from)
        if d:
            qs = qs.filter(duedate__gte=d)
    if due_to:
        d = parse_date(due_to)
        if d:
            qs = qs.filter(duedate__lte=d)

    # --- Tổng hợp số liệu trên tập đã lọc ---
    aggregates = qs.aggregate(
        total_amount=Sum('TONG_CONG'),
        total_paid=Sum('SO_TIEN_DA_TRA'),
        total_debt=Sum('CONG_NO'),
    )

    # --- Phân trang ---
    paginator = Paginator(qs, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    # --- Dữ liệu cho bộ lọc ---
    locations = Location.objects.order_by('id') if request.user.is_superuser else Location.objects.filter(chu=request.user).order_by('id')
    houses = (
        House.objects.filter(loc_id=loc_id).order_by('ten') if loc_id else House.objects.none()
    )
    status_choices = Hoadon._meta.get_field('status').choices

    context = {
        'page_obj': page_obj,
        'aggregates': aggregates,
        'locations': locations,
        'houses': houses,
        'status_choices': status_choices,
        # giữ lại giá trị filter
        'loc_id': str(loc_id),
        'house_id': str(house_id),
        'status_val': status,
        'created_from': created_from,
        'created_to': created_to,
        'due_from': due_from,
        'due_to': due_to,
        'params': request.GET.urlencode(),
    }

    return render(request, 'sms/bill_search.html', context)


@login_required
def invoice_export_excel(request):
    """Xuất Excel theo đúng bộ lọc hiện tại của trang tìm kiếm hóa đơn"""
    from django.utils.dateparse import parse_date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers

    qs = Hoadon.objects.select_related('house', 'house__loc').all().order_by('-ngay_tao', '-id')
    if not request.user.is_superuser:
        qs = qs.filter(house__loc__chu=request.user)

    # Đọc tham số lọc (giống invoice_search)
    loc_id = request.GET.get('loc') or ''
    house_id = request.GET.get('house') or ''
    status = request.GET.get('status') or ''
    created_from = request.GET.get('created_from') or ''
    created_to = request.GET.get('created_to') or ''
    due_from = request.GET.get('due_from') or ''
    due_to = request.GET.get('due_to') or ''

    if loc_id:
        qs = qs.filter(house__loc_id=loc_id)
    if house_id:
        qs = qs.filter(house_id=house_id)
    if status:
        qs = qs.filter(status=status)

    if created_from:
        d = parse_date(created_from)
        if d:
            qs = qs.filter(ngay_tao__date__gte=d)
    if created_to:
        d = parse_date(created_to)
        if d:
            qs = qs.filter(ngay_tao__date__lte=d)

    if due_from:
        d = parse_date(due_from)
        if d:
            qs = qs.filter(duedate__gte=d)
    if due_to:
        d = parse_date(due_to)
        if d:
            qs = qs.filter(duedate__lte=d)

    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'HoaDon'

    # Header
    headers = [
        'STT', 'Vị trí', 'Nhà trọ', 'Mô tả', 'Tổng cộng', 'Đã trả', 'Công nợ', 'Ngày tạo', 'Đến hạn', 'Trạng thái'
    ]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Dữ liệu
    row_index = 2
    for idx, bill in enumerate(qs, start=1):
        ws.cell(row=row_index, column=1, value=idx)
        ws.cell(row=row_index, column=2, value=getattr(getattr(bill.house, 'loc', None), 'diachi', ''))
        ws.cell(row=row_index, column=3, value=getattr(bill.house, 'ten', ''))
        ws.cell(row=row_index, column=4, value=bill.ten)

        c5 = ws.cell(row=row_index, column=5, value=bill.TONG_CONG or 0)
        c6 = ws.cell(row=row_index, column=6, value=bill.SO_TIEN_DA_TRA or 0)
        c7 = ws.cell(row=row_index, column=7, value=bill.CONG_NO or 0)
        for c in (c5, c6, c7):
            c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ws.cell(row=row_index, column=8, value=bill.ngay_tao.strftime('%d/%m/%Y %H:%M') if bill.ngay_tao else '')
        ws.cell(row=row_index, column=9, value=bill.duedate.strftime('%d/%m/%Y') if bill.duedate else '')
        ws.cell(row=row_index, column=10, value=dict(Hoadon._meta.get_field('status').choices).get(bill.status, bill.status))
        row_index += 1

    # Tự động căn chiều rộng đơn giản
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Response
    from datetime import datetime
    filename = f"hoa_don_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def api_houses_by_location(request, loc_id: int):
    """Trả về danh sách houses theo location (JSON) - chỉ chủ sở hữu xem được"""
    if not request.user.is_superuser:
        get_object_or_404(Location, pk=loc_id, chu=request.user)
    houses = House.objects.filter(loc_id=loc_id).order_by('ten').values('id', 'ten')
    return JsonResponse({'results': list(houses)})

@login_required
def update_bill_view(request, bill_id):
    # Chặn renter truy cập
    try:
        if request.user.is_renter:
            messages.error(request, "Bạn không có quyền sửa hóa đơn.")
            return redirect('renter_bills')
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    bill = get_object_or_404(Hoadon, pk=bill_id)
    house = bill.house
    if not request.user.is_superuser and house.loc.chu_id != request.user.id:
        return HttpResponseForbidden("Bạn không có quyền sửa hóa đơn này")
    loc_id = house.loc.id 

    # NOTE: Meter reading fields (chi_so_dien_*, chi_so_nuoc_*) were removed from the Hoadon model.
    # The previous logic attempted to access these legacy attributes causing AttributeError.
    # Update now only saves the direct cost components provided in the form.

    if request.method == 'POST':
        form = CreateHoaDonForm(request.POST, request.FILES, instance=bill, house=house)
        if form.is_valid():
            updated_bill = form.save(commit=False)
            # Ensure house stays consistent (defensive)
            updated_bill.house = house
            updated_bill.save()  # Model save() recalculates TONG_CONG, CONG_NO, status
            messages.success(request, f"Đã cập nhật Hóa đơn {bill.ten} thành công! Tổng tiền mới: {updated_bill.TONG_CONG:,} VNĐ.")
            return redirect('bill_list', loc_id=loc_id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Lỗi ở trường '{field}': {error}")
    else:
        form = CreateHoaDonForm(instance=bill, house=house)
        
    from sms.utils.id_encoder import encode_id
    bill_hashid = encode_id(bill.id) if bill else None
    context = {
        'forms': form,
        'house': house,
        'loc_id': loc_id,
        'bill': bill,
        'bill_hashid': bill_hashid,
    }
    return render(request, 'sms/create_bill.html', context)

@login_required
def bill_detail_view(request, bill_id):
    # Validate bill access
    hoadon, error_response = validate_bill_access(request, bill_id)
    if error_response:
        return error_response
    
    loc_id = hoadon.house.loc.id
    
    # Check if user is renter (read-only mode)
    is_renter = False
    try:
        is_renter = hoadon.renter and hoadon.renter.user_id == request.user.id
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    # Lấy lịch sử thanh toán (Chưa có form ThanhToan, nên tạm thời dùng None)
    lich_su_thanh_toan = Thanhtoan.objects.filter(hoadon=hoadon).order_by('-id')
    lich_su_file = UploadedFile.objects.filter(link_id=bill_id, type=2).order_by('-id')
    # Bình luận hóa đơn
    try:
        from .models import BillComment
        bill_comments = BillComment.objects.filter(bill=hoadon, parent=None).select_related('user').prefetch_related('replies')
    except Exception:
        bill_comments = []
    
    # Chọn template base phù hợp
    base_template = "layouts/renter_base.html" if is_renter else "layouts/rental_base.html"
    
    context = {
        'hoadon': hoadon,
        'loc_id': loc_id,
        'lich_su_thanh_toan': lich_su_thanh_toan,
        'lich_su_file': lich_su_file,
        'payment_form': CreateThanhToanForm(),
        'bill_file_form': CreateUploadFile(),
        'is_renter_view': is_renter,  # Flag to show read-only mode
        'base_template': base_template,  # Template base để extends
        # Bình luận
        'bill_comments': bill_comments,
        'bill_comment_form': BillCommentForm(),
        'is_landlord': hoadon.house.loc.chu_id == request.user.id,
    }
    return render(request, 'sms/bill_detail.html', context)

@login_required
def add_payment(request, bill_id):
    hoadon = get_object_or_404(Hoadon, id=bill_id)
    
    # Kiểm tra quyền: landlord, renter của hóa đơn, hoặc superuser
    is_landlord = hoadon.house.loc.chu_id == request.user.id
    is_renter = False
    try:
        is_renter = hoadon.renter and hoadon.renter.user_id == request.user.id
    except (AttributeError, Renter.DoesNotExist):
        pass
    
    if not request.user.is_superuser and not is_landlord and not is_renter:
        return HttpResponseForbidden("Bạn không có quyền thao tác với hóa đơn này")

    if hoadon.status == 'DaTT':
        messages.warning(request, "Hóa đơn này đã thanh toán đủ, không thể thêm thanh toán mới.")
        return redirect('bill_detail', bill_id=hoadon.id)

    payment_form = CreateThanhToanForm(request.POST)
    
    if payment_form.is_valid():
        
        # LẤY GIÁ TRỊ TIỀN THANH TOÁN (DECIMAL AN TOÀN) TỪ cleaned_data
        tientt = payment_form.cleaned_data['tientt'] 
        
        try:
            # --- LOGIC XÁC ĐỊNH VAI TRÒ VÀ TRẠNG THÁI BAN ĐẦU ---
            # Landlord hoặc superuser: thanh toán tự động được xác nhận
            # Renter: thanh toán cần chờ xác nhận
            if is_landlord or request.user.is_superuser:
                initial_status = 'Daxn' 
                success_msg = f"Đã ghi nhận & XÁC NHẬN thanh toán {tientt:,} VNĐ."
            else:
                initial_status = 'Choxn' 
                success_msg = f"Đã gửi thanh toán {tientt:,} VNĐ. CHỜ CHỦ NHÀ XÁC NHẬN." 
            # --- HẾT LOGIC XÁC ĐỊNH VAI TRÒ ---

            # 1. Kiểm tra logic nghiệp vụ: Công nợ
            if initial_status == 'Daxn' and tientt > hoadon.CONG_NO:
                messages.error(request, f"Số tiền thanh toán ({tientt:,} VNĐ) vượt quá Công nợ còn lại ({hoadon.CONG_NO:,} VNĐ).")
                return redirect('bill_detail', bill_id=hoadon.id)

            # 2. LƯU THANHTOAN VÀ GÁN hoadon
            payment = payment_form.save(commit=False)
            payment.hoadon = hoadon
            payment.user = request.user 
            payment.status = initial_status 

            payment.save() # Kích hoạt logic tự động cập nhật Hoadon

            # Tạo thông báo cho chủ nhà nếu renter gửi thanh toán (chờ xác nhận)
            try:
                if initial_status == 'Choxn':
                    landlord = hoadon.house.loc.chu
                    if landlord:
                        from .models import Notification
                        hashid_converter = HashidConverter()
                        Notification.objects.create(
                            user=landlord,
                            title="Khách thuê gửi thanh toán",
                            message=f"{request.user.get_username()} đã gửi thanh toán {tientt:,} VNĐ cho hóa đơn #{hoadon.id}",
                            type='payment_submitted',
                            target_url=f"/sms/bill/{hashid_converter.to_url(hoadon.id)}/detail/"
                        )
            except Exception:
                pass

            # 3. Thông báo thành công (SỬ DỤNG tientt AN TOÀN)
            messages.success(request, success_msg)
            
            if hoadon.status == 'DaTT':
                messages.info(request, "Hóa đơn đã được thanh toán đủ.")
                
            return redirect('bill_detail', bill_id=hoadon.id)
            
        except Exception as e:
            # Xử lý lỗi trong quá trình lưu/tính toán
            messages.error(request, f"Lỗi xảy ra khi xử lý thanh toán: {e}")
            return redirect('bill_detail', bill_id=hoadon.id)
    else:
        # Nếu Form không hợp lệ, chuyển hướng về trang chi tiết và hiển thị lỗi
        for field, errors in payment_form.errors.items():
            for error in errors:
                label = payment_form.fields[field].label
                messages.error(request, f"Lỗi trường **{label}**: {error}")
                
        return redirect('bill_detail', bill_id=hoadon.id)        

# app_quanly/views.py (Thêm vào cuối file)

@login_required
def confirm_payment_view(request, payment_id):
    # Validate payment access
    payment, error_response = validate_payment_access(request, payment_id)
    if error_response:
        return error_response
    
    hoadon = payment.hoadon
    
    # Chỉ landlord hoặc superuser mới được xác nhận thanh toán
    is_landlord = hoadon.house.loc.chu_id == request.user.id
    if not request.user.is_superuser and not is_landlord:
        messages.error(request, "Chỉ chủ nhà mới có quyền xác nhận thanh toán.")
        return redirect('bill_detail', bill_id=hoadon.id)
    
    if payment.status == 'Choxn':
        payment.status = 'Daxn'
        payment.save() # Kích hoạt Thanhtoan.save() -> cập nhật Hoadon
        messages.success(request, f"Đã XÁC NHẬN thanh toán {payment.tientt:,} VNĐ.")
    elif payment.status == 'Daxn':
        messages.info(request, "Thanh toán này đã được xác nhận trước đó.")
    else:
        messages.error(request, "Bạn không có quyền hoặc thanh toán không hợp lệ để xác nhận.")
        
    return redirect('bill_detail', bill_id=hoadon.id)

@login_required
def delete_payment_view(request, payment_id):
    # Validate payment access
    payment, error_response = validate_payment_access(request, payment_id)
    if error_response:
        return error_response
    
    hoadon = payment.hoadon
    
    # Chỉ landlord hoặc superuser mới được xóa thanh toán
    is_landlord = hoadon.house.loc.chu_id == request.user.id
    
    if is_landlord or request.user.is_superuser:
        payment.delete() # Kích hoạt Thanhtoan.delete() -> cập nhật Hoadon
        messages.success(request, f"Đã xóa bản ghi thanh toán {payment.tientt:,} VNĐ.")
    else:
        messages.error(request, "Bạn không có quyền xóa bản ghi thanh toán này.")

    return redirect('bill_detail', bill_id=hoadon.id)
    
# app_quanly/views.py (Thêm vào cuối file)


# app_quanly/views.py (Thêm hoặc thay thế hàm generate_bill_pdf)
# app_quanly/views.py (Chỉnh sửa hàm generate_bill_pdf)

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont 
from decimal import Decimal
import os
from django.conf import settings 
from django.core.mail import EmailMessage
# Đảm bảo đã import các Models cần thiết
from .models import Hoadon, HouseRenter


def send_bill_email_with_pdf(hoadon):
    """
    Gửi email hóa đơn với PDF đính kèm cho renter
    
    Args:
        hoadon: Đối tượng Hoadon cần gửi email
    
    Returns:
        bool: True nếu gửi thành công, False nếu thất bại
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Kiểm tra renter có email không
    if not hoadon.renter or not hoadon.renter.email:
        logger.info(f"Bill {hoadon.id}: Renter không có email, bỏ qua gửi mail")
        return False
    
    try:
        from reportlab.pdfbase.pdfmetrics import registerFont
        from reportlab.pdfbase.ttfonts import TTFont
        registerFont(TTFont('Arial','ARIAL.ttf'))
        VIETNAMESE_FONT = 'Arial'

        # Tạo PDF trong memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                leftMargin=1.5*cm, rightMargin=1.5*cm, 
                                topMargin=2*cm, bottomMargin=2*cm)
        Story = []

        # Lấy thông tin renter
        renter = hoadon.renter
        
        # Định dạng số
        def f(number):
            return "{:,.0f}".format(Decimal(number or 0)).replace(",", ",")

        # Styles
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Arial', fontName=VIETNAMESE_FONT, fontSize=10, leading=12))
        normal_style = styles['Arial']

        # Tiêu đề
        p_title = Paragraph(f'<font name="{VIETNAMESE_FONT}">HÓA ĐƠN THUÊ NHÀ - KỲ {hoadon.duedate.strftime("%m/%Y")}</font>', normal_style)
        Story.append(p_title)
        Story.append(Spacer(0, 0.5*cm))

        # Thông tin chung
        Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Nhà trọ: {hoadon.house.ten} - {hoadon.house.loc.diachi}</font>', normal_style))
        if renter:
            Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Người thuê: {renter.hoten} - SĐT: {renter.sdt}</font>', normal_style))
        Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Ngày tạo: {hoadon.ngay_tao.strftime("%d/%m/%Y")}</font>', normal_style))
        Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Ngày hết hạn: {hoadon.duedate.strftime("%d/%m/%Y")}</font>', normal_style))
        Story.append(Spacer(0, 1*cm))

        # Bảng chi tiết
        def vn_text(text):
            return Paragraph(f'<font name="{VIETNAMESE_FONT}">{text}</font>', normal_style)

        table_data = [
            [vn_text('STT'), vn_text('Mục'), vn_text('Đơn vị'), vn_text('Số lượng'), vn_text('Đơn giá'), vn_text('Thành tiền (VNĐ)')],
            [1, vn_text('Tiền thuê nhà'), 'Tháng', 1, f(hoadon.house.permonth), f(hoadon.tienthuenha)],
            [2, vn_text('Tiền Điện'), 'kWh', '', '', f(hoadon.tiendien)],
            [3, vn_text('Tiền Nước'), 'm³', '', '', f(hoadon.tiennuoc)],
            [4, vn_text('Chi phí khác'), 'Lần', 1, f(hoadon.tienkhac), f(hoadon.tienkhac)],
            ['', '', '', '', vn_text('TỔNG CỘNG'), f(hoadon.TONG_CONG)]
        ]
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), VIETNAMESE_FONT), 
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'), 
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (4, -1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (4, -1), (-1, -1), colors.yellow),
        ])
        
        t = Table(table_data, colWidths=[0.8*cm, 4*cm, 1.5*cm, 2*cm, 2*cm, 3*cm])
        t.setStyle(table_style)
        Story.append(t)
        Story.append(Spacer(0, 0.5*cm))

        # Tổng kết
        summary_data = [
            [vn_text('TỔNG CỘNG PHẢI THU'), f(hoadon.TONG_CONG)],
            [vn_text('Số tiền đã thanh toán'), f(hoadon.SO_TIEN_DA_TRA)],
            [vn_text('CÔNG NỢ CÒN LẠI'), f(hoadon.CONG_NO)],
        ]
        summary_table = Table(summary_data, colWidths=[5*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), VIETNAMESE_FONT),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.9, 1)),
            ('TEXTCOLOR', (1, -1), (1, -1), colors.red if hoadon.CONG_NO > 0 else colors.green),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        Story.append(summary_table)

        # QR Code (nếu có)
        if hoadon.qr_code_image:
            try:
                from reportlab.platypus import Image as RLImage
                Story.append(Spacer(0, 0.5*cm))
                Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">QUÉT MÃ ĐỂ THANH TOÁN</font>', normal_style))
                Story.append(Spacer(0, 0.3*cm))
                
                qr_path = hoadon.qr_code_image.path
                if os.path.exists(qr_path):
                    qr_img = RLImage(qr_path, width=4*cm, height=4*cm)
                    Story.append(qr_img)
                    Story.append(Spacer(0, 0.2*cm))
                    Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}"><i>Quét mã QR để chuyển khoản trực tiếp đến tài khoản chủ nhà</i></font>', normal_style))
            except Exception as e:
                logger.warning(f"Không thể thêm QR code vào email PDF: {e}")

        # Chữ ký
        Story.append(Spacer(0, 1*cm))
        Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}"><i>Hóa đơn được tạo tự động bởi hệ thống.</i></font>', normal_style))
        
        # Build PDF
        doc.build(Story)
        pdf_content = buffer.getvalue()
        buffer.close()

        # Tạo email
        subject = f'Hóa đơn {hoadon.house.ten} - Tháng {hoadon.duedate.strftime("%m/%Y")}'
        
        # Nội dung email
        message = f"""
Kính gửi {renter.hoten},

Bạn có hóa đơn mới cho nhà trọ {hoadon.house.ten}.

Chi tiết hóa đơn:
- Kỳ thanh toán: {hoadon.duedate.strftime("%m/%Y")}
- Tổng tiền: {f(hoadon.TONG_CONG)} VNĐ
- Đã thanh toán: {f(hoadon.SO_TIEN_DA_TRA)} VNĐ
- Còn nợ: {f(hoadon.CONG_NO)} VNĐ
- Hạn thanh toán: {hoadon.duedate.strftime("%d/%m/%Y")}

Vui lòng kiểm tra file PDF đính kèm để xem chi tiết.

Trân trọng,
Ban quản lý
"""
        
        if getattr(settings, 'SEND_OWNER_EMAIL_NOTIFICATIONS', True):
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.EMAIL_HOST_USER,
                to=[hoadon.renter.email],
            )
            # Đính kèm PDF
            pdf_filename = f'HoaDon_{hoadon.house.ten}_{hoadon.duedate.strftime("%Y%m")}.pdf'
            email.attach(pdf_filename, pdf_content, 'application/pdf')
            # Gửi email
            email.send(fail_silently=False)
            logger.info(f"Đã gửi email hóa đơn {hoadon.id} tới {hoadon.renter.email}")
        else:
            logger.info(f"Không gửi email hóa đơn {hoadon.id} do cấu hình tắt gửi mail")
        return True
        
    except Exception as e:
        logger.error(f"Lỗi khi gửi email hóa đơn {hoadon.id}: {str(e)}")
        return False 


@login_required
def generate_bill_pdf(request, bill_id):

    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.pdfbase.ttfonts import TTFont
    registerFont(TTFont('Arial','ARIAL.ttf'))

    VIETNAMESE_FONT = 'Arial'

    hoadon = get_object_or_404(Hoadon, pk=bill_id)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                            leftMargin=1.5*cm, rightMargin=1.5*cm, 
                            topMargin=2*cm, bottomMargin=2*cm)
    Story = []

    # --- LẤY THÔNG TIN NGƯỜI THUÊ ---
    # Ưu tiên lấy từ hoadon.renter (người được tạo hóa đơn cho)
    # Nếu không có thì tìm active contract
    renter = hoadon.renter
    if not renter:
        active_contract = HouseRenter.objects.filter(
            house=hoadon.house, 
            active=True
        ).select_related('renter').first()
        renter = active_contract.renter if active_contract else None
    
    # --- ĐỊNH DẠNG SỐ VÀ CHUẨN BỊ DỮ LIỆU ---
    def f(number):
        return "{:,.0f}".format(Decimal(number or 0)).replace(",", ",")

    # --- ÁP DỤNG FONT CHO STYLES ---
    styles = getSampleStyleSheet()
    
    # Style thường
    styles.add(ParagraphStyle(name='Arial', fontName=VIETNAMESE_FONT, fontSize=10, leading=12))
    
    # Style Heading
 #   styles.add(ParagraphStyle(name='Arial', fontName=VIETNAMESE_FONT, fontSize=16, alignment=1, spaceAfter=12))
    
    normal_style = styles['Arial']
#    heading_style = styles['Arial']


    # --- 1. TIÊU ĐỀ (ĐÃ XÓA THẺ B) ---
    p_title = Paragraph(f'<font name="{VIETNAMESE_FONT}">HÓA ĐƠN THUÊ NHÀ - KỲ {hoadon.duedate.strftime("%m/%Y")}</font>', normal_style)
    Story.append(p_title)
    Story.append(Spacer(0, 0.5*cm))

    # --- 2. THÔNG TIN CHUNG (ĐÃ XÓA THẺ B) ---
    Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Nhà trọ: {hoadon.house.ten} - {hoadon.house.loc.diachi}</font>', normal_style))
    if renter:
        Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Người thuê: {renter.hoten} - SĐT: {renter.sdt}</font>', normal_style))
    else:
        Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Người thuê: Đang trống</font>', normal_style))
    Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Ngày tạo: {hoadon.ngay_tao.strftime("%d/%m/%Y")}</font>', normal_style))
    Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">Ngày hết hạn: {hoadon.duedate.strftime("%d/%m/%Y")}</font>', normal_style))
    Story.append(Spacer(0, 1*cm))

    # --- 3. BẢNG CHI TIẾT DỊCH VỤ ---
    
    # Hàm bọc text trong font (chỉ font thường)
    def vn_text(text):
        return Paragraph(f'<font name="{VIETNAMESE_FONT}">{text}</font>', normal_style)

    table_data = [
        # ĐÃ XÓA THẺ B (Bold)
        [vn_text('STT'), vn_text('Mục'), vn_text('Đơn vị'), vn_text('Số lượng'), vn_text('Đơn giá'), vn_text('Thành tiền (VNĐ)')],
        
        [1, vn_text('Tiền thuê nhà'), 'Tháng', 1, f(hoadon.house.permonth), f(hoadon.tienthuenha)],
        
        [2, vn_text('Tiền Điện'), 'kWh', 
         '', 
         '', 
         f(hoadon.tiendien)],
         
        [3, vn_text('Tiền Nước'), 'm³', 
         '', 
         '', 
         f(hoadon.tiennuoc)],
         
        [4, vn_text('Chi phí khác'), 'Lần', 1, f(hoadon.tienkhac), f(hoadon.tienkhac)],
        
        # Dòng tổng cộng 
        ['', '', '', '', vn_text('TỔNG CỘNG'), f(hoadon.TONG_CONG)]
    ]
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # ÁP DỤNG FONT CHO TOÀN BỘ BẢNG
        ('FONTNAME', (0, 0), (-1, -1), VIETNAMESE_FONT), 
        
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'), 
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (4, -1), (-1, -1), 'RIGHT'),
        ('BACKGROUND', (4, -1), (-1, -1), colors.yellow),
    ])
    
    t = Table(table_data, colWidths=[0.8*cm, 4*cm, 1.5*cm, 2*cm, 2*cm, 3*cm])
    t.setStyle(table_style)
    Story.append(t)
    Story.append(Spacer(0, 0.5*cm))

    # --- 4. TỔNG KẾT VÀ CÔNG NỢ ---
    summary_data = [
        [vn_text('TỔNG CỘNG PHẢI THU'), f(hoadon.TONG_CONG)],
        [vn_text('Số tiền đã thanh toán'), f(hoadon.SO_TIEN_DA_TRA)],
        [vn_text('CÔNG NỢ CÒN LẠI'), f(hoadon.CONG_NO)],
    ]
    summary_table = Table(summary_data, colWidths=[5*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), VIETNAMESE_FONT),
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.9, 1)),
        ('TEXTCOLOR', (1, -1), (1, -1), colors.red if hoadon.CONG_NO > 0 else colors.green),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    Story.append(summary_table)

    # --- QR CODE (nếu có) ---
    if hoadon.qr_code_image:
        try:
            from reportlab.platypus import Image as RLImage
            import os
            
            Story.append(Spacer(0, 0.5*cm))
            Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}">QUÉT MÃ ĐỂ THANH TOÁN</font>', normal_style))
            Story.append(Spacer(0, 0.3*cm))
            
            # Get absolute path to QR code image
            qr_path = hoadon.qr_code_image.path
            if os.path.exists(qr_path):
                qr_img = RLImage(qr_path, width=4*cm, height=4*cm)
                Story.append(qr_img)
                Story.append(Spacer(0, 0.2*cm))
                Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}"><i>Quét mã QR để chuyển khoản trực tiếp đến tài khoản chủ nhà</i></font>', normal_style))
        except Exception as e:
            # Log error but continue generating PDF
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error adding QR code to PDF: {e}")

    # --- Chữ ký và Ghi chú ---
    Story.append(Spacer(0, 1*cm))
    Story.append(Paragraph(f'<font name="{VIETNAMESE_FONT}"><i>Hóa đơn được tạo tự động bởi hệ thống.</i></font>', normal_style))
    
    # --- Xây dựng PDF ---
    doc.build(Story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="HoaDon_{hoadon.house.ten}_{hoadon.duedate.strftime("%Y%m")}.pdf"'
    response.write(pdf)
    return response

# Đảm bảo các thư viện này đã được import
import pandas as pd # Cần thiết cho việc xử lý dữ liệu
import xlsxwriter # Cần thiết cho việc định dạng chuyên sâu
from io import BytesIO # Cần thiết để tạo file trong bộ nhớ
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from decimal import Decimal # Cần thiết cho việc định dạng số
# ... các models Hoadon, HouseRenter

@login_required
def generate_bill_excel(request, bill_id):
    hoadon = get_object_or_404(Hoadon, pk=bill_id)
    
    # --- 1. Chuẩn bị dữ liệu ---
    # TÌM NGƯỜI THUÊ HIỆN TẠI (Lấy logic từ hàm cũ)
    active_contract = HouseRenter.objects.filter(
        house=hoadon.house, 
        active=True
    ).select_related('renter').first()
    renter = active_contract.renter if active_contract else None

    # Dữ liệu chi tiết bảng dịch vụ
    detail_data = [
        [1, 'Tiền thuê nhà', 'Tháng', 1, hoadon.house.permonth, hoadon.tienthuenha],
        [2, 'Tiền Điện', 'kWh', '', '', hoadon.tiendien], # Cần đảm bảo DON_GIA_DIEN đã được định nghĩa
        [3, 'Tiền Nước', 'm³', '', '', hoadon.tiennuoc], # Cần đảm bảo DON_GIA_NUOC đã được định nghĩa
        [4, 'Chi phí khác', 'Lần', 1, hoadon.tienkhac, hoadon.tienkhac],
    ]
    
    # --- 2. Tạo Workbook và Worksheet ---
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet(f'HoaDon_{hoadon.duedate.strftime("%m-%Y")}')

    # --- 3. Định dạng Styles trong Excel (Giao diện đẹp) ---
    currency_format = workbook.add_format({'num_format': '#,##0', 'align': 'right', 'font_name': 'Arial'})
    bold_format = workbook.add_format({'bold': True, 'font_name': 'Arial'})
    center_bold_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_name': 'Arial'})
    header_format = workbook.add_format({
        'bold': True, 
        'font_size': 16, 
        'align': 'center', 
        'valign': 'vcenter', 
        'font_name': 'Arial'
    })
    table_header_format = workbook.add_format({
        'bold': True, 
        'bg_color': '#D9E1F2', 
        'border': 1, 
        'align': 'center', 
        'valign': 'vcenter', 
        'font_name': 'Arial'
    })
    table_cell_format = workbook.add_format({'border': 1, 'font_name': 'Arial'})
    table_currency_format = workbook.add_format({'border': 1, 'num_format': '#,##0', 'align': 'right', 'font_name': 'Arial'})
    summary_label_format = workbook.add_format({'bold': True, 'bg_color': '#FCE4D6', 'border': 1, 'font_name': 'Arial'})
    summary_value_format = workbook.add_format({'bold': True, 'bg_color': '#FCE4D6', 'num_format': '#,##0', 'align': 'right', 'border': 1, 'font_name': 'Arial'})
    debt_format = workbook.add_format({'bold': True, 'bg_color': '#FCE4D6', 'num_format': '#,##0', 'align': 'right', 'border': 1, 'font_color': 'red', 'font_name': 'Arial'})
    
    # --- 4. Thiết lập Cột và Kích thước ---
    worksheet.set_column('A:A', 15) # STT
    worksheet.set_column('B:B', 25) # Mục
    worksheet.set_column('C:D', 10) # Đơn vị, Số lượng
    worksheet.set_column('E:G', 15) # Đơn giá, Thành tiền, Tóm tắt

    # --- 5. Viết Tiêu đề Hóa đơn (Gộp ô) ---
    worksheet.merge_range('A1:F1', f'HÓA ĐƠN THUÊ NHÀ - KỲ {hoadon.duedate.strftime("%m/%Y")}', header_format)
    worksheet.merge_range('A2:F2', 'CHI TIẾT THANH TOÁN', center_bold_format)

    # --- 6. Viết Thông tin Chung ---
    ROW_START = 4
    worksheet.write(ROW_START, 0, 'Nhà trọ:', bold_format)
    worksheet.write(ROW_START, 1, f'{hoadon.house.ten} - {hoadon.house.loc.diachi}')
    worksheet.write(ROW_START + 1, 0, 'Người thuê:', bold_format)
    worksheet.write(ROW_START + 1, 1, renter.hoten if renter else "Đang trống")
    worksheet.write(ROW_START + 2, 0, 'Ngày tạo:', bold_format)
    worksheet.write(ROW_START + 2, 1, hoadon.ngay_tao.strftime("%d/%m/%Y"))
    worksheet.write(ROW_START + 3, 0, 'Ngày hết hạn:', bold_format)
    worksheet.write(ROW_START + 3, 1, hoadon.duedate.strftime("%d/%m/%Y"))

    # --- 7. Viết Bảng Chi tiết Dịch vụ ---
    TABLE_ROW_START = ROW_START + 5 # Bắt đầu từ dòng 9

    # Viết Header Bảng
    headers = ['STT', 'Mục', 'Đơn vị', 'Số lượng', 'Đơn giá (VNĐ)', 'Thành tiền (VNĐ)']
    worksheet.write_row(TABLE_ROW_START, 0, headers, table_header_format)
    
    # Viết Dữ liệu chi tiết
    row_num = TABLE_ROW_START + 1
    for row in detail_data:
        # Cột STT, Mục, Đơn vị
        worksheet.write(row_num, 0, row[0], table_cell_format)
        worksheet.write(row_num, 1, row[1], table_cell_format)
        worksheet.write(row_num, 2, row[2], table_cell_format)
        # Cột Số lượng (Số)
        worksheet.write(row_num, 3, row[3], table_currency_format)
        # Cột Đơn giá (Tiền)
        worksheet.write(row_num, 4, row[4], table_currency_format)
        # Cột Thành tiền (Tiền)
        worksheet.write(row_num, 5, row[5], table_currency_format)
        row_num += 1

    # Dòng Tổng cộng bảng
    total_row_num = row_num
    worksheet.merge_range(total_row_num, 0, total_row_num, 4, 'TỔNG CỘNG PHẢI THU', summary_label_format)
    worksheet.write(total_row_num, 5, hoadon.TONG_CONG, summary_value_format)
    
    # --- 8. Viết Tóm tắt Công nợ ---
    
    # Dòng Đã thanh toán
    paid_row_num = total_row_num + 1
    worksheet.merge_range(paid_row_num, 0, paid_row_num, 4, 'Số tiền đã thanh toán', summary_label_format)
    worksheet.write(paid_row_num, 5, hoadon.SO_TIEN_DA_TRA, summary_value_format)

    # Dòng Công nợ còn lại (Định dạng màu Đỏ nếu > 0)
    debt_row_num = total_row_num + 2
    debt_style = debt_format if hoadon.CONG_NO > 0 else summary_value_format
    worksheet.merge_range(debt_row_num, 0, debt_row_num, 4, 'CÔNG NỢ CÒN LẠI', summary_label_format)
    worksheet.write(debt_row_num, 5, hoadon.CONG_NO, debt_style)
    
    # --- 9. Kết thúc và Trả về Response ---
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="HoaDon_{hoadon.house.ten}_{hoadon.duedate.strftime("%Y%m")}.xlsx"'
        },
    )
    return response

# ==========================
# Notifications & Issues Views
# ==========================
from django import forms

class IssueReportForm(forms.ModelForm):
    images = forms.FileField(
        required=False,
        widget=forms.ClearableFileInput(attrs={'multiple': True, 'accept': 'image/*'}),
        label='Hình ảnh'
    )
    
    class Meta:
        model = IssueReport
        fields = ['house', 'title', 'description']
        widgets = {
            'description': forms.Textarea(attrs={'rows':4})
        }

    def __init__(self, *args, **kwargs):
        renter = kwargs.pop('renter', None)
        super().__init__(*args, **kwargs)
        if renter is not None:
            # Chỉ cho chọn house có liên quan tới renter (qua hợp đồng hoặc hóa đơn)
            houses_qs = House.objects.filter(hoadon__renter=renter).distinct()
            self.fields['house'].queryset = houses_qs


class IssueCommentForm(forms.ModelForm):
    class Meta:
        model = IssueComment
        fields = ['comment']
        widgets = {
            'comment': forms.Textarea(attrs={'rows': 2, 'placeholder': 'Viết bình luận...', 'class': 'form-control'})
        }

class BillCommentForm(forms.ModelForm):
    class Meta:
        from .models import BillComment
        model = BillComment
        fields = ['comment']
        widgets = {
            'comment': forms.Textarea(attrs={'rows': 2, 'placeholder': 'Viết bình luận về hóa đơn...', 'class': 'form-control'})
        }


@login_required
def notifications_list(request):
    from .models import Notification
    items = Notification.objects.filter(user=request.user).order_by('-created_at')
    # Determine base template by role
    is_landlord = getattr(request.user, 'is_landlord', False)
    base_template = 'layouts/rental_base.html' if is_landlord else 'layouts/renter_base.html'
    return render(request, 'sms/notifications_list.html', { 'items': items, 'base_template': base_template })


@login_required
def mark_notification_read(request, notification_id):
    """Mark a notification as read and redirect to target"""
    from .models import Notification
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        
        # If modal parameter, just return success (called via AJAX)
        if request.GET.get('modal') == '1':
            return HttpResponse('OK')
        
        # Otherwise redirect to target URL or back to notifications
        if notification.target_url:
            return redirect(notification.target_url)
    except Notification.DoesNotExist:
        pass
    
    return redirect('notifications_list')


@login_required
def add_bill_comment(request, bill_id):
    """Thêm bình luận cho Hóa đơn (chủ nhà hoặc khách thuê liên quan)"""
    # Validate bill access using the helper that handles both int and hashid
    hoadon, error_response = validate_bill_access(request, bill_id)
    if error_response:
        return error_response

    is_landlord = hoadon.house.loc.chu_id == request.user.id
    is_renter = False
    try:
        is_renter = hoadon.renter and hoadon.renter.user_id == request.user.id
    except (AttributeError, Renter.DoesNotExist):
        pass

    if request.method == 'POST':
        form = BillCommentForm(request.POST)
        if form.is_valid():
            from .models import BillComment, Notification
            c = form.save(commit=False)
            c.bill = hoadon
            c.user = request.user
            parent_id = request.POST.get('parent_id')
            if parent_id:
                c.parent_id = parent_id
            c.save()

            # Gửi thông báo cho bên còn lại
            try:
                recipient = None
                if is_renter:
                    recipient = hoadon.house.loc.chu
                else:
                    recipient = hoadon.renter.user if hoadon.renter and hoadon.renter.user else None
                if recipient:
                    from django.urls import reverse
                    hashid_converter = HashidConverter()
                    Notification.objects.create(
                        user=recipient,
                        title=f"Bình luận mới về Hóa đơn #{hoadon.id}",
                        message=f"{request.user.username}: {c.comment[:100]}",
                        type='bill_commented',
                        target_url=reverse('bill_detail', args=[{hashid_converter.to_url(hoadon.id)}])
                    )
            except Exception:
                pass

            # HTMX: trả về HTML của comment mới
            if request.headers.get('HX-Request'):
                return render(request, 'sms/partials/comment_item.html', {
                    'comment': c,
                    'is_landlord': is_landlord,
                })

    return redirect('bill_detail', bill_id=hoadon.id)


@login_required
def renter_report_issue(request):
    renter = Renter.objects.filter(user=request.user).first()
    if renter is None:
        messages.error(request, "Chỉ khách thuê mới được ghi nhận sự cố.")
        return redirect('renter_dashboard')

    if request.method == 'POST':
        form = IssueReportForm(request.POST, request.FILES, renter=renter)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.renter = renter
            issue.status = 'new'
            issue.save()
            
            # Handle multiple image uploads
            images = request.FILES.getlist('images')
            for image in images:
                IssueImage.objects.create(
                    issue=issue,
                    image=image,
                    uploaded_by=request.user
                )

            # Notify landlord via notification
            try:
                landlord = issue.house.loc.chu
                if landlord:
                    from django.urls import reverse
                    from .models import Notification
                    Notification.objects.create(
                        user=landlord,
                        title="Sự cố mới được báo",
                        message=f"{renter.hoten} báo sự cố: {issue.title} tại {issue.house.ten}",
                        type='issue_reported',
                        target_url=reverse('issue_detail', kwargs={'issue_id': issue.id})
                    )
                    
                    # Send email notification
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        if getattr(settings, 'SEND_OWNER_EMAIL_NOTIFICATIONS', True):
                            send_mail(
                            subject=f'Sự cố mới: {issue.title}',
                            message=f'{renter.hoten} đã báo sự cố tại {issue.house.ten}:\n\n{issue.description}',
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[landlord.email],
                            fail_silently=True,
                        )
                    except Exception:
                        pass
            except Exception:
                pass

            # HTMX request: return new table row + trigger modal close
            if request.headers.get('HX-Request'):
                response = render(request, 'sms/partials/issue_row.html', {'i': issue})
                response['HX-Trigger'] = 'closeModal, issueCreated'
                return response

            messages.success(request, "Đã ghi nhận sự cố. Chủ nhà sẽ xử lý sớm nhất.")
            return redirect('renter_issues')
    else:
        form = IssueReportForm(renter=renter)

    # If HTMX, return modal content only
    if request.headers.get('HX-Request'):
        return render(request, 'sms/partials/renter_issue_form_modal.html', { 'form': form })

    base_template = 'layouts/renter_base.html'
    return render(request, 'sms/renter_issue_form.html', { 'form': form, 'base_template': base_template })


@login_required
def issue_list(request):
    """Landlord's issue list - landlords only"""
    # Check if user is a landlord (owns at least one location)
    from .models import Location
    if request.user.is_renter:
        messages.error(request, "Chỉ chủ nhà mới có quyền truy cập trang này.")
        return redirect('index')
    
    issues = IssueReport.objects.filter(house__loc__chu=request.user).select_related('house','renter').order_by('-created_at')
    return render(request, 'sms/issue_list.html', { 'issues': issues })


@login_required
def change_issue_status(request, issue_id):
    """Change status of an issue and track history"""
    issue = get_object_or_404(IssueReport, id=issue_id, house__loc__chu=request.user)
    
    new_status = request.POST.get('status', 'resolved')
    old_status = issue.status
    
    if new_status != old_status:
        issue.status = new_status
        if new_status == 'resolved':
            issue.resolved_at = timezone.now()
        issue.save()
        
        # Track status history
        IssueStatusHistory.objects.create(
            issue=issue,
            old_status=old_status,
            new_status=new_status,
            changed_by=request.user,
            note=request.POST.get('note', '')
        )
        
        # Notify renter
        try:
            if issue.renter and issue.renter.user:
                from django.urls import reverse
                status_text = dict(IssueReport.STATUS_CHOICES).get(new_status, new_status)
                Notification.objects.create(
                    user=issue.renter.user,
                    title=f"Sự cố đã chuyển trạng thái: {status_text}",
                    message=f"Sự cố '{issue.title}' tại {issue.house.ten} đã chuyển sang: {status_text}.",
                    type='issue_resolved',
                    target_url=reverse('issue_detail', kwargs={'issue_id': issue.id})
                )
                
                # Send email
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    if getattr(settings, 'SEND_OWNER_EMAIL_NOTIFICATIONS', True):
                        send_mail(
                        subject=f'Cập nhật sự cố: {issue.title}',
                        message=f'Sự cố tại {issue.house.ten} đã chuyển sang trạng thái: {status_text}.\n\n{issue.description}',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[issue.renter.user.email],
                        fail_silently=True,
                    )
                except Exception:
                    pass
        except Exception:
            pass
        
        messages.success(request, f"Đã cập nhật trạng thái thành '{status_text}'.")
    
    # HTMX: refresh detail modal
    if request.headers.get('HX-Request'):
        return redirect('issue_detail', issue_id=issue_id)
    
    return redirect('issue_list')


@login_required
def resolve_issue(request, issue_id):
    """Quick resolve action - marks as pending confirmation"""
    issue = get_object_or_404(IssueReport, id=issue_id, house__loc__chu=request.user)
    old_status = issue.status
    
    issue.status = 'pending_confirmation'
    issue.save()
    
    # Track status history
    IssueStatusHistory.objects.create(
        issue=issue,
        old_status=old_status,
        new_status='pending_confirmation',
        changed_by=request.user,
        note='Chủ nhà đã đánh dấu hoàn thành, chờ khách thuê xác nhận'
    )

    # Notify renter to confirm
    try:
        if issue.renter and issue.renter.user:
            from django.urls import reverse
            Notification.objects.create(
                user=issue.renter.user,
                title="Vui lòng xác nhận sự cố đã xử lý",
                message=f"Chủ nhà đã đánh dấu sự cố '{issue.title}' tại {issue.house.ten} là đã xử lý. Vui lòng kiểm tra và xác nhận.",
                type='issue_resolved',
                target_url=reverse('issue_detail', kwargs={'issue_id': issue.id})
            )
            
            # Send email
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                if getattr(settings, 'SEND_OWNER_EMAIL_NOTIFICATIONS', True):
                    send_mail(
                    subject=f'Xác nhận sự cố đã xử lý: {issue.title}',
                    message=f'Chủ nhà đã đánh dấu sự cố tại {issue.house.ten} là đã xử lý.\n\nVui lòng vào hệ thống để kiểm tra và xác nhận.\n\n{issue.description}',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[issue.renter.user.email],
                    fail_silently=True,
                )
            except Exception:
                pass
    except Exception:
        pass

    messages.success(request, "Đã đánh dấu hoàn thành. Chờ khách thuê xác nhận.")
    return redirect('issue_list')


@login_required
def renter_confirm_issue(request, issue_id):
    """Renter confirms issue is resolved"""
    issue = get_object_or_404(IssueReport, id=issue_id)
    
    # Check renter owns this issue
    if not (issue.renter and issue.renter.user == request.user):
        return HttpResponse("Unauthorized", status=403)
    
    if issue.status != 'pending_confirmation':
        messages.warning(request, "Sự cố này không cần xác nhận.")
        # Return updated issue detail for HTMX
        return issue_detail(request, issue_id)
    
    action = request.POST.get('action')
    old_status = issue.status
    
    if action == 'confirm':
        issue.status = 'resolved'
        issue.resolved_at = timezone.now()
        issue.save()
        
        # Track history
        IssueStatusHistory.objects.create(
            issue=issue,
            old_status=old_status,
            new_status='resolved',
            changed_by=request.user,
            note='Khách thuê xác nhận đã xử lý xong'
        )
        
        # Notify landlord
        try:
            landlord = issue.house.loc.chu
            if landlord:
                from django.urls import reverse
                Notification.objects.create(
                    user=landlord,
                    title="Khách thuê đã xác nhận sự cố",
                    message=f"{issue.renter.hoten} đã xác nhận sự cố '{issue.title}' tại {issue.house.ten} đã được xử lý xong.",
                    type='issue_resolved',
                    target_url=reverse('issue_detail', kwargs={'issue_id': issue.id})
                )
        except Exception:
            pass
        
        messages.success(request, "Cảm ơn bạn đã xác nhận. Sự cố đã được đóng.")
        
    elif action == 'reject':
        issue.status = 'rejected'
        rejection_note = request.POST.get('note', 'Khách thuê cho biết chưa xử lý xong')
        issue.save()
        
        # Track history
        IssueStatusHistory.objects.create(
            issue=issue,
            old_status=old_status,
            new_status='rejected',
            changed_by=request.user,
            note=rejection_note
        )
        
        # Notify landlord
        try:
            landlord = issue.house.loc.chu
            if landlord:
                from django.urls import reverse
                Notification.objects.create(
                    user=landlord,
                    title="Sự cố cần xử lý lại",
                    message=f"{issue.renter.hoten} cho biết sự cố '{issue.title}' tại {issue.house.ten} chưa được xử lý xong: {rejection_note}",
                    type='issue_reported',
                    target_url=reverse('issue_detail', kwargs={'issue_id': issue.id})
                )
                
                # Send email
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    if getattr(settings, 'SEND_OWNER_EMAIL_NOTIFICATIONS', True):
                        send_mail(
                        subject=f'Sự cố cần xử lý lại: {issue.title}',
                        message=f'Khách thuê {issue.renter.hoten} cho biết sự cố tại {issue.house.ten} chưa được xử lý xong.\n\nLý do: {rejection_note}\n\n{issue.description}',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[landlord.email],
                        fail_silently=True,
                    )
                except Exception:
                    pass
        except Exception:
            pass
        
        messages.info(request, "Đã gửi yêu cầu xử lý lại cho chủ nhà.")
    
    # If HTMX request (from modal), return success response
    if request.headers.get('HX-Request'):
        return HttpResponse(status=200)
    
    # Otherwise return to notifications list
    return redirect('notifications_list')


@login_required
def bulk_change_issue_status(request):
    """Bulk change status for multiple issues"""
    if request.method == 'POST':
        issue_ids_str = request.POST.get('issue_ids', '')
        issue_ids = [int(i) for i in issue_ids_str.split(',') if i]
        new_status = request.POST.get('status')
        
        if issue_ids and new_status:
            issues = IssueReport.objects.filter(
                id__in=issue_ids,
                house__loc__chu=request.user
            )
            
            count = 0
            for issue in issues:
                old_status = issue.status
                if old_status != new_status:
                    issue.status = new_status
                    if new_status == 'resolved':
                        issue.resolved_at = timezone.now()
                    issue.save()
                    
                    # Track history
                    IssueStatusHistory.objects.create(
                        issue=issue,
                        old_status=old_status,
                        new_status=new_status,
                        changed_by=request.user,
                        note='Bulk update'
                    )
                    
                    # Notify renter
                    try:
                        if issue.renter and issue.renter.user:
                            status_text = dict(IssueReport.STATUS_CHOICES).get(new_status, new_status)
                            Notification.objects.create(
                                user=issue.renter.user,
                                title=f"Sự cố đã chuyển trạng thái: {status_text}",
                                message=f"Sự cố '{issue.title}' tại {issue.house.ten} đã chuyển sang: {status_text}.",
                                type='issue_resolved',
                                target_url=reverse('issue_detail', kwargs={'issue_id': issue.id})
                            )
                    except Exception:
                        pass
                    
                    count += 1
            
            if count > 0:
                status_text = dict(IssueReport.STATUS_CHOICES).get(new_status, new_status)
                messages.success(request, f"Đã cập nhật {count} sự cố sang trạng thái '{status_text}'.")
        
    return redirect('issue_list')

# ------------------------
# Authentication Modals
# ------------------------
from django.contrib.auth import login as auth_login
from django.conf import settings
from .forms import CustomAuthenticationForm
from django.views.decorators.http import require_http_methods

@require_http_methods(["GET", "POST"])
def login_modal(request):
    """Render and process a login form inside an HTML <dialog> for modal usage.

    GET: returns dialog with login form
    POST: authenticates; if success returns 204 with HX-Trigger to close and refresh/redirect
    """
    next_url = request.GET.get('next') or request.POST.get('next') or request.META.get('HTTP_REFERER') or '/'

    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            # Prefer HX-Redirect to the next_url to refresh the page
            response = HttpResponse(status=204)
            response["HX-Redirect"] = next_url
            # Also trigger closeModal for our dialog system if needed
            response["HX-Trigger"] = "closeModal"
            return response
        # On invalid, fall-through to render form with errors (status 400 for clarity)
    else:
        form = CustomAuthenticationForm(request)

    # Render modal dialog
    from django.shortcuts import render
    context = {
        'form': form,
        'next': next_url,
    }
    status = 400 if request.method == 'POST' else 200
    return render(request, 'sms/partials/login_modal.html', context=context, status=status)


@login_required
def renter_issue_list(request):
    """Renter's own issue list page - renters only"""
    # Check if user is a renter
    try:
        renter = Renter.objects.get(user=request.user)
    except Renter.DoesNotExist:
        messages.error(request, "Chỉ khách thuê mới có quyền truy cập trang này.")
        return redirect('index')
    
    issues = IssueReport.objects.filter(renter=renter).select_related('house', 'renter').order_by('-created_at')
    return render(request, 'sms/renter_issue_list.html', { 'issues': issues })


@login_required
def issue_detail(request, issue_id):
    """View issue detail - accessible by renter or landlord"""
    issue = get_object_or_404(IssueReport, id=issue_id)
    
    # Check access: either renter owns it or user is landlord
    is_renter = issue.renter and issue.renter.user == request.user
    is_landlord = issue.house.loc.chu == request.user
    
    if not (is_renter or is_landlord):
        return HttpResponse("Unauthorized", status=403)
    
    # Get related data
    images = issue.images.all()
    comments = issue.comments.filter(parent=None).prefetch_related('replies', 'user')
    status_history = issue.status_history.select_related('changed_by')
    comment_form = IssueCommentForm()
    
    context = {
        'issue': issue,
        'images': images,
        'comments': comments,
        'status_history': status_history,
        'comment_form': comment_form,
        'is_landlord': is_landlord,
        'is_renter': is_renter,
    }
    
    # If HTMX request, return modal content only
    if request.headers.get('HX-Request'):
        return render(request, 'sms/partials/issue_detail.html', context)
    
    # Otherwise return full page
    return render(request, 'sms/issue_detail_page.html', context)


@login_required
def issue_detail_modal(request, issue_id):
    """View issue detail for modal - always returns partial content"""
    issue = get_object_or_404(IssueReport, id=issue_id)
    
    # Check access: either renter owns it or user is landlord
    is_renter = issue.renter and issue.renter.user == request.user
    is_landlord = issue.house.loc.chu == request.user
    
    if not (is_renter or is_landlord):
        return HttpResponse("Unauthorized", status=403)
    
    # Get related data
    images = issue.images.all()
    comments = issue.comments.filter(parent=None).prefetch_related('replies', 'user')
    status_history = issue.status_history.select_related('changed_by')
    comment_form = IssueCommentForm()
    
    context = {
        'issue': issue,
        'images': images,
        'comments': comments,
        'status_history': status_history,
        'comment_form': comment_form,
        'is_landlord': is_landlord,
        'is_renter': is_renter,
    }
    
    return render(request, 'sms/partials/issue_detail.html', context)


@login_required
def add_issue_comment(request, issue_id):
    """Add comment to an issue"""
    issue = get_object_or_404(IssueReport, id=issue_id)
    
    # Check access
    is_renter = issue.renter and issue.renter.user == request.user
    is_landlord = issue.house.loc.chu == request.user
    
    if not (is_renter or is_landlord):
        return HttpResponse("Unauthorized", status=403)
    
    if request.method == 'POST':
        form = IssueCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.issue = issue
            comment.user = request.user
            
            # Handle reply (parent comment)
            parent_id = request.POST.get('parent_id')
            if parent_id:
                comment.parent_id = parent_id
            
            comment.save()
            
            # Notify the other party
            try:
                recipient = issue.house.loc.chu if is_renter else issue.renter.user
                if recipient:
                    from django.urls import reverse
                    Notification.objects.create(
                        user=recipient,
                        title=f"Bình luận mới về sự cố: {issue.title}",
                        message=f"{request.user.username} đã bình luận: {comment.comment[:100]}...",
                        type='issue_reported',
                        target_url=reverse('issue_detail', args=[issue.id])
                    )
                    
                    # Send email
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        if getattr(settings, 'SEND_OWNER_EMAIL_NOTIFICATIONS', True):
                            send_mail(
                            subject=f'Bình luận mới: {issue.title}',
                            message=f'{request.user.username} đã bình luận:\n\n{comment.comment}',
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[recipient.email],
                            fail_silently=True,
                        )
                    except Exception:
                        pass
            except Exception:
                pass
            
            # HTMX: return comment HTML
            if request.headers.get('HX-Request'):
                return render(request, 'sms/partials/comment_item.html', {
                    'comment': comment,
                    'is_landlord': is_landlord
                })
    
    return redirect('issue_detail', issue_id=issue_id)


@login_required
def notification_badge(request):
    """HTMX endpoint to return notification badge HTML"""
    from .models import Notification
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    return render(request, 'sms/notification_badge.html', {'unread_count': unread_count})

@login_required
def bill_history_view(request, bill_id):
    bill = get_object_or_404(Hoadon, pk=bill_id)
    history_qs = bill.history.order_by('-history_date')
    history_list = []
    for h in history_qs:
        # Tính toán sự khác biệt giữa bản ghi này và bản ghi trước đó
        diff = {}
        prev = h.prev_record
        if prev:
            for field in bill._meta.fields:
                fname = field.name
                old = getattr(prev, fname, None)
                new = getattr(h, fname, None)
                if old != new:
                    diff[fname] = {'old': old, 'new': new}
        history_list.append({'history_date': h.history_date,
                            'history_user': h.history_user,
                            'diff': diff})
    context = {
        'bill': bill,
        'history_list': history_list,
    }
    return render(request, 'sms/bill_history.html', context)

def renter_account_pdf(request, renter_id):
    from utils.generate_renter_pdf import generate_renter_account_pdf
    from django.http import HttpResponse
    from sms.models import Renter
    renter = get_object_or_404(Renter, id=renter_id)
    user = renter.user
    username = user.username if user else 'N/A'
    # Lấy password thực tế (ví dụ: lưu tạm khi tạo user, hoặc lấy từ trường tạm)
    password = getattr(renter, 'plain_password', 'Vui lòng liên hệ quản lý để nhận mật khẩu.')
    full_name = renter.hoten
    system_url = 'https://your-domain.com'  # Thay bằng địa chỉ hệ thống thực tế
    pdf_buffer = generate_renter_account_pdf(username, renter.init_pwd, full_name, system_url)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="renter_{username}_account.pdf"'
    return response

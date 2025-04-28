from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect
from .models import Lop, Ctdt, Hssv, Hsgv, Hsns, SvStatus, HocphiStatus, LopMonhoc, Trungtam, LogDiem
from .models import NsLop, LopHk
from django.urls import reverse
from django.utils import timezone
from django.contrib.auth.decorators import login_required, permission_required

from django.contrib.auth import get_user_model
from django.db.models import Sum

import psycopg2
import openpyxl, xlsxwriter 
import shutil, os
import locale

from django.conf import settings
from django.http import HttpResponse, Http404
from guardian.decorators import permission_required_or_403
from notifications.signals import notify



User = get_user_model()

from django.shortcuts import render, get_object_or_404, redirect
from .models import Ctdt, Diemthanhphan, Hocky, HocphiStatus, Loaidiem, TeacherInfo, Hsgv, Hssv, CtdtMonhoc, Monhoc, Lop, Lichhoc, Hs81, Diemdanh, Diemthanhphan, Hocphi, LopMonhoc, DiemdanhAll
from .models import LopHk, Hp81, Ttgv, UploadedFile
from .forms import CreateDiem, CreateLichhoc, CreateLopMonhoc, CreateTeacher, CreateCtdtMonhoc, CreateDiemdanh, CreateHocphi, CreateCtdt, CreateLop, CreateSv, CreateGv
from .forms import CreateHp81, CreateTtgv,CreateUploadFile
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime, date

import pandas as pd
from django.http import HttpResponseForbidden,HttpResponse

def must_be_supervisor(user):
    return user.groups.filter(name='Supervisors').count()


@login_required
@permission_required('dashboard.view_report',raise_exception=True)
def report_hs81(request, opt = None):
    svs = None
    #lops = Lop.objects.all()
    if request.user.is_superuser:
        lops = Lop.objects.all()
    elif request.user.is_internalstaff:
        ns = Hsns.objects.get(user = request.user)
        nsl = NsLop.objects.filter(ns = ns, status =1)
        lops = Lop.objects.filter(id__in = nsl.values_list('lop_id', flat=True))
    else:
        lops = None
    hks = Hocky.objects.all()
    lop  = None
    hk  = None
    if request.method == "POST":
        lop_id = request.POST.get('lop', None)
        hk_id = request.POST.get('hk', None)
#        svs = Hssv.objects.filter(lop_id=lop_id)
        locale.setlocale(locale.LC_ALL, 'vi_VN')
        #ans = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda ans: locale.strxfrm(ans.ten), reverse=False)
        svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
#        svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: svs.ten, reverse=False)
        hk = Hocky.objects.get(id = hk_id)
        lop = Lop.objects.get(id = lop_id)

        for sv in svs:
            sv.hs = Hs81.objects.filter(sv = sv, hk = hk)[0] if Hs81.objects.filter(sv = sv, hk = hk).exists() else None
            sv.hp = Hp81.objects.filter(sv = sv, hk = hk).select_related('status')[0] if Hp81.objects.filter(sv = sv, hk = hk).exists() else None

        messages.success(request, "Tìm kiếm thành công!")

        #export to excel
        if opt == 2:
            #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)
            exp=[]
            exp.append({"Lớp": lop.ten, 
                        "Học kỳ": hk.ten
                })
            for sv in svs:
                exp.append({"Mã": sv.msv,
                            "Họ tên": sv.hoten, 
                            "Hồ sơ":sv.hs.status if sv.hs else "", 
                            "Học phí":sv.hp.status.ten if sv.hp else "", 
                            "$ Dự kiến":sv.hp.sotien1 if sv.hp else "",
                            "$ Thực nhận":sv.hp.sotien2 if sv.hp else "",
                            })

            # Convert the QuerySet to a DataFrame
            df = pd.DataFrame(list(exp))

            # Define the Excel file response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report81.xlsx'

            # Use Pandas to write the DataFrame to an Excel file
            df.to_excel(response, index=False, engine='openpyxl')

            return response


    context = {
        "svs": svs,
        "lops": lops,
        "lop": lop,
        "hks": hks,
        "hk": hk
    }
    return render(request, "sms/report_hs81.html", context)

@login_required
@permission_required('dashboard.view_report',raise_exception=True)
def export_hs81(request):
    import pandas as pd

    mylist = []

    lh = Lop.objects.all().order_by("ten").select_related("trungtam")
  

    #svs = Hssv.objects.filter(malop__in=lh).select_related("malop").order_by("malop", "msv")
    for l in lh:
        #gvs = l.values()
        data={"Mã":"","Tên":l.ten, "Học kỳ":"","Hồ sơ":"","Học phí":"","Số tiền được giải ngân":"","Số tiền trường nhận":""}
        mylist.append(data)

        svs = Hssv.objects.filter(lop_id=l.id).order_by("msv")
        for sv in svs:
            data={"Mã":sv.msv,"Tên":sv.hoten, "Học kỳ":"","Hồ sơ":"","Học phí":"","Số tiền được giải ngân":"","Số tiền trường nhận":""}
            mylist.append(data)
            hp81s = Hp81.objects.filter(sv=sv).order_by("hk")
            for hp in hp81s:
                data={"Mã":"","Tên":"", "Học kỳ":hp.hk,"Hồ sơ": hp.hs81_st,"Học phí": hp.status,"Số tiền được giải ngân":hp.sotien1,"Số tiền trường nhận":hp.sotien2}
                mylist.append(data)


    df = pd.DataFrame(mylist)

    # Define the Excel file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export_hs81-hp.xlsx'

    # Use Pandas to write the DataFrame to an Excel file
    df.to_excel(response, index=True, engine='openpyxl')

    return response

@login_required
@permission_required('dashboard.view_report',raise_exception=True)
def report_ttgv(request, opt=None):
    if request.user.is_superuser:
        lh = Lop.objects.all()
    elif request.user.is_internalstaff:
        ns = Hsns.objects.get(user = request.user)
        nsl = NsLop.objects.filter(ns = ns, status =1)
        lh = Lop.objects.filter(id__in = nsl.values_list('lop_id', flat=True))
    else:
        lh = None
#    lh = Lop.objects.all()
    #query_tt = None
    lop= None
    lmhs = None
    if request.method == "POST":
        lop_id = request.POST.get('lop', None)
        #svs = Hssv.objects.filter(lop_id=lop_id)
        lop = Lop.objects.get(id=lop_id)

        #lmh = LopMonhoc.objects.get(id = lmh_id)
        lmhs = LopMonhoc.objects.filter(lop = lop).select_related("monhoc", "lop")

        
        for lmh in lmhs:
            lhs = Lichhoc.objects.filter(lmh = lmh).select_related("lop", "monhoc")
            gvs = Hsgv.objects.filter(id__in = lhs.values_list('giaovien_id', flat=True))
            lmh.sotiet = lhs.aggregate(Sum('sotiet'))['sotiet__sum']
            for gv in gvs:

                lhs = Lichhoc.objects.filter(lmh = lmh, giaovien = gv)
                gv.sotiet = lhs.aggregate(Sum('sotiet'))['sotiet__sum']
                if Ttgv.objects.filter(lopmh = lmh, gv = gv).exists():
                    gv.sotien1 = Ttgv.objects.get(lopmh = lmh, gv = gv).sotien1
                    gv.sotien2 = Ttgv.objects.get(lopmh = lmh, gv = gv).sotien2
                else:
                    gv.sotien1 = 0
                    gv.sotien2 = 0
            lmh.ttgvs=gvs


        messages.success(request, "Tìm kiếm thành công!")



        #export to excel
        if opt == 2:
            #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)
            exp=[]
            exp.append({"Lớp": lop.ten})
            for lmh in lmhs:
                exp.append({"Môn học": lmh.monhoc.ten,"Số tín chỉ": lmh.monhoc.sotinchi,"Số tiết": lmh.sotiet })
                for tt in lmh.ttgvs:
                    exp.append({"Giáo viên": tt.ma + "-" + tt.hoten,
                                "Số tiết": tt.sotiet, 
                                "$ Dự kiến": tt.sotien1, 
                                "$ Thực trả": tt.sotien2 
                                })

            # Convert the QuerySet to a DataFrame
            df = pd.DataFrame(list(exp))

            # Define the Excel file response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report-ttgv.xlsx'

            # Use Pandas to write the DataFrame to an Excel file
            df.to_excel(response, index=False, engine='openpyxl')

            return response

    context = {
        "lmhs": lmhs,
        "lh": lh,
        "lop": lop
    }
    return render(request, "sms/report_ttgv.html", context)

@login_required
@permission_required('dashboard.view_report',raise_exception=True)
def report_dd(request, opt=None):
    if request.user.is_superuser:
        lh = Lop.objects.all()
    elif request.user.is_internalstaff:
        ns = Hsns.objects.get(user = request.user)
        nsl = NsLop.objects.filter(ns = ns, status =1)
        lh = Lop.objects.filter(id__in = nsl.values_list('lop_id', flat=True))
    else:
        lh = None
#    lh = Lop.objects.all()
    #query_tt = None
    lop= None
    lmhs = None
    if request.method == "POST":
        lop_id = request.POST.get('lop', None)
        #svs = Hssv.objects.filter(lop_id=lop_id)
        lop = Lop.objects.get(id=lop_id)

        #lmh = LopMonhoc.objects.get(id = lmh_id)
        lmhs = LopMonhoc.objects.filter(lop = lop).select_related("monhoc", "lop")

        
        for lmh in lmhs:
            lhs = Lichhoc.objects.filter(lmh = lmh)
            for lich in lhs:

                dds = Diemdanh.objects.filter(lichhoc = lich, status = 0)
                lich.vangmat = dds
            lmh.lhs=lhs


        messages.success(request, "Tìm kiếm thành công!")



        #export to excel
        if opt == 2:
            #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)
            exp=[]
            exp.append({"Lớp": lop.ten})
            for lmh in lmhs:
                exp.append({"Môn học": lmh.monhoc.ten})
                for lich in lmh.lhs:
                    if lich.vangmat:
                        exp.append({"Lịch học": lich.thoigian.strftime("%d/%m/%Y %H:%M"),
                                    })
                        for v in lich.vangmat:
                            exp.append({"Học viên vắng mặt": v.sv.msv + "-" + v.sv.hoten,
                                        })

            # Convert the QuerySet to a DataFrame
            df = pd.DataFrame(list(exp))

            # Define the Excel file response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report-dd.xlsx'

            # Use Pandas to write the DataFrame to an Excel file
            df.to_excel(response, index=False, engine='openpyxl')

            return response

    context = {
        "lmhs": lmhs,
        "lh": lh,
        "lop": lop
    }
    return render(request, "sms/report_dd.html", context)

@login_required
@permission_required('dashboard.view_report',raise_exception=True)
def report_kqht(request, opt = None):
    if request.user.is_superuser:
        lh = Lop.objects.all()
    elif request.user.is_internalstaff:
        ns = Hsns.objects.get(user = request.user)
        nsl = NsLop.objects.filter(ns = ns, status =1)
        lh = Lop.objects.filter(id__in = nsl.values_list('lop_id', flat=True))
    else:
        lh = None
    #query_tt = None
    svs = None
    lop= None
    if request.method == "POST":
        lop_id = request.POST.get('lop', None)
        locale.setlocale(locale.LC_ALL, 'vi_VN')
        svs = sorted(Hssv.objects.filter(lop_id = lop_id), key=lambda svs: locale.strxfrm(svs.ten), reverse=False)
        lop = Lop.objects.get(id=lop_id)
        lhk = LopHk.objects.filter(lop = lop).select_related('hk').order_by('hk_id')
        next_hk=0

        for l in lhk:
            if l.end_hk and date.today() > l.end_hk:
                next_hk = l.hk.ma
                #break
        if next_hk == 0:
            next_hk = lhk.last().hk.ma + 1
        else:
            next_hk = next_hk + 1
        print(next_hk)
        lds = Loaidiem.objects.all()
        hks = Hocky.objects.filter(ma__lt = next_hk).order_by('ma')
        mhl=[]
        ldl=[]
        dtpl=[]



        # for mh in lmhs:
        #     ldl=[]
        #     for l in ld:
        #         dtpl=[]
        #         for dtp in Diemthanhphan.objects.filter(sv_id = sv_id, monhoc_id = mh.monhoc_id, tp_id = l.id, status=1).order_by('log_id'):
        #             dtpl.append({"id":dtp.log_id,"mark":dtp.diem})

        #         ldl.append({"ma":l.ma,"dtplst": dtpl})

        #     mhl.append({ "ma":mh.monhoc.ten,"ttdiem0": ldl[0], "ttdiem": ldl})

        for sv in svs:
            hkl=[]
            tctl, diem4 = 0,0
            xl = ""
            for hk in hks:
                lml=[]
                tbmhk, tchk = 0,0
                lmhs = LopMonhoc.objects.filter(lop_id = sv.lop_id, hk_id = hk.id).select_related("monhoc").order_by('ngaystart')
                print(hk.ten)
                for mh in lmhs:
                    print(mh.monhoc.ten)
                    ldl=[]
                    tbm1_diem, tbm1_heso, tbm2_diem, tbm2_heso, tbm= 0,0,0,0,0
                    kttx1,kttx2,kttx3,ktdk1,ktdk2,ktdk3,ktkt1,ktkt2 = 0,0,0,0,0,0,0,0
                    for ld in lds:
                        dtpl=[]
                        dtps = Diemthanhphan.objects.filter(sv = sv, monhoc_id = mh.monhoc_id, tp_id = ld.id, status=1).order_by('log_id')

                        i=1
                        for dtp in dtps:
                            if i==1 and ld.ma == 'KTTX':
                                kttx1 = dtp.diem
                            elif i==2 and ld.ma == 'KTTX':
                                kttx2 = dtp.diem
                            elif i==3 and ld.ma == 'KTTX':
                                kttx3 = dtp.diem
                            elif i==1 and ld.ma == 'KTĐK':
                                ktdk1 = dtp.diem
                            elif i==2 and ld.ma == 'KTĐK':
                                kttdk2 = dtp.diem
                            elif i==3 and ld.ma == 'KTĐK':
                                ktdk3 = dtp.diem
                            elif i==1 and ld.ma == 'KTKT':
                                ktkt1 = dtp.diem
                            elif i==2 and ld.ma == 'KTKT':
                                ktkt2 = dtp.diem
                            i=i+1
                            dtpl.append({"id":dtp.log_id,"mark":dtp.diem})
                            if ld.ma == 'KTĐK' or ld.ma == 'KTTX':
                                tbm1_diem = tbm1_diem + dtp.diem * ld.heso
                                tbm1_heso = tbm1_heso + ld.heso
                            elif ld.ma == 'KTKT' and dtp.diem > 0:
                                tbm2_diem = dtp.diem * ld.heso
                                # tbm2_heso = ld.heso
                                # print(tbm2_diem)
                                # print(tbm2_heso)
                        if ld.ma == 'KTKT':
                            tbm2_heso = ld.heso
                        ldl.append({"ma":ld.ma, "dtplst": dtpl})
                        
                    tbm = round(((tbm1_diem/tbm1_heso)*(10-tbm2_heso) + tbm2_diem)/10,1) if tbm1_heso else 0
                    tbmhk= tbmhk+tbm*mh.monhoc.sotinchi
                    tchk=tchk+mh.monhoc.sotinchi
                    lml.append({ "ten":mh.monhoc.ten,
                                "tc": mh.monhoc.sotinchi,
                                "tbm": tbm,
                                "kttx1": kttx1, 
                                "kttx2": kttx2, 
                                "kttx3": kttx3,
                                "ktdk1": ktdk1,
                                "ktdk2": ktdk2, 
                                "ktdk3": ktdk3,
                                "ktkt1": ktkt1,
                                "ktkt2": ktkt2
                                })
                    
                tbmhk = round(tbmhk/tchk,1)
                if tbmhk >=8.5 and tbmhk <=10:
                    tbmhk4 = 4
                elif tbmhk >=7 and tbmhk <=8.4:
                    tbmhk4 = 3
                elif tbmhk >=5.5 and tbmhk <=6.9:
                    tbmhk4 = 2
                elif tbmhk >=4 and tbmhk <=5.4:
                    tbmhk4 = 1
                elif tbmhk  < 4:
                    tbmhk4 = 0

                tctl = tctl + tchk
                diem4 = diem4 +tbmhk4*tchk

                tbctl = round(diem4/tctl,1)

                if tbctl >=3.5 and tbctl <=4:
                    xl = "Xuất xắc"
                elif tbctl >=3 and tbctl <3.5:
                    xl = "Giỏi"
                elif tbctl >=2.5 and tbctl <3:
                    xl = "Khá"
                elif tbctl >=2 and tbctl <2.5:
                    xl = "Trung bình"
                elif tbctl <2:
                    xl = "Yếu"

                hkl.append({"ma":hk.ma, "tbmhk":tbmhk, "tbmhk4":tbmhk4,"tbctl":tbctl, "xl":xl})

                # hk.tchk = tchk
                # hk.tbmhk = round(tbmhk/tchk,1)
            sv.hkl = hkl
        messages.success(request, "Tạo báo cáo theo tiêu chí thành công!")

        #export to excel
        if opt ==1:
            #svs = sorted(svs, key=lambda svs: svs.ten, reverse=False)
            exp=[]
            exp.append({"Lớp": lop.ten
                        })
            for sv in svs:
                tbmhk1,tbmhk41,tbctl1,xl1 = None,None,None,None
                tbmhk2,tbmhk42,tbctl2,xl2 = None,None,None,None
                tbmhk3,tbmhk43,tbctl3,xl3 = None,None,None,None
                tbmhk4,tbmhk44,tbctl4,xl4 = None,None,None,None   
                for hk in sv.hkl:
                    print('printing hk')
                    if hk['ma'] == 1:
                        tbmhk1 = hk['tbmhk']
                        tbmhk41 = hk['tbmhk4']
                        tbctl1 = hk['tbctl']
                        xl1 = hk['xl']
                    elif hk['ma'] == 2:
                        tbmhk2 = hk['tbmhk']
                        tbmhk42 = hk['tbmhk4']
                        tbctl2 = hk['tbctl']
                        xl2 = hk['xl']
                    elif hk['ma'] == 3:
                        tbmhk3 = hk['tbmhk']
                        tbmhk43 = hk['tbmhk4']
                        tbctl3 = hk['tbctl']
                        xl3 = hk['xl']
                    elif hk['ma'] == 4:
                        tbmhk4 = hk['tbmhk']
                        tbmhk44 = hk['tbmhk4']
                        tbctl4 = hk['tbctl']
                        xl4 = hk['xl']

                    
                exp.append({"Mã học tên": sv.msv,
                            "Họ tên": sv.hoten, 
                            "HK1 TBM 10":tbmhk1 , 
                            "HK1 TBM 4":tbmhk41 , 
                            "HK1 TBCTL":tbctl1 , 
                            "HK1 XL":xl1 , 
                            "HK2 TBM 10":tbmhk2, 
                            "HK2 TBM 4":tbmhk42 , 
                            "HK2 TBCTL":tbctl2 , 
                            "HK2 XL":xl2 , 
                            "HK3 TBM 10":tbmhk3, 
                            "HK3 TBM 4":tbmhk43 , 
                            "HK3 TBCTL":tbctl3 , 
                            "HK3 XL":xl3 , 
                            "HK4 TBM 10":tbmhk4, 
                            "HK4 TBM 4":tbmhk44 , 
                            "HK4 TBCTL":tbctl4 , 
                            "HK4 XL":xl4 , 
                            })

            # Convert the QuerySet to a DataFrame
            df = pd.DataFrame(list(exp))

            # Define the Excel file response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=kqht-lop.xlsx'

            # Use Pandas to write the DataFrame to an Excel file
            df.to_excel(response, index=False, engine='openpyxl')

            return response

            
    context = {
        "lh": lh,
        "lop": lop,
        "svs": svs
        # "query_tt": query_tt,
        # "query_lop": query_lop
    }
    return render(request, "sms/report_kqht.html", context)

@login_required
@permission_required('sms.add_hs81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def import_hs81(request, lop_id):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        if 'hs81' not in wb.sheetnames:
            messages.error(request, "File excel khong dung format")
            return redirect("hv_hs81_new_list", lop_id)

        sheet = wb["hs81"]
        #for r in range(3, sheet.max_row+1):
        #maxid=500
        #for r in range(3, 44):
        for r in range(2, sheet.max_row+1):
            #maxid = maxid+1
            v1 = sheet.cell(r,1).value
            v2 = sheet.cell(r,2).value
            v3 = sheet.cell(r,3).value
            v4 = sheet.cell(r,4).value
            v5 = sheet.cell(r,5).value

            v6 = sheet.cell(r,6).value
            v7 = sheet.cell(r,7).value
            v8 = sheet.cell(r,8).value
            v9 = sheet.cell(r,9).value

            v10 = sheet.cell(r,10).value
            v11 = sheet.cell(r,11).value
            v12 = sheet.cell(r,12).value
            v13 = sheet.cell(r,13).value
            v14 = sheet.cell(r,14).value
            #print(type(v4))
            if not v1 or not v2 or not v3:
                messages.error(request, 'Mã: ' + v1+ ' thiếu thông tin bắt buộc')
                continue
            if not Hssv.objects.filter(msv=v1,lop_id = lop_id).exists():
                messages.error(request, 'Mã: ' + v1+ ' không có trong danh sách lóp')
                continue
            else:
                sv=Hssv.objects.get(msv=v1)
                if v4 == 1:
                    status = "Đủ"
                else:
                    status = "Thiếu"
                if type(v5) is not datetime:
                    v5 = None
                if Hs81.objects.filter(sv=sv, hk_id = v3).exists():
                    hs81 = Hs81.objects.filter(sv=sv, hk_id = v3)[0]
                    hs81.status = status
                    hs81.thoigian = v5
                else:
                    hs81 = Hs81(sv=sv, hk_id = v3, status=status, thoigian=v5)
                
                hs81.ddn = True if v6 == 1 else False
                hs81.cntn = True if v7 == 1 else False
                hs81.btn = True if v8 == 1 else False
                hs81.xnct = True if v9 == 1 else False
                hs81.cccd = True if v10 == 1 else False
                hs81.cccdbo = True if v11 == 1 else False
                hs81.cccdme = True if v12 == 1 else False
                hs81.gks = True if v13 == 1 else False
                hs81.ghichu = v14
                hs81.save()

        messages.success(request, "Import thông tin hồ sơ 81 thành công!")
        return redirect("hv_hs81_new_list", lop_id)

@login_required
@permission_required('sms.add_hp81',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def import_hp81(request, lop_id):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        if 'hp81' not in wb.sheetnames:
            messages.error(request, "File excel khong dung format")
            return redirect("hv_hp81_new_list", lop_id)

        sheet = wb["hp81"]
        #for r in range(3, sheet.max_row+1):
        #maxid=500
        #for r in range(3, 44):
        for r in range(2, sheet.max_row+1):
            #maxid = maxid+1
            v1 = sheet.cell(r,1).value
            v2 = sheet.cell(r,2).value
            v3 = sheet.cell(r,3).value
            v4 = sheet.cell(r,4).value
            v5 = sheet.cell(r,5).value
            v6 = sheet.cell(r,6).value
            v7 = sheet.cell(r,7).value
            #print(type(v4))
            if not v1 or not v2 or not v3:
                messages.error(request, 'Mã: ' + v1+ ' thiếu thông tin bắt buộc')
                continue
            if not Hssv.objects.filter(msv=v1, lop_id = lop_id).exists():
                messages.error(request, 'Mã: ' + v1+ ' không có trong danh sách lóp')
                continue
            # elif v5 and type(v5) is not datetime:
            #     messages.error(request, 'Mã: ' + v1 + ' Ngày thu: ' + v5 + ' sai định dạng ngày (m/d/yy)')
            #     continue
            else:
                sv=Hssv.objects.get(msv=v1, lop_id = lop_id)
                if not v4:
                    status = 1
                else:
                    status = v4
                if Hp81.objects.filter(sv=sv, hk_id = v3).exists():
                    hp81 = Hp81.objects.filter(sv=sv, hk_id = v3)[0]
                else:
                    hp81 = Hp81(sv=sv, hk_id = v3)
                
                hp81.status_id = status
                hp81.thoigian = v5
                hp81.sotien1 = "{:,}".format(v6) if v6 else ""
                hp81.sotien2 = "{:,}".format(v7) if v7 else ""
                chu = "{:,}".format(v6)
                print(chu)
                print(type(chu))
                try:
                    hp81.save()
                except Exception as e:
                    messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
                #hp81.save()
                print(sv.hoten)

        messages.success(request, "Import thông tin học phí 81 thành công!")
        return redirect("hv_hp81_new_list", lop_id)
    
@login_required
@permission_required('sms.add_diemthanhphan',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def import_diemtp(request, lop_id, lmh_id, ld_id):
    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("monhoc", "lop")[0]
    ld = Loaidiem.objects.get(id = ld_id)
    stud_list = Hssv.objects.filter(lop_id = lmh.lop_id)
    #teachers = Hsgv.objects.all().order_by('hoten')
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        if 'diemtp' not in wb.sheetnames:
            messages.error(request, "File excel khong dung format")
            return redirect("diemtp-lmh-lst", lmh_id)

        sheet = wb["diemtp"]
        #for r in range(3, sheet.max_row+1):
        #maxid=500
        #for r in range(3, 44):
        log = LogDiem()
        log.ten = request.user.username
        log.save()
        for sv in stud_list:
            dtp = Diemthanhphan(sv_id = sv.id, tp_id = ld_id, monhoc_id=lmh.monhoc.id, status=1, diem = 0, log=log) 
            dtp.save()

        for r in range(2, sheet.max_row+1):
            #maxid = maxid+1
            try:
                v1 = sheet.cell(r,1).value
                v2 = sheet.cell(r,2).value
                v3 = sheet.cell(r,3).value
                if not v1 or not v2 or not v3:
                    messages.error(request, 'Dòng: '+str(r)+' thiếu thông tin bắt buộc')
                    continue
                if not Hssv.objects.filter(msv=v1, hoten=v2, lop_id = lmh.lop.id).exists():
                    messages.error(request, 'Dòng: '+str(r)+' không có trong danh sách lóp')
                    continue
                if not (v3 >=0 and v3<=10):
                    messages.error(request, 'Dòng: '+str(r)+' có điểm sai định dạng')
                    continue
                sv = Hssv.objects.filter(msv=v1, hoten=v2, lop_id = lmh.lop.id)[0]

                mark = Diemthanhphan.objects.filter(sv_id = sv.id, tp_id = ld_id, monhoc_id=lmh.monhoc.id, status=1, log=log)[0]
                mark.diem = v3
                mark.save()
    
                #send notification to Hv
                if sv.user:
                    notify.send(sender=sv.user, recipient= sv.user, verb='Thông tin điểm được cập nhật trên hệ thống', level='info')
            except Exception as e:
                messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
            #mark.save()

        messages.success(request, "Import thông tin điểm thành công!")
        return redirect("diemtp-lmh-lst", lmh_id)

    context = {
        "lmh": lmh,
        "ld": ld
    }
    return render(request, "sms/import_diemtp.html", context)

@login_required
@permission_required('sms.change_diemthanhphan',raise_exception=True)
@permission_required_or_403('sms.assign_lop',(Lop, 'id', 'lop_id'))
def import_edit_diemtp(request, lop_id, lmh_id, ld_id, log_id):
    lmh = LopMonhoc.objects.filter(id = lmh_id).select_related("monhoc", "lop")[0]
    ld = Loaidiem.objects.get(id = ld_id)
    stud_list = Hssv.objects.filter(lop_id = lmh.lop_id)
    #teachers = Hsgv.objects.all().order_by('hoten')
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        if 'diemtp' not in wb.sheetnames:
            messages.error(request, "File excel khong dung format")
            return redirect("diemtp-lmh-lst", lmh_id)

        sheet = wb["diemtp"]
        #for r in range(3, sheet.max_row+1):
        #maxid=500
        #for r in range(3, 44):
        log = LogDiem.objects.get(id = log_id)
        log.ten = request.user.username
        log.capnhat_at = datetime.now()
        log.save()
        #log.save()
        for r in range(2, sheet.max_row+1):
            #maxid = maxid+1
            try:
                v1 = sheet.cell(r,1).value
                v2 = sheet.cell(r,2).value
                v3 = sheet.cell(r,3).value
                #print(type(v4))
                if not v1 or not v2 or not v3:
                    messages.error(request, 'Mã: ' + v1+ ' thiếu thông tin bắt buộc')
                    continue
                if not Hssv.objects.filter(msv=v1, hoten=v2, lop_id = lmh.lop_id).exists():
                    messages.error(request, 'Mã: ' + v1+ ' không có trong danh sách lóp')
                    continue
                if not (v3 >=0 and v3<=10):
                    messages.error(request, 'Mã: ' + v1+ ' không có điểm')
                    continue
                sv = Hssv.objects.filter(msv=v1, hoten=v2, lop_id = lmh.lop_id)[0]
                if Diemthanhphan.objects.filter(sv_id = sv.id, tp_id = ld_id, monhoc_id=lmh.monhoc_id, status=1, log=log).exists():
                    mark = Diemthanhphan.objects.filter(sv_id = sv.id, tp_id = ld_id, monhoc_id=lmh.monhoc_id, log=log)[0]
                #create new
                else:
                    mark = Diemthanhphan(sv_id = sv.id, tp_id = ld_id, monhoc_id=lmh.monhoc.id, status=1, diem = 0, log=log) 

                mark.diem = v3
                mark.save()
                #send notification to Hv
                if sv.user:
                    notify.send(sender=sv.user, recipient= sv.user, verb='Thông tin điểm được cập nhật trên hệ thống', level='info')
                
            except Exception as e:
                messages.error(request, 'Dòng: '+str(r)+' có lỗi:' + str(e))
#                mark.save()

        messages.success(request, "Import thông tin điểm thành công!")
        return redirect("diemtp-lmh-lst", lmh_id)

    context = {
        "lmh": lmh,
        "log_id": log_id,
        "ld": ld
    }
    return render(request, "sms/import_edit_diemtp.html", context)

@login_required
def tb_list(request):
    #students = Hssv.objects.all()
    tbaos = request.user.notifications.all()[0:50]
    paginator = Paginator(tbaos, 60)
    page = request.GET.get('page')
    tbs = paginator.get_page(page)

    context = {
        "tbs": tbs
    }
    return render(request, "sms/tb_list.html", context)
